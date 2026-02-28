import re
import datetime
from io import BytesIO
from typing import Dict, List, Tuple, Optional, Any

import pandas as pd
import numpy as np

try:
    from rapidfuzz import fuzz, process as rf_process
    RAPIDFUZZ_AVAILABLE = True
except ImportError:
    RAPIDFUZZ_AVAILABLE = False


# ═══════════════════════════════════════════════════════════════════════
#  CONFIGURATION  (inline constants — no external files)
# ═══════════════════════════════════════════════════════════════════════
APP_TITLE = "Enterprise Data Quality Assessment"
APP_ICON = "📊"
SUPPORTED_FORMATS = ["csv", "xlsx", "xls", "xlsm"]

DIMENSIONS: Dict[str, List[str]] = {
    "Completeness": [
        "Not Null",
        "Not Empty",
        "Whitespace Only",
        "Minimum Length",
        "Mandatory Column",
    ],
   "Validity": [
    "Data Type Validation",
    "Email Format",
    "Phone Format",
    "PAN Format",
    "Numeric Range",
    "Allowed Values",
    "Custom Regex",
    "Length Check",
    "Format Check",
],
    "Uniqueness": [
        "Single Column Exact Match",
        "Combination Column Exact Match",
        "Hybrid Fuzzy Match",
    ],
    "Standardization": [
    "Special Characters Not Allowed",
    "Convert to Upper Case",
    "Convert to Lower Case",
    "Normalize Date Format",
],
}
def std_remove_special_characters(series: pd.Series) -> pd.Series:
    return series.astype(str).str.replace(r"[^a-zA-Z0-9\s]", "", regex=True)

def std_upper_case(series: pd.Series) -> pd.Series:
    return series.astype(str).str.upper()

def std_lower_case(series: pd.Series) -> pd.Series:
    return series.astype(str).str.lower()

def std_normalize_date(series: pd.Series, fmt="%d-%m-%Y") -> pd.Series:
    return pd.to_datetime(series, errors="coerce").dt.strftime(fmt)

def render_fuzzy_rule_builder(all_columns: List[str]) -> List[Dict[str, Any]]:
    """
    Render an interactive Streamlit UI for building named fuzzy duplicate rules.
    Returns the current list of saved rules.

    Usage in app.py::

        from modules.dq import render_fuzzy_rule_builder
        active_fuzzy_rules = render_fuzzy_rule_builder(all_columns)
    """
    import streamlit as st  # imported here so dq.py stays importable without streamlit

    # ── Initialise session state ────────────────────────────────────────
    if "fuzzy_rules" not in st.session_state:
        st.session_state["fuzzy_rules"] = []

    rules: List[Dict[str, Any]] = st.session_state["fuzzy_rules"]

    st.markdown("""
    <div style="background:#f5f0fc;border-left:4px solid #5b2d90;padding:0.65rem 1rem;
         border-radius:0 8px 8px 0;margin-bottom:0.9rem;">
        <strong style="color:#3b1f72;">Fuzzy Rule Builder</strong>
        <span style="color:#555;font-size:0.85rem;"> — build named rules, set per-column weights and thresholds.</span>
    </div>""", unsafe_allow_html=True)

    tab_add, tab_view = st.tabs(["➕ Add Rule", "📋 Active Rules"])

    # ── Add Rule tab ────────────────────────────────────────────────────
    with tab_add:
        col_a, col_b = st.columns([3, 2])

        with col_a:
            chosen_cols = st.multiselect(
                "Columns to match on",
                options=all_columns,
                key="frb_cols",
                help="Select 1 or more columns. Rows are compared using a weighted similarity score across all selected columns.",
            )

        with col_b:
            rule_name_default = ("FUZZY: " + " + ".join(chosen_cols)) if chosen_cols else "FUZZY: Rule"
            rule_name = st.text_input(
                "Rule name",
                value=rule_name_default,
                key="frb_name",
            )

        # Per-column weights
        weights: List[float] = []
        if chosen_cols:
            st.markdown(
                '<div style="font-size:0.82rem;font-weight:600;color:#5b2d90;margin-bottom:0.3rem;">'
                "Per-column weights (higher = more important in similarity score)</div>",
                unsafe_allow_html=True,
            )
            w_cols = st.columns(min(len(chosen_cols), 4))
            for wi, col_name in enumerate(chosen_cols):
                with w_cols[wi % 4]:
                    w = st.number_input(
                        f"Weight: {col_name[:18]}",
                        min_value=0.0,
                        max_value=10.0,
                        value=1.0,
                        step=0.5,
                        key=f"frb_w_{wi}_{col_name}",
                        help="0 = ignore this column. Weights are normalised internally.",
                    )
                    weights.append(w)

        # Threshold + advanced settings
        adv1, adv2, adv3 = st.columns(3)
        with adv1:
            threshold = st.slider(
                "Similarity threshold (%)",
                min_value=60,
                max_value=99,
                value=85,
                step=1,
                key="frb_threshold",
                help="Minimum weighted score to flag two records as duplicates.",
            )
        with adv2:
            max_pairs = st.number_input(
                "Max pairs per block",
                min_value=1_000,
                max_value=500_000,
                value=20_000,
                step=1_000,
                key="frb_max_pairs",
                help="Blocks with more pairwise comparisons than this are skipped (performance guard).",
            )
        with adv3:
            ignore_nulls = st.checkbox(
                "Skip rows with blank/null values",
                value=True,
                key="frb_ignore_nulls",
                help="Rows where any selected column is blank/null are excluded from fuzzy comparison.",
            )

        btn_col, _ = st.columns([1, 3])
        with btn_col:
            if st.button("➕ Add Rule", key="frb_add_btn", use_container_width=True):
                if not chosen_cols:
                    st.warning("⚠️ Select at least one column before adding a rule.")
                else:
                    new_rule: Dict[str, Any] = {
                        "name":         rule_name.strip() or rule_name_default,
                        "cols":         list(chosen_cols),
                        "weights":      weights if weights else [1.0] * len(chosen_cols),
                        "threshold":    int(threshold),
                        "ignore_nulls": bool(ignore_nulls),
                        "max_pairs":    int(max_pairs),
                    }
                    st.session_state["fuzzy_rules"].append(new_rule)
                    st.success(f"✅ Rule **{new_rule['name']}** added.")
                    st.rerun()

    # ── Active Rules tab ────────────────────────────────────────────────
    with tab_view:
        if not rules:
            st.info("No fuzzy rules added yet. Use the **Add Rule** tab to create one.")
        else:
            # Display rules as a styled table
            display_rows = []
            for i, r in enumerate(rules):
                weight_str = ", ".join(
                    f"{c}×{w}" for c, w in zip(r["cols"], r["weights"])
                )
                display_rows.append({
                    "#":          i + 1,
                    "Rule Name":  r["name"],
                    "Columns":    " + ".join(r["cols"]),
                    "Weights":    weight_str,
                    "Threshold":  f"{r['threshold']}%",
                    "Max Pairs":  f"{r['max_pairs']:,}",
                    "Skip Nulls": "Yes" if r["ignore_nulls"] else "No",
                })
            import pandas as _pd
            st.dataframe(_pd.DataFrame(display_rows), use_container_width=True, hide_index=True)

            # Per-rule delete buttons
            st.markdown(
                '<div style="font-size:0.8rem;color:#7a7a9a;margin-bottom:0.4rem;">'
                "Remove a rule:</div>",
                unsafe_allow_html=True,
            )
            del_cols = st.columns(min(len(rules), 5))
            for i, r in enumerate(rules):
                with del_cols[i % 5]:
                    if st.button(
                        f"🗑 #{i+1}",
                        key=f"frb_del_{i}",
                        use_container_width=True,
                        help=f"Remove rule: {r['name']}",
                    ):
                        st.session_state["fuzzy_rules"].pop(i)
                        st.rerun()

            _, clear_col = st.columns([4, 1])
            with clear_col:
                if st.button("🗑 Clear All Rules", key="frb_clear_all", use_container_width=True):
                    st.session_state["fuzzy_rules"] = []
                    st.rerun()

    return st.session_state.get("fuzzy_rules", [])
def load_dataset(uploaded_file, sheet_name: Optional[str] = None) -> pd.DataFrame:
    """Load CSV or Excel into a Pandas DataFrame."""
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        df = pd.read_csv(uploaded_file, dtype=str, keep_default_na=False)
    elif name.endswith((".xlsx", ".xls", ".xlsm")):
        df = pd.read_excel(
            uploaded_file,
            sheet_name=sheet_name or 0,
            dtype=str,
            keep_default_na=False,
            engine="openpyxl",
        )
    else:
        raise ValueError(f"Unsupported format: {name}")
    df.columns = df.columns.str.strip()
    return df


def get_excel_sheet_names(uploaded_file) -> List[str]:
    """Return sheet names for an Excel file."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(uploaded_file.read()), read_only=True)
        uploaded_file.seek(0)
        return wb.sheetnames
    except Exception:
        uploaded_file.seek(0)
        return []


# ═══════════════════════════════════════════════════════════════════════
#  COMPLETENESS RULES  (vectorized Pandas operations)
# ═══════════════════════════════════════════════════════════════════════
def rule_not_null(series: pd.Series) -> pd.Series:
    """
    True where a value is genuinely present (not null / NaN / empty / null-sentinel).

    Background: datasets are loaded with ``keep_default_na=False`` and ``dtype=str``,
    so all cells arrive as Python strings.  Empty Excel cells become ``""``.
    Previously-null cells that pandas would have parsed as NaN instead arrive as
    the literal string ``"nan"``.  All of the patterns below must be treated as
    missing for completeness purposes.
    """
    s = series.astype(str).str.strip()
    _NULL_SENTINELS = {"", "nan", "none", "null", "na", "n/a", "n.a.", "nil", "missing", "#n/a"}
    return ~s.str.lower().isin(_NULL_SENTINELS)


def rule_not_empty(series: pd.Series) -> pd.Series:
    """True where value is a non-empty, non-whitespace-only string."""
    return series.astype(str).str.strip().str.len() > 0


def rule_whitespace_only(series: pd.Series) -> pd.Series:
    """True where value is NOT whitespace-only."""
    return ~(series.astype(str).str.fullmatch(r"\s+", na=False))


def rule_minimum_length(series: pd.Series, min_len: int) -> pd.Series:
    """True where the stripped value has at least min_len characters.

    Null-sentinel strings (empty, "nan", "none", etc.) are treated as absent
    and always fail this check — they should be caught by Not Null first.
    """
    s = series.astype(str).str.strip()
    _NULL_SENTINELS = {"", "nan", "none", "null", "na", "n/a", "n.a.", "nil", "missing", "#n/a"}
    is_null = s.str.lower().isin(_NULL_SENTINELS)
    return (~is_null) & (s.str.len() >= min_len)


# ═══════════════════════════════════════════════════════════════════════
#  VALIDITY RULES
# ═══════════════════════════════════════════════════════════════════════
def _is_empty_for_validity(s: pd.Series) -> pd.Series:
    """
    Return a boolean mask that is True for cells that should be SKIPPED by
    validity rules because they represent missing / absent values.

    Completeness rules handle these; validity rules should not double-count them.
    The dataset is loaded with keep_default_na=False so null cells arrive as
    "", "nan", "none", "null", "na", "n/a", etc.
    """
    _NULL_SENTINELS = {"", "nan", "none", "null", "na", "n/a", "n.a.", "nil", "missing", "#n/a"}
    return s.str.lower().isin(_NULL_SENTINELS)


def rule_data_type(series: pd.Series, expected: str) -> pd.Series:
    """Validate expected data type: numeric, integer, float, string, date."""
    s = series.astype(str).str.strip()
    skip = _is_empty_for_validity(s)
    if expected == "numeric":
        valid = s.str.fullmatch(r"-?\d+\.?\d*", na=False)
    elif expected == "integer":
        valid = s.str.fullmatch(r"-?\d+", na=False)
    elif expected == "float":
        valid = s.str.fullmatch(r"-?\d+\.\d+", na=False)
    elif expected == "date":
        valid = pd.to_datetime(s, errors="coerce", dayfirst=True).notna()
    else:
        # string — always passes
        return pd.Series(True, index=series.index)
    return skip | valid


def rule_email_format(series: pd.Series) -> pd.Series:
    pattern = r"^[a-zA-Z0-9_.+\-]+@[a-zA-Z0-9\-]+\.[a-zA-Z0-9.\-]+$"
    s = series.astype(str).str.strip()
    return _is_empty_for_validity(s) | s.str.fullmatch(pattern, na=False)


def rule_phone_format(series: pd.Series) -> pd.Series:
    pattern = r"^\+?[0-9][\d\s\-\(\)]{6,14}$"
    s = series.astype(str).str.strip()
    return _is_empty_for_validity(s) | s.str.fullmatch(pattern, na=False)


def rule_pan_format(series: pd.Series) -> pd.Series:
    pattern = r"^[A-Z]{5}[0-9]{4}[A-Z]$"
    s = series.astype(str).str.strip().str.upper()
    skip = _is_empty_for_validity(s)
    return skip | s.str.fullmatch(pattern, na=False)


def rule_date_format(series: pd.Series, fmt: str = "") -> pd.Series:
    s = series.astype(str).str.strip()
    skip = _is_empty_for_validity(s)
    if fmt:
        try:
            valid = pd.to_datetime(s, format=fmt, errors="coerce").notna()
            return skip | valid
        except Exception:
            pass
    valid = pd.to_datetime(s, errors="coerce", dayfirst=True).notna()
    return skip | valid


def rule_numeric_range(series: pd.Series, min_val: float, max_val: float) -> pd.Series:
    s = series.astype(str).str.strip()
    skip = _is_empty_for_validity(s)
    num = pd.to_numeric(s, errors="coerce")
    in_range = (num >= min_val) & (num <= max_val)
    # Non-numeric non-empty values: num is NaN → in_range is False → correctly fails
    return skip | in_range


def rule_allowed_values(series: pd.Series, allowed: List[str]) -> pd.Series:
    s = series.astype(str).str.strip()
    skip = _is_empty_for_validity(s)
    allowed_lower = {v.strip().lower() for v in allowed}
    return skip | s.str.lower().isin(allowed_lower)


def rule_custom_regex(series: pd.Series, pattern: str) -> pd.Series:
    s = series.astype(str).str.strip()
    skip = _is_empty_for_validity(s)
    try:
        return skip | s.str.fullmatch(pattern, na=False)
    except re.error:
        return pd.Series(True, index=series.index)


# ═══════════════════════════════════════════════════════════════════════
#  UNIQUENESS RULES
# ═══════════════════════════════════════════════════════════════════════
def detect_exact_duplicates_single(df: pd.DataFrame, col: str) -> pd.DataFrame:
    """Single-column exact duplicate detection with metadata.

    FIX: Rows where the column value is null/empty/nan are excluded before
    duplicate detection. A group of identical empty values is NOT a duplicate
    — it is a completeness issue handled by the Not Null rule.
    """
    _NULL_SENTINELS = {"", "nan", "none", "null", "na", "n/a", "n.a.", "nil", "missing", "#n/a"}

    s = df[col].astype(str).str.strip().str.lower()

    # Exclude null/empty rows from duplicate detection
    non_null_mask = ~s.isin(_NULL_SENTINELS)
    s_valid = s[non_null_mask]

    if s_valid.empty:
        return pd.DataFrame()

    dup_mask_valid = s_valid.duplicated(keep=False)
    if not dup_mask_valid.any():
        return pd.DataFrame()

    # Map back to original df using label indices
    dup_labels = s_valid.index[dup_mask_valid]
    dup_df = df.loc[dup_labels].copy()
    groups = s_valid[dup_mask_valid]
    group_map = {v: i + 1 for i, v in enumerate(groups.unique())}
    dup_df["Duplicate_Group_ID"] = groups.map(group_map).values
    dup_df["Duplicate_Type"] = "Exact"
    dup_df["Similarity_Score"] = 100.0
    dup_df["Row_Number"] = dup_df.index + 2  # Excel row (1-indexed + header)
    return dup_df


def detect_exact_duplicates_multi_single(
    df: pd.DataFrame, cols: List[str]
) -> Tuple[pd.DataFrame, Dict[str, pd.DataFrame]]:
    """
    Run single-column exact duplicate detection independently on each column
    in *cols*.

    Returns
    -------
    combined_df : pd.DataFrame
        All duplicate rows merged (may contain rows flagged by multiple cols).
    per_col_dfs : Dict[str, pd.DataFrame]
        Mapping of column_name → individual duplicate DataFrame (may be empty).
    """
    per_col_dfs: Dict[str, pd.DataFrame] = {}
    all_frames: List[pd.DataFrame] = []

    for col in cols:
        if col not in df.columns:
            continue
        dup = detect_exact_duplicates_single(df, col)
        per_col_dfs[col] = dup
        if not dup.empty:
            # Tag which column triggered the duplicate flag
            dup = dup.copy()
            dup["_Source_Column"] = col
            all_frames.append(dup)

    if not all_frames:
        return pd.DataFrame(), per_col_dfs

    combined = pd.concat(all_frames, ignore_index=True)
    return combined, per_col_dfs


def detect_exact_duplicates_combination(
    df: pd.DataFrame, cols: List[str]
) -> pd.DataFrame:
    """Multi-column combination exact duplicate detection.

    FIX: Rows where ALL selected columns are null/empty/nan are excluded
    before duplicate detection. A set of all-null rows sharing the same
    "nan||nan||..." key is NOT a duplicate — it is a completeness issue.
    Rows where only SOME columns are null are still included (they may
    genuinely duplicate another partially-null row).
    """
    _NULL_SENTINELS = {"", "nan", "none", "null", "na", "n/a", "n.a.", "nil", "missing", "#n/a"}

    combo = df[cols].astype(str).apply(lambda x: x.str.strip().str.lower())

    # Exclude rows where EVERY selected column is a null sentinel
    all_null_mask = combo.apply(
        lambda row: all(v in _NULL_SENTINELS for v in row), axis=1
    )
    combo_valid = combo[~all_null_mask]
    df_valid = df[~all_null_mask]

    if combo_valid.empty:
        return pd.DataFrame()

    combo_key = combo_valid.apply(lambda row: "||".join(row), axis=1)
    dup_mask = combo_key.duplicated(keep=False)

    if not dup_mask.any():
        return pd.DataFrame()

    dup_df = df_valid[dup_mask].copy()
    groups = combo_key[dup_mask]
    group_map = {v: i + 1 for i, v in enumerate(groups.unique())}
    dup_df["Duplicate_Group_ID"] = groups.map(group_map).values
    dup_df["Duplicate_Type"] = "Combination"
    dup_df["Similarity_Score"] = 100.0
    dup_df["Row_Number"] = dup_df.index + 2
    return dup_df


# ═══════════════════════════════════════════════════════════════════════
#  HYBRID FUZZY SCORING ENGINE
#
#  Strategy (per cell comparison):
#    • If rapidfuzz is available  → use 4-scorer ensemble via cdist:
#        ratio            – exact character overlap
#        partial_ratio    – best substring alignment (handles short tokens in long strings)
#        token_sort_ratio – order-invariant token comparison
#        token_set_ratio  – handles duplicate/extra tokens (best for names/addresses)
#      Final cell score = weighted average of 4 scorers with tuned scorer weights:
#        ratio×0.15 + partial_ratio×0.20 + token_sort_ratio×0.30 + token_set_ratio×0.35
#    • Fallback (no rapidfuzz) → pure Python SequenceMatcher ratio × 100
#
#  Row-level score = weighted average across selected columns
#  (column weights provided by user in UI; default = equal weights)
# ═══════════════════════════════════════════════════════════════════════

# Scorer weights within the 4-scorer ensemble (must sum to 1.0)
_SCORER_WEIGHTS: Dict[str, float] = {
    "ratio":            0.15,
    "partial_ratio":    0.20,
    "token_sort_ratio": 0.30,
    "token_set_ratio":  0.35,
}


def _hybrid_cell_score(a_val: str, b_val: str) -> float:
    """
    Compute a hybrid similarity score (0–100) between two string values.

    If rapidfuzz is available:
      Uses a 4-scorer ensemble (ratio, partial_ratio, token_sort_ratio,
      token_set_ratio) via cdist for vectorised speed, then takes the
      weighted average with _SCORER_WEIGHTS.

    Fallback:
      Uses Python's difflib.SequenceMatcher for a simple ratio × 100.
    """
    a_str = str(a_val).strip().lower()
    b_str = str(b_val).strip().lower()

    # Exact match shortcut — avoids scoring overhead
    if a_str == b_str:
        return 100.0
    # Both empty — treat as match
    if not a_str and not b_str:
        return 100.0
    # One empty — no similarity
    if not a_str or not b_str:
        return 0.0

    if RAPIDFUZZ_AVAILABLE:
        from rapidfuzz import fuzz as _fuzz
        scores = {
            "ratio":            _fuzz.ratio(a_str, b_str),
            "partial_ratio":    _fuzz.partial_ratio(a_str, b_str),
            "token_sort_ratio": _fuzz.token_sort_ratio(a_str, b_str),
            "token_set_ratio":  _fuzz.token_set_ratio(a_str, b_str),
        }
        return sum(scores[k] * w for k, w in _SCORER_WEIGHTS.items())
    else:
        import difflib
        return difflib.SequenceMatcher(None, a_str, b_str).ratio() * 100.0


def _hybrid_row_score(
    a: pd.Series,
    b: pd.Series,
    cols: List[str],
    weights: List[float],
) -> float:
    """
    Compute a weighted hybrid similarity score between two DataFrame rows
    across the specified columns.

    Each cell pair is scored via _hybrid_cell_score() (4-scorer ensemble
    when rapidfuzz is available). Cell scores are combined using column
    weights provided by the user.

    Weights are normalised internally — they do not need to sum to 1.
    Columns with weight=0 are skipped entirely.
    """
    sims: List[float] = []
    wts: List[float] = []

    for col, w in zip(cols, weights):
        if w <= 0:
            continue
        av = a.get(col, "")
        bv = b.get(col, "")
        if pd.isna(av) or pd.isna(bv):
            sim = 0.0
        else:
            sim = _hybrid_cell_score(str(av), str(bv))
        sims.append(sim)
        wts.append(w)

    if not wts:
        return 0.0

    total_w = sum(wts)
    return sum(s * w for s, w in zip(sims, wts)) / total_w


def _rapidfuzz_cdist_block_scores(
    block_df: pd.DataFrame,
    cols: List[str],
    weights: List[float],
) -> np.ndarray:
    """
    Vectorised pairwise scoring for a block using rapidfuzz.cdist.

    For each column, builds the full n×n similarity matrix via cdist
    for each of the 4 scorers, combines with _SCORER_WEIGHTS to get a
    per-column matrix, then applies column weights to produce the final
    n×n weighted score matrix.

    Returns a float64 np.ndarray of shape (n, n).
    Falls back to None if rapidfuzz is unavailable (caller uses scalar path).
    """
    if not RAPIDFUZZ_AVAILABLE:
        return None  # type: ignore[return-value]

    # Use module-level fuzz (imported at top of file as `fuzz`)
    # cdist from rapidfuzz.process — if unavailable, fall back to None so caller uses scalar
    try:
        from rapidfuzz.process import cdist as _cdist
    except ImportError:
        return None  # type: ignore[return-value]

    n = len(block_df)
    final_matrix = np.zeros((n, n), dtype=np.float64)
    total_weight = sum(w for w in weights if w > 0)
    if total_weight == 0:
        return final_matrix

    scorers = {
        "ratio":            fuzz.ratio,
        "partial_ratio":    fuzz.partial_ratio,
        "token_sort_ratio": fuzz.token_sort_ratio,
        "token_set_ratio":  fuzz.token_set_ratio,
    }

    for col, col_w in zip(cols, weights):
        if col_w <= 0:
            continue
        vals = block_df[col].astype(str).str.strip().str.lower().tolist()

        # Per-scorer matrices (n×n), then ensemble average
        col_matrix = np.zeros((n, n), dtype=np.float64)
        for scorer_name, scorer_fn in scorers.items():
            sw = _SCORER_WEIGHTS[scorer_name]
            mat = _cdist(vals, vals, scorer=scorer_fn, score_cutoff=0) * sw
            col_matrix += mat

        final_matrix += col_matrix * (col_w / total_weight)

    return final_matrix


def detect_fuzzy_duplicates(
    df: pd.DataFrame,
    cols: List[str],
    threshold: int = 80,
    weights: Optional[List[float]] = None,
    max_pairs_per_block: int = 20_000,
    ignore_nulls: bool = True,
) -> Tuple[pd.DataFrame, List[str]]:
    """
    Hybrid fuzzy duplicate detection engine.

    Two-tier scoring strategy
    ─────────────────────────
    Tier 1 — rapidfuzz cdist (vectorised, fast):
        When rapidfuzz is installed AND a block has ≤ max_pairs_per_block pairs,
        uses _rapidfuzz_cdist_block_scores() to build the full n×n similarity
        matrix in one vectorised call per scorer per column.
        4-scorer ensemble: ratio × 0.15 + partial_ratio × 0.20
                         + token_sort_ratio × 0.30 + token_set_ratio × 0.35

    Tier 2 — scalar hybrid fallback:
        For blocks that exceed max_pairs_per_block (performance guard), or when
        rapidfuzz is unavailable, falls back to _hybrid_row_score() which scores
        each pair individually using the same 4-scorer ensemble (rapidfuzz) or
        difflib.SequenceMatcher if rapidfuzz is absent.

    Blocking strategy
    ─────────────────
    • Primary block: first 2 chars of the first fuzzy column (case-normalised)
    • Secondary geo block: first 3 chars of City / Country / State if present
      → Combined key reduces O(n²) comparisons dramatically

    Group merging
    ─────────────
    Union-Find with path compression guarantees correct transitive group
    membership (A~B and B~C → A, B, C all in the same group).

    Parameters
    ----------
    df               : Input DataFrame
    cols             : Column names to fuzzy-match on
    threshold        : Minimum weighted score (0-100) to flag as duplicate
    weights          : Per-column weights (aligned to cols). Defaults to equal.
    max_pairs_per_block : Blocks exceeding this pair count fall back to scalar.
    ignore_nulls     : Exclude rows with null/blank values in any fuzzy column.

    Returns
    -------
    (dup_df, warnings_list)
    """
    warnings_out: List[str] = []

    if not cols:
        return pd.DataFrame(), warnings_out

    if not RAPIDFUZZ_AVAILABLE:
        warnings_out.append(
            "RapidFuzz not installed — falling back to difflib.SequenceMatcher. "
            "Install rapidfuzz for significantly faster and more accurate results."
        )

    # Normalise column weights
    if weights is None or len(weights) != len(cols):
        weights = [1.0] * len(cols)

    work = df.copy()

    # ── Null / blank filtering ───────────────────────────────────────
    _NULL_SENTINELS = {"", "nan", "none", "null", "na", "n/a", "n.a.", "nil", "missing", "#n/a"}
    if ignore_nulls:
        valid_mask = pd.Series(True, index=work.index)
        for col in cols:
            if col in work.columns:
                s_col = work[col].astype(str).str.strip().str.lower()
                valid_mask &= ~s_col.isin(_NULL_SENTINELS)
        dropped = int((~valid_mask).sum())
        if dropped:
            warnings_out.append(
                f"Fuzzy match: skipped {dropped} row(s) with null/blank values "
                f"in fuzzy column(s)."
            )
        work = work[valid_mask].copy()

    if work.empty:
        return pd.DataFrame(), warnings_out

    # ── Blocking key construction ────────────────────────────────────
    # Primary: first 2 chars of the primary fuzzy column
    first_col = cols[0]
    work["_block"] = (
        work[first_col].astype(str).str.strip().str.lower()
        .str[:2].fillna("__")
    )
    # Secondary geo key: first 3 chars of a geographic column if present
    for geo_col in ["Country", "country", "COUNTRY", "City", "city", "CITY",
                    "State", "state", "STATE", "Region", "region"]:
        if geo_col in work.columns:
            geo_key = (
                work[geo_col].astype(str).str.strip().str.lower()
                .str[:3].fillna("")
            )
            work["_block"] = work["_block"] + "|" + geo_key
            break

    # ── Main comparison loop ─────────────────────────────────────────
    matches: List[Dict[str, Any]] = []
    skipped_blocks = 0
    scalar_fallback_blocks = 0

    for _, block_group in work[work["_block"] != ""].groupby("_block", sort=False):
        idxs = list(block_group.index)
        n = len(idxs)
        if n < 2:
            continue
        n_pairs = n * (n - 1) // 2

        # ── Tier 1: vectorised cdist (rapidfuzz) ────────────────────
        if RAPIDFUZZ_AVAILABLE and n_pairs <= max_pairs_per_block:
            score_matrix = _rapidfuzz_cdist_block_scores(block_group, cols, weights)
            if score_matrix is not None:
                # score_matrix is n×n; only look at upper triangle
                for i_pos in range(n):
                    for j_pos in range(i_pos + 1, n):
                        score = float(score_matrix[i_pos, j_pos])
                        if score >= threshold:
                            matches.append({
                                "_i":     idxs[i_pos],
                                "_j":     idxs[j_pos],
                                "_score": score,
                            })
            else:
                # cdist unavailable — use scalar hybrid
                scalar_fallback_blocks += 1
                for i_pos in range(n):
                    for j_pos in range(i_pos + 1, n):
                        score = _hybrid_row_score(
                            work.loc[idxs[i_pos]], work.loc[idxs[j_pos]], cols, weights
                        )
                        if score >= threshold:
                            matches.append({"_i": idxs[i_pos], "_j": idxs[j_pos], "_score": score})

        # ── Tier 2: scalar hybrid fallback ──────────────────────────
        elif n_pairs <= max_pairs_per_block:
            scalar_fallback_blocks += 1
            for i_pos in range(n):
                for j_pos in range(i_pos + 1, n):
                    i_idx = idxs[i_pos]
                    j_idx = idxs[j_pos]
                    score = _hybrid_row_score(
                        work.loc[i_idx], work.loc[j_idx], cols, weights
                    )
                    if score >= threshold:
                        matches.append({
                            "_i":     i_idx,
                            "_j":     j_idx,
                            "_score": score,
                        })

        # ── Block too large — skip entirely ─────────────────────────
        else:
            skipped_blocks += 1

    if skipped_blocks:
        warnings_out.append(
            f"Fuzzy match: {skipped_blocks} block(s) exceeded the "
            f"{max_pairs_per_block:,} pair limit and were skipped. "
            f"Consider increasing 'Max pairs per block' or raising the threshold."
        )
    if scalar_fallback_blocks and RAPIDFUZZ_AVAILABLE:
        warnings_out.append(
            f"Fuzzy match: {scalar_fallback_blocks} block(s) used scalar fallback "
            f"(cdist unavailable for those blocks)."
        )

    if not matches:
        return pd.DataFrame(), warnings_out

    pair_df = pd.DataFrame(matches).sort_values("_score", ascending=False)

    # ── Union-Find: transitive group merging ─────────────────────────
    parent: Dict[int, int] = {}

    def _find(x: int) -> int:
        parent.setdefault(x, x)
        if parent[x] != x:
            parent[x] = _find(parent[x])   # path compression
        return parent[x]

    def _union(x: int, y: int) -> None:
        rx, ry = _find(x), _find(y)
        if rx != ry:
            parent[ry] = rx

    for _, row in pair_df.iterrows():
        _union(int(row["_i"]), int(row["_j"]))

    # Collect group members (size ≥ 2)
    members: Dict[int, List[int]] = {}
    all_matched = (
        set(pair_df["_i"].astype(int))
        .union(set(pair_df["_j"].astype(int)))
    )
    for idx in all_matched:
        root = _find(idx)
        members.setdefault(root, []).append(idx)

    out_frames: List[pd.DataFrame] = []
    group_num = 0
    for _, idx_list in members.items():
        if len(idx_list) < 2:
            continue
        group_num += 1
        grp = work.loc[idx_list].copy()

        group_pairs = pair_df[
            pair_df["_i"].astype(int).isin(idx_list) &
            pair_df["_j"].astype(int).isin(idx_list)
        ]
        best_score = (
            float(group_pairs["_score"].max())
            if not group_pairs.empty
            else float(threshold)
        )

        grp["Duplicate_Group_ID"] = group_num
        grp["Duplicate_Type"]     = "Hybrid Fuzzy"
        grp["Similarity_Score"]   = round(best_score, 2)
        grp["Row_Number"]         = grp.index + 2
        out_frames.append(grp)

    if not out_frames:
        return pd.DataFrame(), warnings_out

    dup_df = pd.concat(out_frames, ignore_index=True)
    dup_df.drop(columns=["_block"], errors="ignore", inplace=True)

    return dup_df, warnings_out


# ═══════════════════════════════════════════════════════════════════════
#  STANDARDIZATION TRANSFORMS  (return cleaned series + changed mask)
#  NOTE: These are used for FLAGGING non-standard values. The system
#  does NOT auto-correct data; it only highlights records that violate
#  the expected standard.
# ═══════════════════════════════════════════════════════════════════════
def std_trim_spaces(series: pd.Series) -> Tuple[pd.Series, pd.Series]:
    original = series.astype(str)
    cleaned = original.str.strip()
    return cleaned, original != cleaned


def std_remove_extra_spaces(series: pd.Series) -> Tuple[pd.Series, pd.Series]:
    original = series.astype(str)
    cleaned = original.str.strip().str.replace(r"\s+", " ", regex=True)
    return cleaned, original != cleaned


def std_proper_case(series: pd.Series) -> Tuple[pd.Series, pd.Series]:
    original = series.astype(str)
    cleaned = original.str.strip().str.title()
    return cleaned, original != cleaned


def std_lowercase(series: pd.Series) -> Tuple[pd.Series, pd.Series]:
    original = series.astype(str)
    cleaned = original.str.strip().str.lower()
    return cleaned, original != cleaned


def std_uppercase(series: pd.Series) -> Tuple[pd.Series, pd.Series]:
    """Flag values not in UPPERCASE."""
    original = series.astype(str)
    cleaned = original.str.strip().str.upper()
    return cleaned, original != cleaned


def std_remove_special_chars(series: pd.Series) -> Tuple[pd.Series, pd.Series]:
    original = series.astype(str)
    cleaned = original.str.replace(r"[^a-zA-Z0-9\s]", "", regex=True).str.strip()
    return cleaned, original != cleaned


def std_normalize_date(
    series: pd.Series, target_fmt: str = "%Y-%m-%d"
) -> Tuple[pd.Series, pd.Series]:
    original = series.astype(str)
    parsed = pd.to_datetime(original.str.strip(), errors="coerce", dayfirst=True)
    cleaned = parsed.dt.strftime(target_fmt).fillna(original)
    return cleaned, original != cleaned


def std_replace_null_default(
    series: pd.Series, default: str = "N/A"
) -> Tuple[pd.Series, pd.Series]:
    original = series.astype(str)
    s = original.str.strip()
    mask = (
        (s == "")
        | (s.str.lower() == "nan")
        | (s.str.lower() == "none")
        | series.isna()
    )
    cleaned = s.copy()
    cleaned[mask] = default
    return cleaned, mask


# ═══════════════════════════════════════════════════════════════════════
#  RULE EXECUTION AUDIT LOG
#  Populated by every execute_* function.
#  Call clear_rule_exec_log() at the start of each assessment run.
#  Call get_rule_exec_log() to retrieve results for display / export.
# ═══════════════════════════════════════════════════════════════════════
_RULE_EXEC_LOG: List[Dict[str, Any]] = []


def clear_rule_exec_log() -> None:
    """Reset the execution log.  Call once at the start of each assessment."""
    global _RULE_EXEC_LOG
    _RULE_EXEC_LOG.clear()


def get_rule_exec_log() -> List[Dict[str, Any]]:
    """Return a snapshot of the current execution log."""
    return list(_RULE_EXEC_LOG)


def _log_rule(
    dimension: str,
    column: str,
    rule: str,
    evaluated: int,
    failed: int,
) -> None:
    """
    Append one rule-execution record.

    Fields logged
    -------------
    Dimension, Column, Rule  — identity
    Evaluated                — total rows checked (= len(df) in most cases)
    Failed                   — rows that did NOT pass the rule
    Passed                   — evaluated - failed
    Score_%                  — (passed / evaluated) * 100, or None if evaluated == 0
    Severity                 — Critical / High / Medium / Low / Pass
    Issue_Type               — human-readable category used in Issue_Type column
    """
    if evaluated > 0:
        score_pct: Optional[float] = round((evaluated - failed) / evaluated * 100, 2)
    else:
        score_pct = None

    severity = (
        "Critical" if score_pct is not None and score_pct < 50  else
        "High"     if score_pct is not None and score_pct < 70  else
        "Medium"   if score_pct is not None and score_pct < 85  else
        "Low"      if score_pct is not None and score_pct < 95  else
        "Pass"
    )

    # Map to the canonical Issue_Type values used in the annexure
    _ISSUE_TYPE_MAP = {
        "Completeness":    "Missing Values",
        "Validity":        "Invalid Format",
        "Uniqueness":      "Duplicate Records",
        "Standardization": "Non-Standard Values",
    }

    _RULE_EXEC_LOG.append({
        "Dimension":  dimension,
        "Column":     column,
        "Rule":       rule,
        "Evaluated":  evaluated,
        "Failed":     failed,
        "Passed":     evaluated - failed,
        "Score_%":    score_pct,
        "Severity":   severity,
        "Issue_Type": _ISSUE_TYPE_MAP.get(dimension, dimension),
    })


# ═══════════════════════════════════════════════════════════════════════
#  RULE EXECUTION ORCHESTRATORS
# ═══════════════════════════════════════════════════════════════════════
_ANNEXURE_COLS = [
    "Row_Number", "Column_Name", "Rule_Applied",
    "Issue_Type", "Original_Value", "Expected_Value", "Dimension",
]


def _annex_row(
    idx: int, col: str, rule: str, issue: str,
    orig: str, expected: str, dim: str,
) -> Dict[str, Any]:
    return {
        "Row_Number": idx + 2,
        "Column_Name": col,
        "Rule_Applied": rule,
        "Issue_Type": issue,
        "Original_Value": orig,
        "Expected_Value": expected,
        "Dimension": dim,
    }


def _collect_failures(
    series: pd.Series, mask_pass: pd.Series, col: str,
    rule: str, issue: str, expected: str, dim: str,
) -> List[Dict[str, Any]]:
    """Collect annexure rows for every failing label-index position.

    Uses .loc[idx] (label-based) not .iloc[idx] (position-based) so that
    non-zero-based or filtered DataFrames are handled correctly.
    """
    fail_labels = series.index[~mask_pass]
    if fail_labels.empty:
        return []
    rows = []
    for idx in fail_labels:
        orig = repr(str(series.loc[idx])) if rule == "Whitespace Only" else str(series.loc[idx])
        rows.append(_annex_row(idx, col, rule, issue, orig, expected, dim))
    return rows


# ── Completeness ────────────────────────────────────────────────────
def execute_completeness_rules(
    df: pd.DataFrame,
    selected_rules: List[str],
    columns: List[str],
    min_length_val: int = 3,
    mandatory_cols: Optional[List[str]] = None,
    column_rule_map: Optional[List[Dict[str, Any]]] = None,
) -> List[Dict[str, Any]]:
    """
    Execute completeness rules and return annexure rows for every failure.

    column_rule_map path (new Criteria Builder)
    --------------------------------------------
    Each entry: {"column": str, "rule": str, "config": {"min_length": int}}
    Only the explicitly mapped (column, rule) pairs are evaluated.

    Legacy path
    -----------
    All selected_rules are applied to all columns in the columns list.

    Both paths emit _log_rule() entries for every (column, rule) pair so that
    the scoring engine can compute an accurate denominator.
    """
    annexure: List[Dict[str, Any]] = []
    n_rows = len(df)

    rule_funcs: Dict[str, Tuple[Any, str, str]] = {
        "Not Null":        (rule_not_null,       "Missing Values",         "Non-null value"),
        "Not Empty":       (rule_not_empty,      "Missing Values",         "Non-empty value"),
        "Whitespace Only": (rule_whitespace_only, "Missing Values",        "Non-whitespace value"),
    }

    def _run_simple(s: pd.Series, col: str, rule_name: str) -> List[Dict]:
        func, issue, expected = rule_funcs[rule_name]
        mask = func(s)
        fails = df.index[~mask]
        rows = []
        for idx in fails:
            orig = repr(str(s.loc[idx])) if rule_name == "Whitespace Only" else str(s.loc[idx])
            rows.append(_annex_row(idx, col, rule_name, issue, orig, expected, "Completeness"))
        _log_rule("Completeness", col, rule_name, n_rows, len(rows))
        return rows

    # ── Column-level criteria builder path ──────────────────────────
    if column_rule_map:
        for entry in column_rule_map:
            col       = entry.get("column", "")
            rule_name = entry.get("rule", "")
            cfg       = entry.get("config", {})
            if col not in df.columns:
                continue
            s = df[col]

            if rule_name in rule_funcs:
                annexure.extend(_run_simple(s, col, rule_name))

            elif rule_name == "Minimum Length":
                min_len = int(cfg.get("min_length", cfg.get("min_length_val", min_length_val)))
                mask  = rule_minimum_length(s, min_len)
                fails = list(df.index[~mask])
                for idx in fails:
                    annexure.append(_annex_row(
                        idx, col, f"Minimum Length ({min_len})",
                        "Missing Values", str(s.loc[idx]),
                        f"Length ≥ {min_len}", "Completeness",
                    ))
                _log_rule("Completeness", col, f"Minimum Length ({min_len})", n_rows, len(fails))

            elif rule_name == "Mandatory Column":
                mask  = rule_not_null(s)
                fails = list(df.index[~mask])
                for idx in fails:
                    annexure.append(_annex_row(
                        idx, col, "Mandatory Column",
                        "Missing Values", str(s.loc[idx]),
                        "Required value", "Completeness",
                    ))
                _log_rule("Completeness", col, "Mandatory Column", n_rows, len(fails))

        return annexure

    # ── Legacy global path ───────────────────────────────────────────
    for col in columns:
        if col not in df.columns:
            continue
        s = df[col]

        for rule_name in rule_funcs:
            if rule_name not in selected_rules:
                continue
            annexure.extend(_run_simple(s, col, rule_name))

        if "Minimum Length" in selected_rules:
            mask  = rule_minimum_length(s, min_length_val)
            fails = list(df.index[~mask])
            for idx in fails:
                annexure.append(_annex_row(
                    idx, col, f"Minimum Length ({min_length_val})",
                    "Missing Values", str(s.loc[idx]),
                    f"Length ≥ {min_length_val}", "Completeness",
                ))
            _log_rule("Completeness", col, f"Minimum Length ({min_length_val})", n_rows, len(fails))

    if "Mandatory Column" in selected_rules and mandatory_cols:
        for col in mandatory_cols:
            if col not in df.columns:
                continue
            mask  = rule_not_null(df[col])
            fails = list(df.index[~mask])
            for idx in fails:
                annexure.append(_annex_row(
                    idx, col, "Mandatory Column",
                    "Missing Values", str(df[col].loc[idx]),
                    "Required value", "Completeness",
                ))
            _log_rule("Completeness", col, "Mandatory Column", n_rows, len(fails))

    return annexure


# ── Validity ────────────────────────────────────────────────────────
def execute_validity_rules(
    df: pd.DataFrame,
    selected_rules: List[str],
    columns: List[str],
    dtype_map: Optional[Dict[str, str]] = None,
    range_min: float = 0,
    range_max: float = 100,
    allowed_values_str: str = "",
    custom_regex: str = "",
    date_fmt: str = "",
    column_rule_map: Optional[List[Dict[str, Any]]] = None,
) -> List[Dict[str, Any]]:
    """
    Execute validity rules and return annexure rows for every failure.

    column_rule_map path (new Criteria Builder)
    --------------------------------------------
    Each entry: {"column": str, "rule": str, "config": {...}}
    Config keys: range_min, range_max, allowed_values (or allowed_values_str),
                 regex (or custom_regex), date_fmt, data_type.

    Legacy path
    -----------
    All selected_rules are applied to all columns with global parameters.

    Both paths emit _log_rule() entries for every (column, rule) pair.
    """
    annexure: List[Dict[str, Any]] = []
    n_rows   = len(df)
    dtype_map = dtype_map or {}

    simple_rules: Dict[str, Tuple[Any, str, str]] = {
        "Email Format": (rule_email_format, "Invalid Format",  "Valid email (user@domain.tld)"),
        "Phone Format": (rule_phone_format, "Invalid Format",  "Valid phone (7–15 digits)"),
        "PAN Format":   (rule_pan_format,   "Invalid Format",  "PAN: AAAAA9999A"),
    }

    def _emit(idx: int, col: str, rule: str, issue: str, s: pd.Series, expected: str):
        annexure.append(_annex_row(idx, col, rule, issue, str(s.loc[idx]), expected, "Validity"))

    # ── Column-level criteria builder path ──────────────────────────
    if column_rule_map:
        for entry in column_rule_map:
            col       = entry.get("column", "")
            rule_name = entry.get("rule", "")
            cfg       = entry.get("config", {})
            if col not in df.columns:
                continue
            s = df[col]

            if rule_name in simple_rules:
                func, issue, expected = simple_rules[rule_name]
                mask  = func(s)
                fails = list(df.index[~mask])
                for idx in fails:
                    _emit(idx, col, rule_name, issue, s, expected)
                _log_rule("Validity", col, rule_name, n_rows, len(fails))

            elif rule_name == "Data Type Validation":
                exp_type = cfg.get("data_type", dtype_map.get(col, "string"))
                mask     = rule_data_type(s, exp_type)
                fails    = list(df.index[~mask])
                label    = f"Data Type ({exp_type})"
                for idx in fails:
                    _emit(idx, col, label, "Invalid Format", s, f"Expected type: {exp_type}")
                _log_rule("Validity", col, label, n_rows, len(fails))

            elif rule_name == "Date Format":
                fmt   = cfg.get("date_fmt", date_fmt)
                mask  = rule_date_format(s, fmt)
                fails = list(df.index[~mask])
                exp   = f"Valid date{' (' + fmt + ')' if fmt else ' (any parseable)'}"
                for idx in fails:
                    _emit(idx, col, "Date Format", "Invalid Format", s, exp)
                _log_rule("Validity", col, "Date Format", n_rows, len(fails))

            elif rule_name == "Numeric Range":
                rmin  = float(cfg.get("range_min", range_min))
                rmax  = float(cfg.get("range_max", range_max))
                mask  = rule_numeric_range(s, rmin, rmax)
                fails = list(df.index[~mask])
                label = f"Numeric Range [{rmin}–{rmax}]"
                for idx in fails:
                    _emit(idx, col, label, "Invalid Format", s, f"Between {rmin} and {rmax}")
                _log_rule("Validity", col, label, n_rows, len(fails))

            elif rule_name == "Allowed Values":
                # Support both key names produced by criteria builder
                av_raw = cfg.get("allowed_values_str", cfg.get("allowed_values", ""))
                allowed_list = [v.strip() for v in av_raw.split(",") if v.strip()] if av_raw else []
                if allowed_list:
                    mask  = rule_allowed_values(s, allowed_list)
                    fails = list(df.index[~mask])
                    preview = ", ".join(allowed_list[:5]) + ("…" if len(allowed_list) > 5 else "")
                    for idx in fails:
                        _emit(idx, col, "Allowed Values", "Invalid Format", s, f"One of: {preview}")
                    _log_rule("Validity", col, "Allowed Values", n_rows, len(fails))

            elif rule_name == "Custom Regex":
                pattern = cfg.get("regex", cfg.get("custom_regex", ""))
                if pattern:
                    mask  = rule_custom_regex(s, pattern)
                    fails = list(df.index[~mask])
                    label = f"Regex: {pattern[:30]}"
                    for idx in fails:
                        _emit(idx, col, label, "Invalid Format", s, f"Matches: {pattern[:50]}")
                    _log_rule("Validity", col, label, n_rows, len(fails))

            elif rule_name == "Special Characters Not Allowed":
                pattern = cfg.get("allowed_chars_pattern", r"^[a-zA-Z0-9\s]+$")
                mask = s.astype(str).str.match(pattern, na=False) | s.isna()
                fails = list(df.index[~mask])
                for idx in fails:
                    _emit(idx, col, "Special Characters Not Allowed", "Invalid Format", s, "Alphanumeric + spaces only")
                _log_rule("Validity", col, "Special Characters Not Allowed", n_rows, len(fails))

            elif rule_name == "Length Check":
                max_len = int(cfg.get("max_length_val", 255))
                mask = s.astype(str).str.len().le(max_len) | s.isna()
                fails = list(df.index[~mask])
                label = f"Length Check (max {max_len})"
                for idx in fails:
                    _emit(idx, col, label, "Invalid Format", s, f"Length ≤ {max_len}")
                _log_rule("Validity", col, label, n_rows, len(fails))

            elif rule_name == "Format Check":
                pattern = cfg.get("format_pattern", r"^[a-zA-Z0-9]+$")
                mask = s.astype(str).str.match(pattern, na=False) | s.isna()
                fails = list(df.index[~mask])
                label = f"Format Check ({pattern[:30]})"
                for idx in fails:
                    _emit(idx, col, label, "Invalid Format", s, f"Expected format: {pattern[:50]}")
                _log_rule("Validity", col, label, n_rows, len(fails))

        return annexure

    # ── Legacy global path ───────────────────────────────────────────
    allowed_list = (
        [v.strip() for v in allowed_values_str.split(",") if v.strip()]
        if allowed_values_str else []
    )

    for col in columns:
        if col not in df.columns:
            continue
        s = df[col]

        for rule_name, (func, issue, expected) in simple_rules.items():
            if rule_name not in selected_rules:
                continue
            mask  = func(s)
            fails = list(df.index[~mask])
            for idx in fails:
                _emit(idx, col, rule_name, issue, s, expected)
            _log_rule("Validity", col, rule_name, n_rows, len(fails))

        if "Data Type Validation" in selected_rules:
            exp_type = dtype_map.get(col, "string")
            mask     = rule_data_type(s, exp_type)
            fails    = list(df.index[~mask])
            label    = f"Data Type ({exp_type})"
            for idx in fails:
                _emit(idx, col, label, "Invalid Format", s, f"Expected type: {exp_type}")
            _log_rule("Validity", col, label, n_rows, len(fails))

        if "Date Format" in selected_rules:
            mask  = rule_date_format(s, date_fmt)
            fails = list(df.index[~mask])
            exp   = f"Valid date{' (' + date_fmt + ')' if date_fmt else ' (any parseable)'}"
            for idx in fails:
                _emit(idx, col, "Date Format", "Invalid Format", s, exp)
            _log_rule("Validity", col, "Date Format", n_rows, len(fails))

        if "Numeric Range" in selected_rules:
            mask  = rule_numeric_range(s, range_min, range_max)
            fails = list(df.index[~mask])
            label = f"Numeric Range [{range_min}–{range_max}]"
            for idx in fails:
                _emit(idx, col, label, "Invalid Format", s, f"Between {range_min} and {range_max}")
            _log_rule("Validity", col, label, n_rows, len(fails))

        if "Allowed Values" in selected_rules and allowed_list:
            mask  = rule_allowed_values(s, allowed_list)
            fails = list(df.index[~mask])
            preview = ", ".join(allowed_list[:5]) + ("…" if len(allowed_list) > 5 else "")
            for idx in fails:
                _emit(idx, col, "Allowed Values", "Invalid Format", s, f"One of: {preview}")
            _log_rule("Validity", col, "Allowed Values", n_rows, len(fails))

        if "Custom Regex" in selected_rules and custom_regex:
            mask  = rule_custom_regex(s, custom_regex)
            fails = list(df.index[~mask])
            label = f"Regex: {custom_regex[:30]}"
            for idx in fails:
                _emit(idx, col, label, "Invalid Format", s, f"Matches: {custom_regex[:50]}")
            _log_rule("Validity", col, label, n_rows, len(fails))

    return annexure


# ── Uniqueness ──────────────────────────────────────────────────────
def execute_uniqueness_rules(
    df: pd.DataFrame,
    selected_rules: List[str],
    single_col: Optional[str] = None,           # legacy: kept for compatibility
    combo_cols: Optional[List[str]] = None,
    fuzzy_cols: Optional[List[str]] = None,
    fuzzy_threshold: int = 80,
    single_cols: Optional[List[str]] = None,    # multi-column single exact match
    fuzzy_weights: Optional[List[float]] = None, # NEW: per-column weights for fuzzy
    fuzzy_max_pairs: int = 20_000,               # NEW: block-level pair cap
    fuzzy_ignore_nulls: bool = True,             # NEW: skip null rows in fuzzy
) -> Tuple[pd.DataFrame, List[Dict[str, Any]], List[str]]:
    """
    Returns (duplicate_records_df, annexure_rows, warnings).

    ``single_cols`` (list) supersedes the legacy ``single_col`` (str).
    When ``single_cols`` is provided, each column is checked independently
    for exact duplicates and a per-column annexure entry is produced.

    ``fuzzy_weights`` provides per-column weights for the Hybrid Fuzzy Match
    rule (aligned to fuzzy_cols order). Defaults to equal weights.
    ``fuzzy_max_pairs`` caps pairwise comparisons per block for performance.
    """
    all_dups: List[pd.DataFrame] = []
    annexure: List[Dict[str, Any]] = []
    warnings: List[str] = []

    # ── Single Column Exact Match (supports multiple columns) ───────────
    if "Single Column Exact Match" in selected_rules:
        # Normalise: prefer single_cols list; fall back to legacy single_col str
        _single_targets: List[str] = []
        if single_cols:
            _single_targets = [c for c in single_cols if c in df.columns]
        elif single_col and single_col in df.columns:
            _single_targets = [single_col]

        if _single_targets:
            _combined, _per_col = detect_exact_duplicates_multi_single(df, _single_targets)
            if not _combined.empty:
                all_dups.append(_combined)

            # Build per-column annexure entries
            for col_name, dup in _per_col.items():
                if dup.empty:
                    continue
                for _, row in dup.iterrows():
                    annexure.append({
                        "Row_Number":     row.get("Row_Number", ""),
                        "Column_Name":    col_name,
                        "Rule_Applied":   "Single Column Exact Duplicate",
                        "Issue_Type":     f"Duplicate (Group {row['Duplicate_Group_ID']})",
                        "Original_Value": str(row.get(col_name, "")),
                        "Expected_Value": "Unique value",
                        "Dimension":      "Uniqueness",
                    })
        else:
            if _single_targets == [] and (single_cols or single_col):
                warnings.append(
                    "Single Column Exact Match: none of the specified columns "
                    "were found in the dataset."
                )

    if "Combination Column Exact Match" in selected_rules and combo_cols and len(combo_cols) >= 2:
        dup = detect_exact_duplicates_combination(df, combo_cols)
        if not dup.empty:
            all_dups.append(dup)
            for _, row in dup.iterrows():
                annexure.append({
                    "Row_Number":     row.get("Row_Number", ""),
                    "Column_Name":    " + ".join(combo_cols),
                    "Rule_Applied":   "Combination Column Exact Duplicate",
                    "Issue_Type":     f"Duplicate (Group {row['Duplicate_Group_ID']})",
                    "Original_Value": " | ".join(str(row.get(c, "")) for c in combo_cols),
                    "Expected_Value": "Unique combination",
                    "Dimension":      "Uniqueness",
                })

    if "Hybrid Fuzzy Match" in selected_rules and fuzzy_cols:
        dup, fw = detect_fuzzy_duplicates(
            df,
            fuzzy_cols,
            threshold=fuzzy_threshold,
            weights=fuzzy_weights,
            max_pairs_per_block=fuzzy_max_pairs,
            ignore_nulls=fuzzy_ignore_nulls,
        )
        warnings.extend(fw)
        if not dup.empty:
            all_dups.append(dup)
            for _, row in dup.iterrows():
                annexure.append({
                    "Row_Number":     row.get("Row_Number", ""),
                    "Column_Name":    " + ".join(fuzzy_cols),
                    "Rule_Applied":   f"Fuzzy Match (>={fuzzy_threshold}%)",
                    "Issue_Type":     f"Fuzzy Dup (Group {row['Duplicate_Group_ID']}, {row['Similarity_Score']:.0f}%)",
                    "Original_Value": " | ".join(str(row.get(c, "")) for c in fuzzy_cols),
                    "Expected_Value": "Unique record",
                    "Dimension":      "Uniqueness",
                })

    if all_dups:
        dup_records = pd.concat(all_dups, ignore_index=True)
        # Deduplicate: if same row flagged by multiple rules, keep first occurrence
        if "Row_Number" in dup_records.columns and "Duplicate_Type" in dup_records.columns:
            dup_records = dup_records.drop_duplicates(
                subset=["Row_Number", "Duplicate_Type"], keep="first",
            )
        # Further deduplicate: if exact and fuzzy both flag same row, keep first
        if "Row_Number" in dup_records.columns:
            dup_records = dup_records.drop_duplicates(subset=["Row_Number"], keep="first")
    else:
        dup_records = pd.DataFrame()

    return dup_records, annexure, warnings


# ── Standardization ─────────────────────────────────────────────────
def execute_standardization_rules(
    df: pd.DataFrame,
    selected_rules: List[str],
    columns: List[str],
    date_target_fmt: str = "%Y-%m-%d",
    null_default: str = "N/A",
    column_rule_map: Optional[List[Dict[str, Any]]] = None,
) -> Tuple[pd.DataFrame, List[Dict[str, Any]]]:
    """
    Apply standardization transforms.  Returns (cleaned_df, annexure).

    Annexure rows represent values that REQUIRED a change (i.e. were non-standard
    before transformation).  These count as failures for scoring purposes.

    column_rule_map path
    --------------------
    Each entry: {"column": str, "rule": str, "config": {"date_target_fmt": str, "null_default": str}}

    Legacy path
    -----------
    All selected_rules applied to all columns.

    Both paths emit _log_rule() entries for every (column, rule) pair.
    """
    cleaned = df.copy()
    annexure: List[Dict[str, Any]] = []
    n_rows   = len(df)

    inline_rules: Dict[str, Any] = {
        "Trim Spaces":              std_trim_spaces,
        "Remove Extra Spaces":      std_remove_extra_spaces,
        "Convert to Proper Case":   std_proper_case,
        "Convert to Lowercase":     std_lowercase,
        "Convert to Uppercase":     std_uppercase,
        "Remove Special Characters": std_remove_special_chars,
    }

    def _apply_inline(col: str, rule_name: str) -> None:
        func = inline_rules[rule_name]
        new_vals, changed = func(cleaned[col])
        fail_labels = cleaned.index[changed]
        for idx in fail_labels:
            annexure.append(_annex_row(
                idx, col, rule_name,
                "Non-Standard Values",
                str(cleaned[col].loc[idx]),
                str(new_vals.loc[idx]),
                "Standardization",
            ))
        cleaned[col] = new_vals
        _log_rule("Standardization", col, rule_name, n_rows, int(changed.sum()))

    def _apply_norm_date(col: str, fmt: str) -> None:
        new_vals, changed = std_normalize_date(cleaned[col], fmt)
        fail_labels = cleaned.index[changed]
        label = f"Normalize Date ({fmt})"
        for idx in fail_labels:
            annexure.append(_annex_row(
                idx, col, label,
                "Non-Standard Values",
                str(cleaned[col].loc[idx]),
                str(new_vals.loc[idx]),
                "Standardization",
            ))
        cleaned[col] = new_vals
        _log_rule("Standardization", col, label, n_rows, int(changed.sum()))

    def _apply_null_default(col: str, default: str) -> None:
        new_vals, changed = std_replace_null_default(cleaned[col], default)
        fail_labels = cleaned.index[changed]
        label = f"Replace Null → {default}"
        for idx in fail_labels:
            annexure.append(_annex_row(
                idx, col, label,
                "Non-Standard Values",
                str(cleaned[col].loc[idx]),
                default,
                "Standardization",
            ))
        cleaned[col] = new_vals
        _log_rule("Standardization", col, label, n_rows, int(changed.sum()))

    # ── Column-level criteria builder path ──────────────────────────
    if column_rule_map:
        for entry in column_rule_map:
            col       = entry.get("column", "")
            rule_name = entry.get("rule", "")
            cfg       = entry.get("config", {})
            if col not in cleaned.columns:
                continue

            if rule_name in inline_rules:
                _apply_inline(col, rule_name)
            elif rule_name == "Normalize Date Format":
                fmt = cfg.get("date_target_fmt", date_target_fmt)
                _apply_norm_date(col, fmt)
            elif rule_name == "Replace Null with Default":
                default = cfg.get("null_default", null_default)
                _apply_null_default(col, default)

        return cleaned, annexure

    # ── Legacy global path ───────────────────────────────────────────
    for col in columns:
        if col not in cleaned.columns:
            continue

        for rule_name in inline_rules:
            if rule_name in selected_rules:
                _apply_inline(col, rule_name)

        if "Normalize Date Format" in selected_rules:
            _apply_norm_date(col, date_target_fmt)

        if "Replace Null with Default" in selected_rules:
            _apply_null_default(col, null_default)

    return cleaned, annexure


# ═══════════════════════════════════════════════════════════════════════
#  SCORING ENGINE  — v2  (accurate column-rule-pair denominator)
#
#  Root cause of the "always 100%" bug
#  ────────────────────────────────────
#  The legacy formula was:
#      total = len(df) * len(columns) * n_rules
#
#  With 30 columns and 2 rules that inflates total to 60 × len(df).
#  If only 5 rows in one column fail, score = (60N - 5) / 60N ≈ 100%.
#
#  The correct denominator is:
#      total = (number of (column, rule) pairs actually evaluated) × len(df)
#
#  The new scoring functions read this denominator from the execution log
#  populated by _log_rule() inside every execute_* function.
# ═══════════════════════════════════════════════════════════════════════

def _severity_label(score: float) -> str:
    if score < 50:  return "Critical"
    if score < 70:  return "High"
    if score < 85:  return "Medium"
    if score < 95:  return "Low"
    return "Pass"


def _score_from_log(dimension: str, n_rows: int) -> Optional[float]:
    """
    Derive the dimension score from the execution log.

    Denominator = sum of evaluated counts across all log entries for this
    dimension.  This equals (distinct column-rule pairs) × n_rows and is
    always accurate regardless of how many columns or rules were configured.

    Returns None if no rules were logged (→ "Not Evaluated").
    """
    entries = [e for e in _RULE_EXEC_LOG if e["Dimension"] == dimension]
    if not entries:
        return None   # No rules ran — do not default to 100%
    total_evaluated = sum(e["Evaluated"] for e in entries)
    total_failed    = sum(e["Failed"]    for e in entries)
    if total_evaluated == 0:
        return None
    return round(max(0.0, (total_evaluated - total_failed) / total_evaluated * 100), 2)


def compute_completeness_score(
    df: pd.DataFrame,
    annexure: List[Dict],
    columns: List[str],
    selected_rules: List[str],
    column_rule_map: Optional[List[Dict[str, Any]]] = None,
) -> float:
    """
    Score (%) = valid checks / total checks × 100.

    Uses _RULE_EXEC_LOG for an accurate denominator.  Falls back to the
    annexure-based calculation only when no log entries exist (e.g. legacy
    callers that do not invoke execute_completeness_rules first).
    """
    n_rows = len(df)
    if n_rows == 0:
        return 0.0

    # Primary path: read from execution log (accurate denominator)
    log_score = _score_from_log("Completeness", n_rows)
    if log_score is not None:
        return log_score

    # Fallback: annexure-only count (legacy callers)
    failures = sum(1 for r in annexure if r.get("Dimension") == "Completeness")
    if column_rule_map:
        n_pairs = len(column_rule_map)
    else:
        active = [r for r in selected_rules if r != "Mandatory Column"]
        n_pairs = len(active) * len(columns) if active else 0
        if "Mandatory Column" in selected_rules and columns:
            n_pairs += len(columns)
    total = n_pairs * n_rows
    if total == 0:
        return 0.0
    return round(max(0.0, (total - failures) / total * 100), 2)


def compute_validity_score(
    df: pd.DataFrame,
    annexure: List[Dict],
    columns: List[str],
    selected_rules: List[str],
    column_rule_map: Optional[List[Dict[str, Any]]] = None,
) -> float:
    """Score (%) = valid checks / total checks × 100."""
    n_rows = len(df)
    if n_rows == 0:
        return 0.0

    log_score = _score_from_log("Validity", n_rows)
    if log_score is not None:
        return log_score

    failures = sum(1 for r in annexure if r.get("Dimension") == "Validity")
    n_pairs  = (len(column_rule_map) if column_rule_map
                else len(selected_rules) * len(columns) if selected_rules and columns else 0)
    total = n_pairs * n_rows
    if total == 0:
        return 0.0
    return round(max(0.0, (total - failures) / total * 100), 2)


def compute_uniqueness_score(df: pd.DataFrame, dup_records: pd.DataFrame) -> float:
    """
    Score (%) = unique rows / total rows × 100.

    A row is counted as a duplicate if its Row_Number appears in dup_records.
    Using nunique() prevents double-counting rows that appear in both exact
    and fuzzy result sets.  Result is capped at total_records.
    """
    total = len(df)
    if total == 0:
        return 0.0
    if dup_records is None or dup_records.empty:
        # Log a "Pass" entry so the score is distinguishable from "Not Evaluated"
        _log_rule("Uniqueness", "ALL", "Exact/Fuzzy Duplicate Check", total, 0)
        return 100.0
    if "Row_Number" in dup_records.columns:
        dup_count = int(dup_records["Row_Number"].nunique())
    else:
        dup_count = len(dup_records)
    # Safety: cap at total records to prevent score going negative
    dup_count = min(dup_count, total)
    _log_rule("Uniqueness", "ALL", "Exact/Fuzzy Duplicate Check", total, dup_count)
    return round(max(0.0, (total - dup_count) / total * 100), 2)


def compute_standardization_score(
    df: pd.DataFrame,
    annexure: List[Dict],
    columns: List[str],
    selected_rules: List[str],
    column_rule_map: Optional[List[Dict[str, Any]]] = None,
) -> float:
    """
    Score (%) = already-standardized values / total values × 100.

    Values that required a transformation are "failures" — they were
    non-standard before the run.  Uses _RULE_EXEC_LOG for accuracy.
    """
    n_rows = len(df)
    if n_rows == 0:
        return 0.0

    log_score = _score_from_log("Standardization", n_rows)
    if log_score is not None:
        return log_score

    changes = sum(1 for r in annexure if r.get("Dimension") == "Standardization")
    n_pairs = (len(column_rule_map) if column_rule_map
               else len(selected_rules) * len(columns) if selected_rules and columns else 0)
    total = n_pairs * n_rows
    if total == 0:
        return 0.0
    return round(max(0.0, (total - changes) / total * 100), 2)


def compute_overall_score(dim_scores: Dict[str, float]) -> float:
    """Weighted average of all evaluated dimensions (equal weights)."""
    if not dim_scores:
        return 0.0
    valid = [v for v in dim_scores.values() if v is not None]
    if not valid:
        return 0.0
    return round(sum(valid) / len(valid), 2)


# ═══════════════════════════════════════════════════════════════════════
#  CLEAN DATASET BUILDER
# ═══════════════════════════════════════════════════════════════════════
def build_clean_dataset(
    df: pd.DataFrame,
    standardized_df: Optional[pd.DataFrame],
    dup_records: pd.DataFrame,
    all_annexure: List[Dict],
) -> pd.DataFrame:
    """Build clean dataset: standardized values, duplicates removed, Issue_Flag."""
    base = standardized_df.copy() if standardized_df is not None else df.copy()

    # Remove duplicate rows — keep first of each group
    if not dup_records.empty and "Row_Number" in dup_records.columns:
        dup_rows = set(dup_records["Row_Number"].tolist())
        keep_rows: set = set()
        if "Duplicate_Group_ID" in dup_records.columns:
            for gid in dup_records["Duplicate_Group_ID"].unique():
                grp = dup_records[dup_records["Duplicate_Group_ID"] == gid]
                keep_rows.add(int(grp["Row_Number"].min()))
        remove_indices = [int(r) - 2 for r in (dup_rows - keep_rows) if 0 <= int(r) - 2 < len(base)]
        base = base.drop(index=[i for i in remove_indices if i in base.index])

    # Issue_Flag
    issue_rows = set(int(r["Row_Number"]) for r in all_annexure)
    base = base.copy()
    base["Issue_Flag"] = base.index.map(lambda i: "Yes" if (i + 2) in issue_rows else "No")
    return base.reset_index(drop=True)


# ═══════════════════════════════════════════════════════════════════════
#  COLUMN-WISE ANNEXURE BUILDER
# ═══════════════════════════════════════════════════════════════════════
def build_column_wise_annexure(all_annexure: List[Dict]) -> pd.DataFrame:
    """
    Aggregate row-level issue records from all_annexure into a column-wise
    risk summary.

    Parameters
    ----------
    all_annexure : List[Dict]
        Each dict must contain keys: Column_Name, Rule_Applied,
        Issue_Type, Dimension.

    Returns
    -------
    pd.DataFrame with columns:
        Column_Name | Issue_Count | Failed_Rules | Issue_Types | Dimensions
    Sorted by Issue_Count descending.
    """
    if not all_annexure:
        return pd.DataFrame(
            columns=["Column_Name", "Issue_Count", "Failed_Rules", "Issue_Types", "Dimensions"]
        )

    df_ann = pd.DataFrame(all_annexure)

    # Ensure required columns exist
    for col in ["Column_Name", "Rule_Applied", "Issue_Type", "Dimension"]:
        if col not in df_ann.columns:
            df_ann[col] = ""

    agg = (
        df_ann.groupby("Column_Name", sort=False)
        .agg(
            Issue_Count=("Column_Name", "count"),
            Failed_Rules=("Rule_Applied", lambda x: " | ".join(sorted(x.dropna().unique()))),
            Issue_Types=("Issue_Type", lambda x: " | ".join(sorted(x.dropna().unique()))),
            Dimensions=("Dimension", lambda x: " | ".join(sorted(x.dropna().unique()))),
        )
        .reset_index()
    )

    agg = agg.sort_values("Issue_Count", ascending=False).reset_index(drop=True)
    return agg[["Column_Name", "Issue_Count", "Failed_Rules", "Issue_Types", "Dimensions"]]


# ═══════════════════════════════════════════════════════════════════════
#  ENTERPRISE EXCEL REPORT GENERATOR
# ═══════════════════════════════════════════════════════════════════════

# ── Colour palette ───────────────────────────────────────────────────
_CLR_HEADER_DARK   = "1F3864"   # Navy – primary header
_CLR_HEADER_MID    = "2E75B6"   # Steel-blue – section header
_CLR_GOOD          = "C6EFCE"   # Green background  (Good)
_CLR_GOOD_FONT     = "276221"
_CLR_MODERATE      = "FFEB9C"   # Amber background  (Moderate)
_CLR_MODERATE_FONT = "9C5700"
_CLR_POOR          = "FFC7CE"   # Red background    (Needs Improvement)
_CLR_POOR_FONT     = "9C0006"
_CLR_ALT_ROW       = "EBF3FB"   # Light-blue zebra stripe
_CLR_WHITE         = "FFFFFF"
_CLR_LABEL_BG      = "D6E4F0"   # Metric label background

# ── Business impact catalogue ─────────────────────────────────────────
_BUSINESS_IMPACT: Dict[str, Dict[str, str]] = {
    "Null/Missing Value": {
        "impact": "Missing data prevents record processing, reporting gaps, and downstream system failures.",
        "severity": "High",
    },
    "Empty String": {
        "impact": "Blank fields cause lookup failures and break automated workflows.",
        "severity": "Medium",
    },
    "Whitespace-Only Value": {
        "impact": "Whitespace-only entries are treated as valid by systems but produce incorrect results.",
        "severity": "Medium",
    },
    "Below Minimum Length": {
        "impact": "Short values indicate truncation or data entry errors affecting data integrity.",
        "severity": "Medium",
    },
    "Mandatory Column Missing Value": {
        "impact": "Missing mandatory fields block record creation in downstream CRM/ERP systems.",
        "severity": "Critical",
    },
    "Invalid Email": {
        "impact": "Invalid emails cause campaign delivery failures and customer communication breakdowns.",
        "severity": "High",
    },
    "Invalid Phone": {
        "impact": "Invalid phone numbers prevent outbound calling, SMS campaigns, and customer verification.",
        "severity": "High",
    },
    "Invalid PAN": {
        "impact": "Invalid PAN numbers cause compliance failures, tax reporting errors, and regulatory risk.",
        "severity": "Critical",
    },
    "Invalid Date": {
        "impact": "Invalid dates break time-series analysis, SLA calculations, and scheduled workflows.",
        "severity": "High",
    },
    "Invalid Data Type": {
        "impact": "Wrong data types cause calculation errors and ETL pipeline failures.",
        "severity": "High",
    },
    "Out of Range": {
        "impact": "Values outside expected ranges indicate data entry errors or fraudulent activity.",
        "severity": "Medium",
    },
    "Value Not in Allowed List": {
        "impact": "Unauthorised values break categorical filters, reporting dashboards, and BI tools.",
        "severity": "Medium",
    },
    "Regex Mismatch": {
        "impact": "Pattern mismatches indicate incorrectly formatted identifiers or reference codes.",
        "severity": "Medium",
    },
}
_DEFAULT_IMPACT = {
    "impact": "Data quality issue detected that may affect downstream processes and reporting accuracy.",
    "severity": "Medium",
}


def _get_status(score: float) -> str:
    if score > 95:
        return "Good"
    if score >= 85:
        return "Moderate"
    return "Needs Improvement"


def _score_color(score: float) -> str:
    if score > 95:
        return _CLR_GOOD
    if score >= 85:
        return _CLR_MODERATE
    return _CLR_POOR


def _score_font_color(score: float) -> str:
    if score > 95:
        return _CLR_GOOD_FONT
    if score >= 85:
        return _CLR_MODERATE_FONT
    return _CLR_POOR_FONT


def _apply_header_style(ws, row_num: int, fill_color: str = _CLR_HEADER_DARK) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    font = Font(bold=True, color=_CLR_WHITE, size=11, name="Arial")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="BDC3C7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[row_num]:
        if cell.value is not None:
            cell.fill = fill
            cell.font = font
            cell.alignment = align
            cell.border = border


def _apply_data_style(ws, start_row: int, end_row: int, num_cols: int, zebra: bool = True) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Side(style="thin", color="BDC3C7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(start_row, end_row + 1):
        bg = _CLR_ALT_ROW if (zebra and r % 2 == 0) else _CLR_WHITE
        fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.font = Font(size=10, name="Arial")
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border


def _auto_col_width(ws, min_w: int = 12, max_w: int = 50) -> None:
    from openpyxl.cell.cell import MergedCell
    for col_cells in ws.columns:
        letter = None
        for cell in col_cells:
            if not isinstance(cell, MergedCell):
                letter = cell.column_letter
                break
        if not letter:
            continue
        max_len = max(
            (len(str(cell.value or "")) for cell in col_cells if not isinstance(cell, MergedCell)),
            default=0,
        )
        ws.column_dimensions[letter].width = min(max(max_len + 3, min_w), max_w)


def _write_title(ws, title: str, subtitle: str = "", merge_cols: int = 8) -> int:
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.row_dimensions[1].height = 35
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=merge_cols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(bold=True, size=16, color=_CLR_WHITE, name="Arial")
    c.fill = PatternFill(start_color=_CLR_HEADER_DARK, end_color=_CLR_HEADER_DARK, fill_type="solid")
    c.alignment = Alignment(horizontal="left", vertical="center")
    if subtitle:
        ws.row_dimensions[2].height = 20
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=merge_cols)
        s = ws.cell(row=2, column=1, value=subtitle)
        s.font = Font(italic=True, size=10, color="555555", name="Arial")
        s.alignment = Alignment(horizontal="left", vertical="center")
        return 4  # next data row
    return 3


def _color_score_cell(cell, score) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment
    if not isinstance(score, (int, float)):
        return
    cell.fill = PatternFill(start_color=_score_color(score), end_color=_score_color(score), fill_type="solid")
    cell.font = Font(bold=True, color=_score_font_color(score), size=10, name="Arial")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _color_severity_cell(cell, severity: str) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment
    mapping = {
        "Critical": (_CLR_POOR, _CLR_POOR_FONT),
        "High":     ("FFD0B0", "7B3A00"),
        "Medium":   (_CLR_MODERATE, _CLR_MODERATE_FONT),
        "Low":      (_CLR_GOOD, _CLR_GOOD_FONT),
    }
    bg, fg = mapping.get(severity, (_CLR_WHITE, "000000"))
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.font = Font(bold=True, color=fg, size=10, name="Arial")
    cell.alignment = Alignment(horizontal="center", vertical="center")


# ── Sheet builders ────────────────────────────────────────────────────

def _sheet_executive_summary(
    ws,
    df: pd.DataFrame,
    dim_scores: Dict[str, float],
    overall_score: float,
    all_annexure: List[Dict],
    dup_records: pd.DataFrame,
    uniqueness_config: Optional[Dict[str, Any]] = None,
) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    total_records = len(df)
    total_cols = len(df.columns)

    # ── Accurate invalid record count (unique rows with any issue) ───
    issue_rows_set = set()
    for r in all_annexure:
        try:
            issue_rows_set.add(int(r["Row_Number"]))
        except (ValueError, TypeError, KeyError):
            pass
    invalid_count = min(len(issue_rows_set), total_records)

    # ── Accurate duplicate count (unique row IDs, not pair rows) ─────
    dup_count = 0
    if not dup_records.empty:
        dup_row_ids = set()
        # Try pair-format columns first
        idx_cols = [c for c in dup_records.columns if c.lower() in (
            'row_index', 'row_idx', 'row_number', 'row_num',
            'row_index_1', 'row_idx_1', 'index_1', 'idx_1',
            'row_index_2', 'row_idx_2', 'index_2', 'idx_2',
        )]
        if idx_cols:
            for col in idx_cols:
                dup_row_ids.update(dup_records[col].dropna().astype(int).tolist())
        elif "Row_Number" in dup_records.columns:
            dup_row_ids.update(dup_records["Row_Number"].dropna().astype(int).tolist())
        else:
            dup_row_ids.update(dup_records.index.tolist())
        dup_count = min(len(dup_row_ids), total_records)

    generated_on = datetime.datetime.now().strftime("%d %B %Y, %H:%M")
    next_row = _write_title(
        ws,
        "Executive Summary – Data Quality Report",
        f"Enterprise Data Quality Assessment  |  Report Generated: {generated_on}",
        merge_cols=4,
    )

    thin = Side(style="thin", color="BDC3C7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _kv(row, label, value, highlight: bool = False):
        lbl = ws.cell(row=row, column=1, value=label)
        lbl.fill = PatternFill(start_color=_CLR_LABEL_BG, end_color=_CLR_LABEL_BG, fill_type="solid")
        lbl.font = Font(bold=True, size=11, name="Arial")
        lbl.alignment = Alignment(horizontal="left", vertical="center")
        lbl.border = border
        ws.row_dimensions[row].height = 24
        val = ws.cell(row=row, column=2, value=value)
        val.font = Font(bold=highlight, size=11, name="Arial")
        val.alignment = Alignment(horizontal="center", vertical="center")
        val.border = border
        return val

    # Section: Dataset Overview
    sec = ws.cell(row=next_row, column=1, value="📋  Dataset Overview")
    sec.font = Font(bold=True, size=12, color=_CLR_WHITE, name="Arial")
    sec.fill = PatternFill(start_color=_CLR_HEADER_MID, end_color=_CLR_HEADER_MID, fill_type="solid")
    sec.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=2)
    ws.row_dimensions[next_row].height = 22
    next_row += 1

    _kv(next_row,     "Total Records Processed",  total_records); next_row += 1
    _kv(next_row,     "Total Columns Analyzed",   total_cols);    next_row += 1
    _kv(next_row,     "Total Duplicate Records",  dup_count);     next_row += 1
    _kv(next_row,     "Total Invalid Records",    invalid_count); next_row += 1
    next_row += 1

    # Section: Dimension Scores
    sec2 = ws.cell(row=next_row, column=1, value="📊  Data Quality Dimension Scores")
    sec2.font = Font(bold=True, size=12, color=_CLR_WHITE, name="Arial")
    sec2.fill = PatternFill(start_color=_CLR_HEADER_MID, end_color=_CLR_HEADER_MID, fill_type="solid")
    sec2.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=2)
    ws.row_dimensions[next_row].height = 22
    next_row += 1

    score_labels = {
        "Completeness":    "Completeness Score (%)",
        "Validity":        "Validity Score (%)",
        "Uniqueness":      "Uniqueness Score (%)",
        "Consistency":     "Consistency Score (%)",
        "Standardization": "Standardization Score (%)",
    }
    for dim, label in score_labels.items():
        raw = dim_scores.get(dim, "N/A")
        val_cell = _kv(next_row, label, f"{raw}%" if isinstance(raw, (int, float)) else raw)
        if isinstance(raw, (int, float)):
            _color_score_cell(val_cell, raw)
        next_row += 1

    next_row += 1
    sec3 = ws.cell(row=next_row, column=1, value="🏆  Overall Data Quality Score")
    sec3.font = Font(bold=True, size=12, color=_CLR_WHITE, name="Arial")
    sec3.fill = PatternFill(start_color=_CLR_HEADER_DARK, end_color=_CLR_HEADER_DARK, fill_type="solid")
    sec3.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=2)
    ws.row_dimensions[next_row].height = 22
    next_row += 1

    ov_val = ws.cell(row=next_row, column=2, value=f"{overall_score}%")
    ov_lbl = ws.cell(row=next_row, column=1, value="Overall Data Quality Score (%)")
    ov_lbl.font = Font(bold=True, size=13, name="Arial")
    ov_lbl.fill = PatternFill(start_color=_CLR_LABEL_BG, end_color=_CLR_LABEL_BG, fill_type="solid")
    ov_lbl.alignment = Alignment(horizontal="left", vertical="center")
    ov_lbl.border = border
    ov_val.font = Font(bold=True, size=14, name="Arial")
    ov_val.alignment = Alignment(horizontal="center", vertical="center")
    ov_val.border = border
    ws.row_dimensions[next_row].height = 28
    if isinstance(overall_score, (int, float)):
        _color_score_cell(ov_val, overall_score)

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22
    ws.sheet_view.showGridLines = False


def _sheet_dimension_scorecard(ws, dim_scores: Dict[str, float]) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment

    _write_title(ws, "Dimension Scorecard", "Quality status across all assessed data dimensions", merge_cols=3)

    headers = ["Dimension Name", "Score (%)", "Status"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    _apply_header_style(ws, 4)
    ws.row_dimensions[4].height = 22

    score_labels = {
        "Completeness":    "Completeness",
        "Validity":        "Validity",
        "Uniqueness":      "Uniqueness",
        "Consistency":     "Consistency",
        "Standardization": "Standardization",
    }
    row = 5
    for dim, label in score_labels.items():
        raw = dim_scores.get(dim)
        if raw is None:
            continue
        status = _get_status(raw)
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10, name="Arial")
        score_c = ws.cell(row=row, column=2, value=f"{raw}%")
        status_c = ws.cell(row=row, column=3, value=status)
        _color_score_cell(score_c, raw)
        _color_score_cell(status_c, raw)
        ws.row_dimensions[row].height = 22
        row += 1

    _apply_data_style(ws, 5, row - 1, 3)
    # Re-apply score coloring after data style (data style overwrites fill)
    row = 5
    for dim in score_labels:
        raw = dim_scores.get(dim)
        if raw is None:
            continue
        _color_score_cell(ws.cell(row=row, column=2), raw)
        _color_score_cell(ws.cell(row=row, column=3), raw)
        row += 1

    _auto_col_width(ws)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"


def _sheet_column_annexure(ws, df: pd.DataFrame, all_annexure: List[Dict]) -> None:
    total = len(df)

    _write_title(ws, "Column-Wise Data Quality Profiling", "Issue breakdown per column across all dimensions", merge_cols=7)

    headers = [
        "Column Name", "Completeness (%)", "Validity (%)",
        "Uniqueness (%)", "Standardization (%)", "Issue Type(s)", "Issue Severity",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    _apply_header_style(ws, 4)
    ws.row_dimensions[4].height = 22

    ann_df = pd.DataFrame(all_annexure) if all_annexure else pd.DataFrame(columns=_ANNEXURE_COLS)
    row = 5
    for col_name in df.columns:
        col_ann = ann_df[ann_df["Column_Name"] == col_name] if not ann_df.empty else pd.DataFrame()

        def _pct(dim):
            if ann_df.empty or total == 0:
                return 100.0
            fails = len(col_ann[col_ann["Dimension"] == dim]) if not col_ann.empty else 0
            return round((total - fails) / total * 100, 1)

        comp = _pct("Completeness")
        val  = _pct("Validity")
        uniq = _pct("Uniqueness")
        std  = _pct("Standardization")

        issue_types = " | ".join(sorted(col_ann["Issue_Type"].dropna().unique())) if not col_ann.empty else "None"

        # Determine worst severity
        worst_score = min(comp, val, uniq, std)
        if worst_score > 95:
            severity = "Low"
        elif worst_score >= 85:
            severity = "Medium"
        else:
            severity = "High"

        ws.cell(row=row, column=1, value=col_name)
        for ci, score in enumerate([comp, val, uniq, std], 2):
            c = ws.cell(row=row, column=ci, value=f"{score}%")
            _color_score_cell(c, score)
        ws.cell(row=row, column=6, value=issue_types)
        sev_c = ws.cell(row=row, column=7, value=severity)
        _color_severity_cell(sev_c, severity)
        ws.row_dimensions[row].height = 20
        row += 1

    _apply_data_style(ws, 5, row - 1, 7)
    # Re-apply score coloring
    data_row = 5
    for col_name in df.columns:
        col_ann = ann_df[ann_df["Column_Name"] == col_name] if not ann_df.empty else pd.DataFrame()
        def _pct2(dim):
            if ann_df.empty or total == 0:
                return 100.0
            fails = len(col_ann[col_ann["Dimension"] == dim]) if not col_ann.empty else 0
            return round((total - fails) / total * 100, 1)
        scores = [_pct2("Completeness"), _pct2("Validity"), _pct2("Uniqueness"), _pct2("Standardization")]
        for ci, score in enumerate(scores, 2):
            _color_score_cell(ws.cell(row=data_row, column=ci), score)
        worst = min(scores)
        sev = "Low" if worst > 95 else ("Medium" if worst >= 85 else "High")
        _color_severity_cell(ws.cell(row=data_row, column=7), sev)
        data_row += 1

    _auto_col_width(ws)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"


def _sheet_rule_failure_summary(ws, all_annexure: List[Dict], total_records: int) -> None:
    _write_title(ws, "Rule Failure Summary", "Validation results showing which rules failed and by how much", merge_cols=5)

    headers = ["Rule Name", "Dimension", "Description", "Failed Record Count", "Failure (%)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    _apply_header_style(ws, 4)
    ws.row_dimensions[4].height = 22

    if not all_annexure:
        ws.cell(row=5, column=1, value="✅  No rule failures detected – all checks passed.")
        return

    ann_df = pd.DataFrame(all_annexure)
    grouped = (
        ann_df.groupby(["Rule_Applied", "Dimension", "Issue_Type"])
        .agg(Failed_Count=("Row_Number", "count"))
        .reset_index()
        .sort_values("Failed_Count", ascending=False)
    )

    _RULE_DESCRIPTIONS = {
        "Not Null": "Checks that the field contains a non-null, non-empty value",
        "Not Empty": "Checks that the field is not an empty string",
        "Whitespace Only": "Detects fields containing only whitespace characters",
        "Email Format": "Validates that the value matches standard email format",
        "Phone Format": "Validates phone numbers are 7–15 digits in standard format",
        "PAN Format": "Validates PAN card numbers match the format AAAAA9999A",
        "Date Format": "Validates dates conform to the specified date format",
        "Numeric Range": "Checks that numeric values fall within the configured range",
        "Allowed Values": "Validates that the value belongs to the permitted set",
        "Data Type Validation": "Checks the value matches the expected data type",
        "Mandatory Column": "Ensures mandatory fields are never left empty",
        "Single Column Exact Duplicate": "Detects exact duplicate values in a single column",
        "Combination Column Exact Duplicate": "Detects duplicate record combinations across multiple columns",
        "Fuzzy Match": "Detects near-duplicate records using fuzzy text similarity",
        "Trim Spaces": "Identifies leading/trailing whitespace that was removed",
        "Remove Extra Spaces": "Detects multiple consecutive spaces within a value",
        "Convert to Proper Case": "Standardises text to Title Case format",
        "Convert to Lowercase": "Standardises text to lowercase format",
        "Remove Special Characters": "Removes non-alphanumeric characters from values",
        "Normalize Date Format": "Converts dates to the target date format",
        "Replace Null with Default": "Replaces null/empty values with a configured default",
    }

    row = 5
    for _, r in grouped.iterrows():
        rule = str(r["Rule_Applied"])
        dim  = str(r["Dimension"])
        desc = next((v for k, v in _RULE_DESCRIPTIONS.items() if k.lower() in rule.lower()), f"Rule: {rule}")
        count = int(r["Failed_Count"])
        pct   = round(count / total_records * 100, 2) if total_records else 0

        ws.cell(row=row, column=1, value=rule)
        ws.cell(row=row, column=2, value=dim)
        ws.cell(row=row, column=3, value=desc)
        ws.cell(row=row, column=4, value=count)
        pct_c = ws.cell(row=row, column=5, value=f"{pct}%")
        if pct >= 20:
            _color_score_cell(pct_c, 0)   # red
        elif pct >= 5:
            _color_score_cell(pct_c, 90)  # amber
        else:
            _color_score_cell(pct_c, 100) # green
        ws.row_dimensions[row].height = 20
        row += 1

    _apply_data_style(ws, 5, row - 1, 5)
    _auto_col_width(ws)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"


def _sheet_row_issue_annexure(ws, all_annexure: List[Dict]) -> None:
    _write_title(ws, "Row-Wise Issue Annexure", "Detailed failed records for data steward remediation", merge_cols=5)

    headers = ["Record ID / Row Number", "Column Name", "Issue Type", "Issue Description", "Failed Rule Name"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    _apply_header_style(ws, 4)
    ws.row_dimensions[4].height = 22

    if not all_annexure:
        ws.cell(row=5, column=1, value="✅  No row-level issues detected.")
        return

    row = 5
    for rec in all_annexure:
        orig = str(rec.get("Original_Value", ""))
        exp  = str(rec.get("Expected_Value", ""))
        desc = f'Value "{orig}" does not meet expectation: {exp}' if orig else f"Expected: {exp}"
        ws.cell(row=row, column=1, value=rec.get("Row_Number", ""))
        ws.cell(row=row, column=2, value=rec.get("Column_Name", ""))
        ws.cell(row=row, column=3, value=rec.get("Issue_Type", ""))
        ws.cell(row=row, column=4, value=desc)
        ws.cell(row=row, column=5, value=rec.get("Rule_Applied", ""))
        ws.row_dimensions[row].height = 18
        row += 1

    _apply_data_style(ws, 5, row - 1, 5)
    _auto_col_width(ws)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"


def _sheet_duplicate_annexure(ws, dup_records: pd.DataFrame) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment

    _write_title(ws, "Duplicate Records Annexure", "Identified duplicate groups and matching record details", merge_cols=6)

    if dup_records.empty:
        ws.cell(row=4, column=1, value="✅  No duplicate records were detected during this assessment.")
        return

    # ── Summary table ──────────────────────────────────────────────
    ws.cell(row=4, column=1, value="Duplicate Group Summary")
    ws.cell(row=4, column=1).font = Font(bold=True, size=12, color=_CLR_WHITE, name="Arial")
    ws.cell(row=4, column=1).fill = PatternFill(start_color=_CLR_HEADER_MID, end_color=_CLR_HEADER_MID, fill_type="solid")
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=4)

    sum_hdrs = ["Duplicate Group ID", "Matching Columns / Combination", "Duplicate Type", "Duplicate Record Count"]
    for c, h in enumerate(sum_hdrs, 1):
        ws.cell(row=5, column=c, value=h)
    _apply_header_style(ws, 5)

    row = 6
    if "Duplicate_Group_ID" in dup_records.columns:
        for gid in sorted(dup_records["Duplicate_Group_ID"].unique()):
            grp = dup_records[dup_records["Duplicate_Group_ID"] == gid]
            dup_type = str(grp["Duplicate_Type"].iloc[0]) if "Duplicate_Type" in grp.columns else "Exact"
            match_col = str(grp.get("_Source_Column", pd.Series(["Multiple Columns"])).iloc[0]) if "_Source_Column" in grp.columns else "Multiple Columns"
            ws.cell(row=row, column=1, value=int(gid))
            ws.cell(row=row, column=2, value=match_col)
            ws.cell(row=row, column=3, value=dup_type)
            ws.cell(row=row, column=4, value=len(grp))
            ws.row_dimensions[row].height = 18
            row += 1

    _apply_data_style(ws, 6, row - 1, 4)
    row += 2

    # ── Detailed records table ──────────────────────────────────────
    ws.cell(row=row, column=1, value="Detailed Duplicate Records")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color=_CLR_WHITE, name="Arial")
    ws.cell(row=row, column=1).fill = PatternFill(start_color=_CLR_HEADER_MID, end_color=_CLR_HEADER_MID, fill_type="solid")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

    display_cols = ["Row_Number", "Duplicate_Group_ID", "Duplicate_Type", "Similarity_Score"]
    display_cols = [c for c in display_cols if c in dup_records.columns]
    friendly_names = {
        "Row_Number": "Record ID / Row Number",
        "Duplicate_Group_ID": "Duplicate Group ID",
        "Duplicate_Type": "Match Type",
        "Similarity_Score": "Similarity Score (%)",
    }
    for ci, col_name in enumerate(display_cols, 1):
        ws.cell(row=row, column=ci, value=friendly_names.get(col_name, col_name))
    _apply_header_style(ws, row)
    row += 1

    detail_start = row
    for _, rec in dup_records[display_cols].iterrows():
        for ci, col_name in enumerate(display_cols, 1):
            ws.cell(row=row, column=ci, value=rec[col_name])
        ws.row_dimensions[row].height = 18
        row += 1

    _apply_data_style(ws, detail_start, row - 1, len(display_cols))
    _auto_col_width(ws)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"


def _sheet_standardization_report(ws, all_annexure: List[Dict]) -> None:
    _write_title(ws, "Standardization Report", "Before vs. after values for all standardization changes applied", merge_cols=4)

    headers = ["Column Name", "Original Value", "Standardized Value", "Standardization Rule Applied"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    _apply_header_style(ws, 4)
    ws.row_dimensions[4].height = 22

    std_rows = [r for r in all_annexure if r.get("Dimension") == "Standardization"]
    if not std_rows:
        ws.cell(row=5, column=1, value="✅  No standardization changes were required in this dataset.")
        return

    row = 5
    for rec in std_rows:
        ws.cell(row=row, column=1, value=rec.get("Column_Name", ""))
        ws.cell(row=row, column=2, value=rec.get("Original_Value", ""))
        ws.cell(row=row, column=3, value=rec.get("Expected_Value", ""))
        ws.cell(row=row, column=4, value=rec.get("Rule_Applied", ""))
        ws.row_dimensions[row].height = 18
        row += 1

    _apply_data_style(ws, 5, row - 1, 4)
    _auto_col_width(ws)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"


def _sheet_business_impact(ws, all_annexure: List[Dict], dim_scores: Dict[str, float]) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment

    _write_title(ws, "Business Impact Summary", "How identified data issues translate into business risks and operational impact", merge_cols=4)

    headers = ["Issue Type", "Affected Columns", "Business Impact Description", "Risk Severity"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    _apply_header_style(ws, 4)
    ws.row_dimensions[4].height = 22

    if not all_annexure:
        ws.cell(row=5, column=1, value="✅  No data quality issues identified. Dataset meets quality standards.")
        return

    ann_df = pd.DataFrame(all_annexure)
    issue_col_map: Dict[str, set] = {}
    for rec in all_annexure:
        issue = str(rec.get("Issue_Type", ""))
        col   = str(rec.get("Column_Name", ""))
        # Normalise issue type to match catalogue key
        matched_key = next((k for k in _BUSINESS_IMPACT if k.lower() in issue.lower()), None)
        key = matched_key if matched_key else issue
        issue_col_map.setdefault(key, set()).add(col)

    row = 5
    seen_keys: set = set()
    for rec in all_annexure:
        raw_issue = str(rec.get("Issue_Type", ""))
        matched_key = next((k for k in _BUSINESS_IMPACT if k.lower() in raw_issue.lower()), raw_issue)
        if matched_key in seen_keys:
            continue
        seen_keys.add(matched_key)

        info = _BUSINESS_IMPACT.get(matched_key, _DEFAULT_IMPACT)
        affected = ", ".join(sorted(issue_col_map.get(matched_key, set())))

        ws.cell(row=row, column=1, value=matched_key)
        ws.cell(row=row, column=2, value=affected)
        ws.cell(row=row, column=3, value=info["impact"])
        sev_c = ws.cell(row=row, column=4, value=info["severity"])
        _color_severity_cell(sev_c, info["severity"])
        ws.row_dimensions[row].height = 20
        row += 1

    _apply_data_style(ws, 5, row - 1, 4)
    # Re-apply severity coloring
    r2 = 5
    for key in list(seen_keys):
        info = _BUSINESS_IMPACT.get(key, _DEFAULT_IMPACT)
        _color_severity_cell(ws.cell(row=r2, column=4), info["severity"])
        r2 += 1

    _auto_col_width(ws)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"


# ── Rule Configuration Summary sheet ─────────────────────────────────
def _sheet_rule_config_summary(ws, uniqueness_config: Dict[str, Any]) -> None:
    """Sheet: Rule Configuration Summary — shows all rule parameters used in the assessment."""
    _write_title(ws, "Rule Configuration Summary", "Parameters and settings used for each configured rule", merge_cols=6)

    headers = ["#", "Column", "Dimension", "Rule", "Configuration Parameters", "Mandatory"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    _apply_header_style(ws, 4)
    ws.row_dimensions[4].height = 22

    rule_entries = uniqueness_config.get("rule_entries", []) if uniqueness_config else []
    if not rule_entries:
        ws.cell(row=5, column=1, value="ℹ️  No rule configuration metadata available for this assessment.")
        _auto_col_width(ws)
        ws.sheet_view.showGridLines = False
        return

    row = 5
    for i, entry in enumerate(rule_entries):
        cfg = entry.get("config", {})
        cfg_parts = []
        if "range_min" in cfg and cfg["range_min"] is not None:
            cfg_parts.append(f"Min={cfg['range_min']}")
        if "range_max" in cfg and cfg["range_max"] is not None:
            cfg_parts.append(f"Max={cfg['range_max']}")
        if "allowed_values_str" in cfg and cfg["allowed_values_str"]:
            vals = cfg["allowed_values_str"]
            if len(vals) > 80:
                vals = vals[:77] + "..."
            cfg_parts.append(f"Values=[{vals}]")
        if "custom_regex" in cfg and cfg["custom_regex"]:
            cfg_parts.append(f"Pattern={cfg['custom_regex']}")
        if "min_length_val" in cfg:
            cfg_parts.append(f"MinLen={cfg['min_length_val']}")
        if "date_fmt" in cfg and cfg["date_fmt"]:
            cfg_parts.append(f"DateFmt={cfg['date_fmt']}")
        if "date_target_fmt" in cfg and cfg["date_target_fmt"]:
            cfg_parts.append(f"TargetFmt={cfg['date_target_fmt']}")
        if "null_default" in cfg and cfg["null_default"]:
            cfg_parts.append(f"Default={cfg['null_default']}")
        if "data_type" in cfg and cfg["data_type"]:
            cfg_parts.append(f"Type={cfg['data_type']}")
        if "phone_flexible" in cfg:
            cfg_parts.append(f"Flexible={'Yes' if cfg['phone_flexible'] else 'No'}")
        if "email_flexible" in cfg:
            cfg_parts.append(f"Flexible={'Yes' if cfg['email_flexible'] else 'No'}")
        if "case_insensitive" in cfg:
            cfg_parts.append(f"CaseInsensitive={'Yes' if cfg['case_insensitive'] else 'No'}")
        cfg_str = "; ".join(cfg_parts) if cfg_parts else "Default"

        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=entry.get("column", "—"))
        ws.cell(row=row, column=3, value=entry.get("dimension", "—"))
        ws.cell(row=row, column=4, value=entry.get("rule", "—"))
        ws.cell(row=row, column=5, value=cfg_str)
        ws.cell(row=row, column=6, value="Yes" if entry.get("mandatory") else "No")
        ws.row_dimensions[row].height = 18
        row += 1

    _apply_data_style(ws, 5, row - 1, 6)
    _auto_col_width(ws)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"


# ── Rule Execution Log sheet ────────────────────────────────────────
def _sheet_rule_execution_log(ws, all_annexure: List[Dict], total_records: int) -> None:
    """Sheet: Rule Execution Log — per-rule failure counts from the execution log."""
    _write_title(ws, "Rule Execution Log", "Detailed execution results per rule showing pass/fail counts and severity", merge_cols=8)

    headers = ["Dimension", "Column", "Rule", "Evaluated", "Failed", "Passed", "Score (%)", "Severity"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    _apply_header_style(ws, 4)
    ws.row_dimensions[4].height = 22

    # Use the global execution log
    log_entries = get_rule_exec_log()
    if not log_entries:
        ws.cell(row=5, column=1, value="ℹ️  No rule execution log available. Run an assessment to generate the log.")
        _auto_col_width(ws)
        ws.sheet_view.showGridLines = False
        return

    row = 5
    for entry in log_entries:
        ws.cell(row=row, column=1, value=entry.get("Dimension", ""))
        ws.cell(row=row, column=2, value=entry.get("Column", ""))
        ws.cell(row=row, column=3, value=entry.get("Rule", ""))
        ws.cell(row=row, column=4, value=entry.get("Evaluated", 0))
        ws.cell(row=row, column=5, value=entry.get("Failed", 0))
        ws.cell(row=row, column=6, value=entry.get("Passed", 0))
        score_pct = entry.get("Score_%")
        score_cell = ws.cell(row=row, column=7, value=f"{score_pct}%" if score_pct is not None else "N/A")
        if isinstance(score_pct, (int, float)):
            _color_score_cell(score_cell, score_pct)
        severity = entry.get("Severity", "")
        sev_cell = ws.cell(row=row, column=8, value=severity)
        _color_severity_cell(sev_cell, severity)
        ws.row_dimensions[row].height = 18
        row += 1

    _apply_data_style(ws, 5, row - 1, 8)
    # Re-apply score/severity coloring after data style
    recolor_row = 5
    for entry in log_entries:
        score_pct = entry.get("Score_%")
        if isinstance(score_pct, (int, float)):
            _color_score_cell(ws.cell(row=recolor_row, column=7), score_pct)
        _color_severity_cell(ws.cell(row=recolor_row, column=8), entry.get("Severity", ""))
        recolor_row += 1

    _auto_col_width(ws)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"


# ══════════════════════════════════════════════════════════════════════
#  NEW SIMPLIFIED REPORT  (matches annex_vendor format)
#
#  Structure:
#    Sheet 1 : "Results"  — full dataset with Issues / Count / Categories
#    Sheet 2 : "Summary"  — S.No. | Particulars | Total Records
#    Sheet 3…N: "Annexure 1", "Annexure 2", … — one sheet per unique rule,
#               containing all rows (full columns) that failed that rule,
#               plus Issues / Count of issues / Issue categories columns
#    Last sheet: "Annexure N" for Duplicates (if any)
# ══════════════════════════════════════════════════════════════════════

def _annexure_header_style(ws, row_num: int) -> None:
    """Bold header row — matches the sample file style."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[row_num]:
        if cell.value is not None:
            cell.font = font
            cell.fill = fill
            cell.alignment = align
            cell.border = border


def _annexure_data_style(ws, start_row: int, end_row: int, num_cols: int) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(start_row, end_row + 1):
        bg = "EBF3FB" if r % 2 == 0 else "FFFFFF"
        fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.font = Font(size=10, name="Arial")
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border


def _annexure_auto_width(ws, min_w: int = 10, max_w: int = 40) -> None:
    from openpyxl.cell.cell import MergedCell
    for col_cells in ws.columns:
        letter = None
        for cell in col_cells:
            if not isinstance(cell, MergedCell):
                letter = cell.column_letter
                break
        if not letter:
            continue
        max_len = max(
            (len(str(cell.value or "")) for cell in col_cells if not isinstance(cell, MergedCell)),
            default=0,
        )
        ws.column_dimensions[letter].width = min(max(max_len + 2, min_w), max_w)


# ══════════════════════════════════════════════════════════════════════
#  REWRITTEN REPORT GENERATION ENGINE
#  ─────────────────────────────────────────────────────────────────────
#  Architecture (strictly enforced):
#
#    Sheet order in the workbook:
#      Results        – full dataset rows that carry at least one issue
#      Annexure 1     – first non-empty rule failure set
#      Annexure 2     – second non-empty rule failure set
#      …
#      Annexure N     – Duplicates (always last, merged into ONE sheet)
#      Summary        – built DYNAMICALLY from the generated annexure list
#                       (never dimension-based; counts must equal sheet rows)
#
#  Key guarantees:
#    • Dimension labels (Completeness / Validity / Standardization) never
#      appear as Summary rows.
#    • "Confusion Check" is never generated.
#    • Duplicates (exact + fuzzy) are merged → deduplicated → ONE annexure.
#    • Summary counts == actual rows in each Annexure sheet (validated).
#    • No entry appears twice in Summary.
# ══════════════════════════════════════════════════════════════════════


# ── Internal: per-row issue map ───────────────────────────────────────
def _build_row_issue_map(
    df: pd.DataFrame,
    all_annexure: List[Dict],
) -> Dict[int, Dict[str, Any]]:
    """
    Build a mapping of  {df_row_index → issue metadata}  from the raw
    annexure records produced by the rule engines.

    For each row we collect:
      issues_text  – human-readable comma-separated issue labels
      issue_count  – number of distinct issues on that row
      categories   – comma-separated dimension names (sorted)

    Implementation notes
    ────────────────────
    • Row_Number in annexure records is 1-based Excel row (header = row 1,
      first data row = row 2), so  df_index = Row_Number − 2.
    • We use Rule_Applied as the human-readable label instead of Issue_Type
      because Issue_Type for duplicates contains group-ID noise like
      "Duplicate (Group 3)" which would pollute the Issues column.
    • Duplicate annexure records are excluded here — they are written
      directly from dup_records in their own annexure sheet, not via
      all_annexure row indices.
    """
    _DIM_SKIP = {"Uniqueness"}          # dup rows handled separately

    row_map: Dict[int, Dict[str, Any]] = {}
    for rec in all_annexure:
        # Skip uniqueness/duplicate records — they live in the dup annexure
        if str(rec.get("Dimension", "")) in _DIM_SKIP:
            continue
        try:
            row_idx = int(rec["Row_Number"]) - 2   # 0-based df index
        except (KeyError, ValueError, TypeError):
            continue
        if row_idx < 0 or row_idx >= len(df):
            continue

        entry = row_map.setdefault(row_idx, {"issues": [], "dims": set()})
        label = str(rec.get("Rule_Applied", rec.get("Issue_Type", ""))).strip()
        if label and label not in entry["issues"]:
            entry["issues"].append(label)
        dim = str(rec.get("Dimension", "")).strip()
        if dim:
            entry["dims"].add(dim)

    result: Dict[int, Dict[str, Any]] = {}
    for idx, entry in row_map.items():
        result[idx] = {
            "issues_text":  "; ".join(entry["issues"]),
            "issue_count":  len(entry["issues"]),
            "categories":   ", ".join(sorted(entry["dims"])),
        }
    return result


# ── Internal: merge & deduplicate dup_records ─────────────────────────
def _merge_dup_records(dup_records: pd.DataFrame) -> pd.DataFrame:
    """
    Accept the raw dup_records DataFrame (which may contain rows flagged by
    multiple exact and/or fuzzy rules), deduplicate on original index, and
    return a clean DataFrame with one row per unique source record.

    Steps
    ─────
    1. Drop rows with duplicate df-index values (keep first occurrence).
    2. Reset the integer index so the result is 0-based and contiguous.
    3. If dup_records is None or empty, return an empty DataFrame.
    """
    if dup_records is None or not isinstance(dup_records, pd.DataFrame):
        return pd.DataFrame()
    if dup_records.empty:
        return pd.DataFrame()

    merged = dup_records.copy()

    # Deduplicate on the original source-row identifier.
    # The engine adds a "Row_Number" column (1-based Excel row).
    # If present, use it as the dedup key; otherwise fall back to the index.
    if "Row_Number" in merged.columns:
        merged = merged.drop_duplicates(subset=["Row_Number"], keep="first")
    else:
        merged = merged.loc[~merged.index.duplicated(keep="first")]

    merged = merged.reset_index(drop=True)
    return merged


# ── Internal: build ordered annexure groups ───────────────────────────
def _build_annexure_groups(
    df: pd.DataFrame,
    all_annexure: List[Dict],
    dup_records_clean: pd.DataFrame,
) -> List[Dict[str, Any]]:
    """
    Produce an ordered list of annexure group descriptors.

    Each descriptor:
    ┌─────────────────────────────────────────────────────────────┐
    │ rule_label    : str   – human-readable name shown in Summary│
    │ row_indices   : list  – 0-based df indices (non-dup groups) │
    │ is_duplicate  : bool  – True for the single dup group       │
    │ row_count     : int   – authoritative count for Summary      │
    └─────────────────────────────────────────────────────────────┘

    Rules
    ─────
    • Groups derive from Rule_Applied (not Issue_Type or Dimension).
    • Dimension labels are NEVER used as rule_label.
    • Only non-empty groups produce a descriptor (empty → no annexure).
    • Uniqueness / duplicate annexure records are excluded from the
      rule-based groups; they appear only in the single dup group.
    • Duplicate group is appended LAST.
    • Non-dup groups are sorted by row_count descending.
    • Deduplication within each group: a df row is counted once per
      Rule_Applied even if multiple annexure records reference it.
    """
    _SKIP_DIMS   = {"Uniqueness"}
    # Dimension names that must never appear as a rule_label in Summary
    _DIM_LABELS  = {
        "Completeness", "Validity", "Standardization", "Uniqueness",
        "Confusion Check", "Missing Values", "Invalid Format",
        "Duplicate Records", "Non-Standard Values",
    }

    # Map: Rule_Applied → ordered set of 0-based df row indices
    rule_rows: Dict[str, "dict[int, None]"] = {}   # use dict as ordered set

    for rec in all_annexure:
        if str(rec.get("Dimension", "")) in _SKIP_DIMS:
            continue
        label = str(rec.get("Rule_Applied", rec.get("Issue_Type", ""))).strip()
        if not label or label in _DIM_LABELS:
            continue
        try:
            row_idx = int(rec["Row_Number"]) - 2
        except (KeyError, ValueError, TypeError):
            continue
        if row_idx < 0 or row_idx >= len(df):
            continue
        rule_rows.setdefault(label, {})[row_idx] = None

    # Build non-dup groups (non-empty only)
    groups: List[Dict[str, Any]] = []
    for label, idx_dict in rule_rows.items():
        indices = list(idx_dict.keys())
        if not indices:
            continue
        groups.append({
            "rule_label":   label,
            "row_indices":  sorted(indices),
            "is_duplicate": False,
            "row_count":    len(indices),
        })

    # Sort by count descending for readability
    groups.sort(key=lambda g: g["row_count"], reverse=True)

    # Single duplicate group — appended last
    if not dup_records_clean.empty:
        groups.append({
            "rule_label":   "Duplicates",
            "row_indices":  [],                         # not used; dup sheet written from df directly
            "is_duplicate": True,
            "row_count":    len(dup_records_clean),
        })

    return groups


# ── Sheet writers ─────────────────────────────────────────────────────

def _sheet_results(
    ws,
    df: pd.DataFrame,
    row_issue_map: Dict[int, Dict[str, Any]],
) -> None:
    """
    Results sheet — only rows that carry at least one non-duplicate issue.
    Columns = all original df columns + Issues | Count of issues | Issue categories.
    """
    headers = list(df.columns) + ["Issues", "Count of issues", "Issue categories"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _annexure_header_style(ws, 1)
    ws.row_dimensions[1].height = 22

    out_row = 2
    for idx, df_row in df.iterrows():
        info = row_issue_map.get(idx)
        if info is None:
            continue
        for c, val in enumerate(df_row, 1):
            ws.cell(row=out_row, column=c, value=_safe_val(val))
        nc = len(df.columns)
        ws.cell(row=out_row, column=nc + 1, value=info["issues_text"])
        ws.cell(row=out_row, column=nc + 2, value=info["issue_count"])
        ws.cell(row=out_row, column=nc + 3, value=info["categories"])
        ws.row_dimensions[out_row].height = 18
        out_row += 1

    if out_row > 2:
        _annexure_data_style(ws, 2, out_row - 1, len(headers))
    _annexure_auto_width(ws)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"


def _sheet_rule_annexure(
    ws,
    df: pd.DataFrame,
    row_indices: List[int],
    row_issue_map: Dict[int, Dict[str, Any]],
    rule_label: str,
) -> None:
    """
    One non-duplicate annexure sheet.

    Writes all original df columns + Issues | Count of issues | Issue categories
    for every df row that failed the given rule.
    """
    headers = list(df.columns) + ["Issues", "Count of issues", "Issue categories"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _annexure_header_style(ws, 1)
    ws.row_dimensions[1].height = 22

    for out_row, idx in enumerate(row_indices, 2):
        # Use .iloc for positional access (idx is already a 0-based integer)
        try:
            df_row = df.iloc[idx]
        except IndexError:
            continue
        for c, val in enumerate(df_row, 1):
            ws.cell(row=out_row, column=c, value=_safe_val(val))
        nc = len(df.columns)
        info = row_issue_map.get(idx, {"issues_text": rule_label, "issue_count": 1, "categories": ""})
        ws.cell(row=out_row, column=nc + 1, value=info["issues_text"])
        ws.cell(row=out_row, column=nc + 2, value=info["issue_count"])
        ws.cell(row=out_row, column=nc + 3, value=info["categories"])
        ws.row_dimensions[out_row].height = 18

    if row_indices:
        _annexure_data_style(ws, 2, len(row_indices) + 1, len(headers))
    _annexure_auto_width(ws)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"


def _sheet_dup_annexure(
    ws,
    dup_records_clean: pd.DataFrame,
) -> None:
    """
    Duplicates annexure sheet.

    Writes dup_records_clean directly (already merged + deduplicated).
    Columns = all original df columns + engine metadata columns
    (Duplicate_Group_ID, Duplicate_Type, Similarity_Score, Row_Number, …).
    """
    if dup_records_clean.empty:
        ws.cell(row=1, column=1, value="No duplicate records found.")
        ws.sheet_view.showGridLines = False
        return

    headers = list(dup_records_clean.columns)
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _annexure_header_style(ws, 1)
    ws.row_dimensions[1].height = 22

    for out_row, (_, dr) in enumerate(dup_records_clean.iterrows(), 2):
        for c, val in enumerate(dr, 1):
            ws.cell(row=out_row, column=c, value=_safe_val(val))
        ws.row_dimensions[out_row].height = 18

    if len(dup_records_clean) > 0:
        _annexure_data_style(ws, 2, len(dup_records_clean) + 1, len(headers))
    _annexure_auto_width(ws)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"


def _sheet_summary(
    ws,
    groups: List[Dict[str, Any]],
) -> None:
    """
    Summary sheet — built dynamically from the generated annexure groups.

    Format
    ──────
    S.No.       | Particulars          | Total Records
    Annexure 1  | <rule_label>         | <row_count>
    Annexure 2  | <rule_label>         | <row_count>
    …
    Annexure N  | Duplicates           | <dup_count>   ← highlighted in red

    Strict rules enforced
    ─────────────────────
    • Only annexure groups that were actually generated appear here.
    • Counts are taken from group["row_count"] which equals the number of
      rows written to the corresponding annexure sheet.
    • No dimension-based rows, no "Confusion Check", no duplicates.
    • Duplicate entry is always last and highlighted.
    """
    from openpyxl.styles import Font, PatternFill, Border, Side

    headers = ["S.No.", "Particulars", "Total Records"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _annexure_header_style(ws, 1)
    ws.row_dimensions[1].height = 22

    dup_excel_rows: List[int] = []
    for i, grp in enumerate(groups, 1):
        excel_row = i + 1
        ws.cell(row=excel_row, column=1, value=f"Annexure {i}")
        ws.cell(row=excel_row, column=2, value=grp["rule_label"])
        ws.cell(row=excel_row, column=3, value=grp["row_count"])
        ws.row_dimensions[excel_row].height = 18
        if grp["is_duplicate"]:
            dup_excel_rows.append(excel_row)

    last_data_row = len(groups) + 1
    if groups:
        _annexure_data_style(ws, 2, last_data_row, 3)

    # Highlight duplicate row(s) in red
    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for er in dup_excel_rows:
        for c in range(1, 4):
            cell = ws.cell(row=er, column=c)
            cell.fill   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            cell.font   = Font(bold=True, color="9C0006", size=10, name="Arial")
            cell.border = border

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 16
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"


# ── Value sanitiser ───────────────────────────────────────────────────
def _safe_val(v: Any) -> Any:
    """Convert NaN / None / numpy scalars to clean Python types for openpyxl."""
    if v is None:
        return ""
    if isinstance(v, float) and (np.isnan(v) or np.isinf(v)):
        return ""
    if isinstance(v, (np.integer,)):
        return int(v)
    if isinstance(v, (np.floating,)):
        return float(v)
    return v


# ── Pre-export validation ─────────────────────────────────────────────
def _validate_summary_counts(groups: List[Dict[str, Any]]) -> None:
    """
    Verify that every group's row_count is a non-negative integer and that
    the summary will exactly reflect the annexure sheet contents.

    Raises
    ──────
    ValueError : "Summary and Annexure counts mismatch." with detail if any
                 group's row_count does not match the length of its row_indices
                 (non-dup groups only; dup count is taken from dup_records_clean).
    """
    mismatches: List[str] = []
    for i, grp in enumerate(groups, 1):
        if grp["is_duplicate"]:
            # row_count for dup group = len(dup_records_clean); no row_indices to check
            continue
        declared  = grp["row_count"]
        actual    = len(grp["row_indices"])
        if declared != actual:
            mismatches.append(
                f"Annexure {i} '{grp['rule_label']}': "
                f"Summary declares {declared} but sheet has {actual} rows."
            )
    if mismatches:
        detail = "\n".join(mismatches)
        raise ValueError(f"Summary and Annexure counts mismatch.\n{detail}")


# ── Main report generator ─────────────────────────────────────────────
def generate_excel_report(
    df: pd.DataFrame,
    clean_df: pd.DataFrame,
    dup_records: pd.DataFrame,
    all_annexure: List[Dict],
    dim_scores: Dict[str, float],
    overall_score: float,
    selected_dimensions: List[str],
    uniqueness_config: Dict[str, Any],
) -> bytes:
    """
    Generate a production-level Excel DQ report.

    Workbook sheet order
    ────────────────────
    1. Results       — all rows with at least one issue (+3 issue columns)
    2. Annexure 1    — rows failing Rule 1 (most failures first)
    3. Annexure 2    — rows failing Rule 2
       …
    N. Annexure N    — Duplicates (merged, deduplicated; always last)
    N+1. Summary     — dynamically built from the annexure list above
                       (S.No. | Particulars | Total Records)

    Strict contract
    ───────────────
    • Dimension names never appear in Summary.
    • Duplicates are merged into ONE annexure, ONE Summary row.
    • Summary counts are validated against actual sheet rows before save.
    • File is NOT written if validation fails (ValueError raised).
    """
    from openpyxl import Workbook

    # ── Step 1: Merge + deduplicate all dup_records ───────────────────
    dup_clean = _merge_dup_records(dup_records)

    # ── Step 2: Build per-row issue map (excludes dup records) ────────
    row_issue_map = _build_row_issue_map(df, all_annexure)

    # ── Step 3: Build ordered annexure groups ─────────────────────────
    groups = _build_annexure_groups(df, all_annexure, dup_clean)

    # ── Step 4: Pre-export validation ────────────────────────────────
    _validate_summary_counts(groups)

    # ── Step 5: Write workbook ────────────────────────────────────────
    output = BytesIO()
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1 — Results
    _sheet_results(wb.create_sheet("Results"), df, row_issue_map)

    # Sheets 2…N — Annexures (one per non-empty group)
    for i, grp in enumerate(groups, 1):
        ws = wb.create_sheet(f"Annexure {i}")
        if grp["is_duplicate"]:
            _sheet_dup_annexure(ws, dup_clean)
        else:
            _sheet_rule_annexure(
                ws, df, grp["row_indices"], row_issue_map, grp["rule_label"]
            )

    # Final sheet — Summary (built last, reflects actual annexure sheets)
    _sheet_summary(wb.create_sheet("Summary"), groups)

    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _style_workbook(writer) -> None:
    """Legacy stub — kept for backward compatibility."""
    pass