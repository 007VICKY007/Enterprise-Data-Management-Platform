"""
app.py — Enterprise Data Management Platform
Fixed: data editor / dataframe table width issues
"""

import traceback
import datetime
import zipfile
from io import BytesIO
from pathlib import Path
from typing import List

import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.patches import Wedge

st.set_page_config(
    page_title="Enterprise Data Management Platform",
    page_icon="🏛️",
    layout="wide",
)


from modules.dq import (
    APP_TITLE, APP_ICON, SUPPORTED_FORMATS, DIMENSIONS,
    load_dataset, get_excel_sheet_names,
    execute_completeness_rules, execute_validity_rules,
    execute_uniqueness_rules, execute_standardization_rules,
    compute_completeness_score, compute_validity_score,
    compute_uniqueness_score, compute_standardization_score,
    compute_overall_score, build_clean_dataset, generate_excel_report,
)
from modules.config        import AppConfig
from modules.ui_components import UIComponents

def _lottie_upload_fixed(caption="Upload both files above to begin"):
    st.markdown(f'<div class="lottie-slot"><div class="lottie-frame lottie-upload-fallback"></div><p class="lottie-caption">{caption}</p></div>', unsafe_allow_html=True)
def _arrow_down_fixed():
    st.markdown('<div class="guidance-arrow-down">⬇</div>', unsafe_allow_html=True)
def _upload_hint_fixed(kind="dataset"):
    label = "📂 Master Dataset" if kind=="dataset" else "📜 Rules / Rulebook"
    tip   = "CSV, Excel, JSON, Parquet, ODS, XML" if kind=="dataset" else "CSV/Excel: column_name·rule·dimension·message — or JSON"
    st.markdown(f'<p style="font-size:0.82rem;color:#64748b;margin-bottom:0.3rem;">{label} · {tip}</p>', unsafe_allow_html=True)
def _results_header_fixed(score: float):
    if score>=80:   cls,em,lbl="dq-score-excellent","🏆","Excellent"
    elif score>=60: cls,em,lbl="dq-score-good","✅","Good"
    elif score>=40: cls,em,lbl="dq-score-fair","⚠️","Fair"
    else:           cls,em,lbl="dq-score-poor","❌","Poor"
    st.markdown(f'<div class="{cls}"><h2 style="margin:0;">{em} {lbl} — {score:.1f}%</h2></div>', unsafe_allow_html=True)

UIComponents.render_lottie_upload  = _lottie_upload_fixed
UIComponents.render_arrow_down     = _arrow_down_fixed
UIComponents.render_upload_hint    = _upload_hint_fixed
UIComponents.render_results_header = _results_header_fixed

from DataMaturity.config import (
    UNIQU_PURPLE, UNIQU_MAGENTA, UNIQU_LAVENDER, UNIQU_LIGHT_BG, UNIQU_TEXT, UNIQU_GREY,
    RATING_LABELS, RATING_TO_SCORE, DEFAULT_MASTER_OBJECTS, MATURITY_DIMS, QUESTION_BANK,
)
from DataMaturity.helpers import (
    dq_score_to_maturity_level, init_maturity_state, build_question_df,
    sync_response_tables, autofill_dq_dimension, compute_all_scores,
    validate_responses, to_excel_bytes,
)
from DataMaturity.visualizations   import render_slide_png
from DataMaturity.report_generator import build_pdf_bytes


# ══════════════════════════════════════════════════════════════════════════════
#  GLOBAL STYLES
#  KEY FIX: removed overflow-x:hidden on stDataFrame (was crushing table width)
#           added proper width rules for stDataEditor and stTabPanel
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:ital,wght@0,300;0,400;0,500;0,600;0,700;0,800;1,400;1,600&display=swap');
/* Primary font: Plus Jakarta Sans — clean, modern, clearly legible at all sizes */
:root {
    --font-primary: 'Plus Jakarta Sans', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
}
html, body, [class*="css"], .stMarkdown, .stText, button, input, select, textarea,
[data-testid="stSidebar"], [data-testid="stMain"], p, td, th, li, label, span {
    font-family: var(--font-primary) !important;
    -webkit-font-smoothing: antialiased !important;
    -moz-osx-font-smoothing: grayscale !important;
}
h1,h2,h3,h4,h5,h6 {
    font-family: var(--font-primary) !important;
    font-weight: 700 !important;
    letter-spacing: -0.01em !important;
}
h1 a,h2 a,h3 a,h4 a,h5 a,h6 a,
[data-testid="stMarkdownContainer"] h1 a,[data-testid="stMarkdownContainer"] h2 a,
[data-testid="stMarkdownContainer"] h3 a,[data-testid="stMarkdownContainer"] h4 a,
[data-testid="stMarkdownContainer"] h5 a,[data-testid="stMarkdownContainer"] h6 a {
    display: none !important; pointer-events: none !important;
}

/* ═══════════════════════════════════════════════════════════
   FULL WIDTH LAYOUT — remove all side padding/max-width
   ═══════════════════════════════════════════════════════════ */
.block-container, [data-testid="block-container"],
section[data-testid="stMain"] > div,
div[data-testid="stAppViewBlockContainer"] {
    max-width: 100% !important;
    padding-left: 1.5rem !important;
    padding-right: 1.5rem !important;
    padding-top: 0 !important;
    margin-top: 0 !important;
}
/* Kill Streamlit's internal top spacing */
[data-testid="stAppViewContainer"] {
    padding-top: 0 !important;
    margin-top: 0 !important;
}
[data-testid="stAppViewContainer"] > section > div:first-child {
    padding-top: 0 !important;
    margin-top: 0 !important;
}
/* Remove the gap Streamlit adds above first element */
.main > div:first-child {
    padding-top: 0 !important;
    margin-top: 0 !important;
}
div[data-testid="stVerticalBlock"] > div:first-child {
    margin-top: 0 !important;
    padding-top: 0 !important;
}
/* Remove sidebar-induced left offset when sidebar is collapsed */
[data-testid="stMain"] {
    padding-left: 0 !important;
}
/* Ensure all columns stretch fully */
[data-testid="stHorizontalBlock"] {
    width: 100% !important;
    gap: 0.75rem !important;
}


/* ═══════════════════════════════════════════════════════════
   TABLE & DATA EDITOR WIDTH FIX
   ROOT CAUSE FIXED:
   - Removed overflow-x: hidden (was clipping table content)
   - Added min-width: 0 so flex children can shrink properly
   - overflow-x: auto allows horizontal scroll when needed
   - Tab panels must not constrain child width
   ═══════════════════════════════════════════════════════════ */

/* stDataFrame — the standard read-only table */
[data-testid="stDataFrame"] {
    width: 100% !important;
    min-width: 0 !important;
    overflow: visible !important;
}
[data-testid="stDataFrame"] > div {
    width: 100% !important;
    min-width: 0 !important;
    overflow-x: auto !important;   /* scroll, never clip */
    overflow-y: visible !important;
}
[data-testid="stDataFrame"] iframe {
    width: 100% !important;
    min-width: 0 !important;
}

/* stDataEditor — the editable glide-data-grid table */
[data-testid="stDataEditor"] {
    width: 100% !important;
    min-width: 0 !important;
    display: block !important;
    overflow: visible !important;
}
[data-testid="stDataEditor"] > div {
    width: 100% !important;
    min-width: 0 !important;
    overflow: visible !important;
}
[data-testid="stDataEditor"] iframe {
    width: 100% !important;
    min-width: 0 !important;
}
/* Glide data grid canvas stretches to fill parent */
[data-testid="stDataEditor"] canvas {
    max-width: 100% !important;
    display: block !important;
}
/* Remove ALL extra/overlay scrollbars — only the glide grid's own scrollbar should show */
[data-testid="stDataEditor"] > div > div {
    overflow: hidden !important;
}
[data-testid="stDataEditor"] > div::-webkit-scrollbar { display: none !important; }
[data-testid="stDataEditor"] > div { -ms-overflow-style: none !important; scrollbar-width: none !important; }
[data-testid="stDataEditor"] > div > div::-webkit-scrollbar { display: none !important; }
[data-testid="stDataEditor"] > div > div { -ms-overflow-style: none !important; scrollbar-width: none !important; }
/* Let only the innermost glide-data-grid scroller work */
[data-testid="stDataEditor"] .dvn-scroller {
    overflow-x: auto !important;
    overflow-y: auto !important;
}

/* Tab panels: must NOT clip or constrain child elements */
[data-testid="stTabPanel"] {
    width: 100% !important;
    overflow: visible !important;
    padding-left: 0.25rem !important;
    padding-right: 0.25rem !important;
    box-sizing: border-box !important;
}
[data-testid="stTabs"],
[data-testid="stTabs"] > div {
    width: 100% !important;
}

/* Columns inside tabs must not collapse child width */
[data-testid="stTabPanel"] [data-testid="stHorizontalBlock"] > [data-testid="stColumn"] > div {
    min-width: 0 !important;
    width: 100% !important;
}
/* ═══════════════════════════════════════════════════════════ */

.mat-table-panel { background:#ffffff; border:1.5px solid #d9cef0; border-radius:16px;
    padding:1.2rem 1.4rem 1rem; margin-bottom:1.2rem; box-shadow:0 2px 12px rgba(91,45,144,0.07); }
.mat-table-title { font-size:1rem; font-weight:700; color:#3b1f72; margin-bottom:0.75rem;
    padding-bottom:0.5rem; border-bottom:2px solid #ede8f7; }
.dl-card { background:#ffffff; border:1.5px solid #d9cef0; border-radius:16px;
    padding:1rem 1.1rem 0.8rem; margin-bottom:0.6rem; min-height:120px;
    box-shadow:0 2px 8px rgba(91,45,144,0.07); }
.dl-card-icon { font-size:1.6rem; margin-bottom:0.3rem; }
.dl-card-title { font-size:0.9rem; font-weight:700; color:#3b1f72; margin-bottom:0.3rem; }
.dl-card-desc { font-size:0.76rem; color:#7a7a9a; line-height:1.4; }
.mat-edit-btn-wrap .stButton > button {
    background:linear-gradient(135deg,#5b2d90,#7c4dbb) !important; color:white !important;
    font-weight:700 !important; font-size:0.82rem !important; padding:0.45rem 1.2rem !important;
    border-radius:999px !important; border:none !important;
    box-shadow:0 4px 16px rgba(91,45,144,0.35) !important; }
.dim-header { background:#f5f0fc; border-left:4px solid #5b2d90;
    padding:0.6rem 1rem; border-radius:0 8px 8px 0;
    font-weight:700; color:#5b2d90; margin:0.5rem 0; }
.score-card { background:#ffffff; border:1.5px solid #d9cef0; border-radius:12px;
    padding:1rem 1.25rem; text-align:center; }
.score-card .val { font-size:1.8rem; font-weight:800; color:#5b2d90; }
.score-card .lbl { font-size:0.78rem; color:#7a7a9a; font-weight:600;
    text-transform:uppercase; letter-spacing:0.04em; }
.score-excellent { color:#059669; } .score-good { color:#5b2d90; }
.score-fair { color:#d97706; } .score-poor { color:#dc2626; }
</style>
""", unsafe_allow_html=True)

# Light-theme dropdown fix
st.markdown("""<style>
div[data-baseweb="popover"],div[data-baseweb="popover"]>div,div[data-baseweb="menu"]{background:#ffffff!important;border:1px solid #d9cef0!important;}
div[data-baseweb="popover"] ul,div[data-baseweb="menu"] ul{background:#ffffff!important;}
div[data-baseweb="popover"] li *,div[data-baseweb="menu"] li *{color:#1a1a2e!important;-webkit-text-fill-color:#1a1a2e!important;}
div[data-baseweb="popover"] [role="option"]{background:#f9f8fc!important;color:#1a1a2e!important;font-weight:600!important;}
div[data-baseweb="popover"] [role="option"]:hover{background:#ede8f7!important;}
div[data-baseweb="popover"] [aria-selected="true"]{background:rgba(124,58,237,0.15)!important;color:#3b1f72!important;}
</style>""", unsafe_allow_html=True)


st.markdown("""<style>
@import url('https://fonts.googleapis.com/css2?family=Material+Symbols+Rounded:opsz,wght,FILL,GRAD@24,400,0,0&display=block');
.material-symbols-rounded,[class*="material-symbols"],button span,
[data-testid="stSidebarCollapsedControl"] span,
[data-testid="stSidebarNavCollapseButton"] span,
section[data-testid="stSidebar"] button span,
div[data-testid="stFileUploaderDropzoneIcon"] span,
[data-testid="stExpanderToggleIcon"] span,
button[data-testid="stBaseButton-minimal"] span,
details > summary span,
details summary span {
    font-family:'Material Symbols Rounded' !important;
    font-variation-settings:'FILL' 0,'wght' 400,'GRAD' 0,'opsz' 24 !important;
    font-size:22px !important;line-height:1 !important;letter-spacing:normal !important;
    text-transform:none !important;display:inline-block !important;white-space:nowrap !important;
    -webkit-font-feature-settings:'liga' !important;font-feature-settings:'liga' !important;
    -webkit-font-smoothing:antialiased !important;font-style:normal !important;font-weight:normal !important;
}
[data-testid="stSidebarCollapsedControl"]{
    background:linear-gradient(135deg,#5b2d90 0%,#7c4dbb 100%) !important;
    border:none !important;border-radius:0 10px 10px 0 !important;
    width:36px !important;height:36px !important;padding:0 !important;
    display:flex !important;align-items:center !important;justify-content:center !important;
    box-shadow:2px 0 12px rgba(91,45,144,0.28) !important;cursor:pointer !important;
}
[data-testid="stSidebarCollapsedControl"]:hover{width:42px !important;background:linear-gradient(135deg,#4a2278 0%,#6d3eab 100%) !important;}
[data-testid="stSidebarCollapsedControl"] span{color:#ffffff !important;-webkit-text-fill-color:#ffffff !important;}
button[data-testid="stBaseButton-headerNoPadding"] span,
button[data-testid="stSidebarNavCollapseButton"] span{color:#5b2d90 !important;-webkit-text-fill-color:#5b2d90 !important;}
div[data-testid="stFileUploaderDropzoneIcon"] span{color:#5b2d90 !important;-webkit-text-fill-color:#5b2d90 !important;font-size:28px !important;}
[data-testid="stExpanderToggleIcon"] span,
details > summary span,
details summary span{color:#5b2d90 !important;-webkit-text-fill-color:#5b2d90 !important;font-size:20px !important;}
</style>""", unsafe_allow_html=True)




st.markdown("""
<style>
/* ── DQ Steps Row (01 / 02 / 03) ── */
.dq-steps-row {
    display: flex;
    align-items: center;
    justify-content: center;
    gap: 0;
    margin: 1.5rem 0 0.5rem;
    flex-wrap: nowrap;
}
.dq-step {
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    background: #ffffff;
    border: 2px solid #d9cef0;
    border-radius: 16px;
    padding: 1.1rem 1.6rem;
    min-width: 160px;
    box-shadow: 0 2px 12px rgba(91,45,144,0.08);
    transition: box-shadow 0.2s, border-color 0.2s;
    position: relative;
}
.dq-step:hover {
    box-shadow: 0 6px 24px rgba(91,45,144,0.16);
    border-color: #b39ddb;
}
.dq-step.active {
    border-color: #7c3aed;
    background: linear-gradient(135deg, #f5f0ff 0%, #fdf2f8 100%);
    box-shadow: 0 4px 20px rgba(124,58,237,0.18);
}
.dq-step-num {
    font-size: 0.72rem;
    font-weight: 800;
    letter-spacing: 0.12em;
    color: #a78bcc;
    margin-bottom: 0.35rem;
    text-transform: uppercase;
}
.dq-step.active .dq-step-num {
    color: #7c3aed;
}
.dq-step-icon {
    font-size: 1.8rem;
    margin-bottom: 0.4rem;
    line-height: 1;
}
.dq-step-title {
    font-size: 0.92rem;
    font-weight: 700;
    color: #2d1b58;
    margin-bottom: 0.15rem;
    text-align: center;
}
.dq-step-desc {
    font-size: 0.76rem;
    color: #7a7a9a;
    text-align: center;
}
.dq-step-line {
    flex: 1;
    height: 2px;
    background: linear-gradient(90deg, #d9cef0, #b39ddb, #d9cef0);
    min-width: 32px;
    max-width: 72px;
    border-radius: 2px;
}
</style>
""", unsafe_allow_html=True)

_GDG_LIGHT_STYLE = """<style>
:root,[data-testid="stDataEditor"],[data-testid="stDataEditor"]>div{
    --gdg-bg-cell:#ffffff!important;--gdg-bg-cell-medium:#f7f4fc!important;
    --gdg-bg-header:#ede8f7!important;--gdg-bg-header-has-focus:#e0d9f2!important;
    --gdg-bg-header-hovered:#d4cced!important;--gdg-border-color:#e8e2f5!important;
    --gdg-accent-color:#7c3aed!important;--gdg-text-dark:#1a1028!important;
    --gdg-text-header:#3b1f72!important;--gdg-cell-text-color:#1a1028!important;
    --gdg-header-font-style:700 13px 'Plus Jakarta Sans',sans-serif!important;
    --gdg-base-font-style:500 13px 'Plus Jakarta Sans',sans-serif!important;}
[data-testid="stDataEditor"] canvas{background-color:#ffffff!important;}
[data-testid="stDataEditor"] input{background:#ffffff!important;color:#1a1028!important;border:2px solid #7c3aed!important;border-radius:5px!important;}
</style>"""
def inject_gdg_light(): st.markdown(_GDG_LIGHT_STYLE, unsafe_allow_html=True)


def load_css():
    try:
        with open("assets/styles.css", encoding="utf-8") as f:
            st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)
    except FileNotFoundError:
        pass


def _init_state():
    if "page" not in st.session_state: st.session_state["page"] = "home"
    for k,v in {"dq_score":None,"dq_dim_scores":None,"dq_results_df":None,
                "dq_object_name":"Customer","dq_excel_bytes":None}.items():
        if k not in st.session_state: st.session_state[k] = v
    # Dynamic criteria builder state
    for k,v in {"dq_rule_entries":[],"dq_dataset_type":"Customer","dq_cb_search":""}.items():
        if k not in st.session_state: st.session_state[k] = v
    _mat = {"mat_dims":list(MATURITY_DIMS),"mat_objects":list(DEFAULT_MASTER_OBJECTS),
            "mat_responses":{},"mat_submitted":False,"mat_payload":{},
            "mat_client_name":"","mat_benchmark":3.0,"mat_target":3.0,
            "mat_low_thr":2.0,"dq_autofilled":False,"mat_masters_applicable":True}
    for k,v in _mat.items():
        if k not in st.session_state: st.session_state[k] = v
    try: init_maturity_state()
    except: pass
    if "mat_dims" not in st.session_state or not st.session_state["mat_dims"]:
        st.session_state["mat_dims"] = list(MATURITY_DIMS)
    if "policies" not in st.session_state: st.session_state["policies"] = []


def get_timestamp_filename(prefix, ext):
    return f"{prefix}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.{ext}"

def _build_zip(files: dict) -> bytes:
    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for fn, data in files.items(): zf.writestr(fn, data)
    return buf.getvalue()

def _score_cls(score):
    if score>=80: return "score-excellent"
    if score>=60: return "score-good"
    if score>=40: return "score-fair"
    return "score-poor"

def _page_banner(icon, badge_text, title, subtitle, gradient="135deg,#3d1d63 0%,#5b2d90 55%,#b10f74 100%"):
    st.markdown(f"""
    <div style="background:linear-gradient({gradient});
         padding:2rem 2.5rem 1.8rem;border-radius:16px;margin-bottom:1rem;color:#fff;
         box-shadow:0 6px 28px rgba(91,45,144,0.22);">
        <div style="display:flex;align-items:center;gap:1rem;margin-bottom:0.75rem;">
            <span style="font-size:2.2rem;filter:drop-shadow(0 2px 8px rgba(0,0,0,0.18));">{icon}</span>
            <div>
                <div style="font-size:0.7rem;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;
                     color:rgba(255,255,255,0.6);margin-bottom:0.25rem;">{badge_text}</div>
                <h1 style="margin:0;font-size:1.95rem;font-weight:800;color:#fff;line-height:1.1;">{title}</h1>
            </div>
        </div>
        <p style="font-size:0.97rem;color:rgba(255,255,255,0.82);margin:0;line-height:1.6;">{subtitle}</p>
    </div>""", unsafe_allow_html=True)

def _page_nav(current_page):
    """Renders a styled horizontal pill navigation bar below the banner."""
    pages = [
        ("home",     "🏠 Home"),
        ("maturity", "📈 Maturity"),
        ("dq",       "🔍 Data Quality"),
        ("policy",   "📋 Policies"),
        ("case",     "📁 Case Mgmt"),
    ]

    # CSS to style the nav row buttons as pills
    st.markdown(f"""
    <style>
    div[data-testid="stHorizontalBlock"]:has(> div > div > div[data-testid="stButton"].nav-btn-wrap) {{
        background: #f5f0fc;
        border: 1.5px solid #d9cef0;
        border-radius: 14px;
        padding: 0.45rem 0.6rem;
        margin-bottom: 1.5rem;
        margin-top: -0.5rem;
        gap: 0.4rem !important;
        box-shadow: 0 2px 8px rgba(91,45,144,0.07);
    }}
    </style>
    """, unsafe_allow_html=True)

    cols = st.columns(len(pages))
    for i, (key, label) in enumerate(pages):
        with cols[i]:
            is_active = key == current_page
            if is_active:
                # Active pill — styled differently, not clickable
                st.markdown(
                    f'<div style="text-align:center;padding:0.42rem 0.5rem;border-radius:999px;'
                    f'background:linear-gradient(135deg,#5b2d90,#b10f74);color:#fff;'
                    f'font-size:0.84rem;font-weight:700;box-shadow:0 3px 12px rgba(91,45,144,0.35);'
                    f'cursor:default;">{label}</div>',
                    unsafe_allow_html=True
                )
            else:
                if st.button(label, key=f"navpill_{current_page}_{key}", use_container_width=True):
                    st.session_state["page"] = key
                    st.rerun()

    # Style the inactive nav buttons as pills
    st.markdown("""
    <style>
    /* Target buttons inside the nav row */
    div[data-testid="stHorizontalBlock"]:has(div[data-testid="stButton"]) button[kind="secondary"] {
        background: transparent !important;
        border: 1.5px solid #d9cef0 !important;
        color: #5b2d90 !important;
        border-radius: 999px !important;
        font-size: 0.84rem !important;
        font-weight: 600 !important;
        padding: 0.38rem 0.5rem !important;
        box-shadow: none !important;
        transition: background 0.18s, border-color 0.18s !important;
    }
    div[data-testid="stHorizontalBlock"]:has(div[data-testid="stButton"]) button[kind="secondary"]:hover {
        background: #ede8f7 !important;
        border-color: #b09dd6 !important;
        color: #3b1f72 !important;
    }
    </style>
    """, unsafe_allow_html=True)

def _svg_gauge(score, label):
    r=45; circ=3.14159*r; dash=score/100*circ; gap=circ-dash
    col="#5b2d90" if score>=80 else "#b10f74" if score>=60 else "#d97706" if score>=40 else "#dc2626"
    return f"""<div style="text-align:center;">
        <svg width="120" height="70" viewBox="0 0 120 70" xmlns="http://www.w3.org/2000/svg">
            <path d="M 15 55 A {r} {r} 0 0 1 105 55" fill="none" stroke="#e9e4f5" stroke-width="10" stroke-linecap="round"/>
            <path d="M 15 55 A {r} {r} 0 0 1 105 55" fill="none" stroke="{col}" stroke-width="10" stroke-linecap="round" stroke-dasharray="{dash:.1f} {gap:.1f}"/>
            <text x="60" y="50" text-anchor="middle" font-size="15" font-weight="800" fill="{col}" font-family="sans-serif">{score:.0f}%</text>
        </svg>
        <div style="font-size:0.72rem;font-weight:600;color:#7a7a9a;text-transform:uppercase;letter-spacing:0.05em;margin-top:-6px;">{label}</div>
    </div>"""

# Enhancement 4: Maturity scale colors for bar graph
_MATURITY_BAR_COLORS_MAP = {1:"#64748b", 2:"#b45309", 3:"#1d4ed8", 4:"#5b2d90", 5:"#0f766e"}
def _maturity_level_color(score):
    level=max(1,min(5,int(round(float(score)))))
    return _MATURITY_BAR_COLORS_MAP.get(level,"#5b2d90")

def _mat_bar_png(dim_vals):
    if not dim_vals: return None
    dims=list(dim_vals.keys()); scores=[dim_vals[d] for d in dims]
    # Enhancement 4: each bar gets color based on its maturity level
    cols=[_maturity_level_color(s) for s in scores]
    fig,ax=plt.subplots(figsize=(10,max(3,len(dims)*0.9)),dpi=140)
    fig.patch.set_facecolor('#f5f0fc'); ax.set_facecolor('#f9f8fc')
    # Enhancement 1: Aptos font in bar chart
    plt.rcParams.update({"font.family":"DejaVu Sans"})
    bars=ax.barh(dims,scores,color=cols,height=0.6,edgecolor="white",linewidth=2)
    ax.set_xlim(0,6.0); ax.set_xlabel("Maturity Score (1=Adhoc → 5=Optimised)",color="#3d1d63",fontsize=11,weight=600)
    ax.axvline(3.0,color="#38bdf8",lw=1.5,ls="--",alpha=0.7,label="Defined (3)")
    ax.axvline(4.0,color="#0284c7",lw=1.5,ls="--",alpha=0.7,label="Managed (4)")
    # Enhancement 4: add maturity scale legend
    from matplotlib.patches import Patch
    legend_handles=[Patch(facecolor=c,label=l) for c,l in zip(
        ["#64748b","#b45309","#1d4ed8","#5b2d90","#0f766e"],
        ["1-Adhoc","2-Repeatable","3-Defined","4-Managed","5-Optimised"]
    )]
    ax.legend(handles=legend_handles,fontsize=8,loc="lower right",frameon=True,title="Maturity Scale",title_fontsize=8)
    ax.tick_params(colors="#0c4a6e",labelsize=10)
    ax.spines[["top","right","bottom"]].set_visible(False)
    ax.spines["left"].set_color("#c4b0e0"); ax.spines["left"].set_linewidth(2)
    for bar,sc in zip(bars,scores):
        ax.text(bar.get_width()+0.1,bar.get_y()+bar.get_height()/2,f"{sc:.2f}",va="center",fontsize=11,fontweight="bold",color="#3d1d63")
    plt.tight_layout()
    buf=BytesIO(); fig.savefig(buf,format="png",bbox_inches="tight",facecolor='#f5f0fc'); plt.close(fig); return buf.getvalue()




# ══════════════════════════════════════════════════════════════════════════════
#  PAGE: HOME
# ══════════════════════════════════════════════════════════════════════════════
def page_home():
    st.markdown("""
    <div style="background:linear-gradient(135deg,#3d1d63 0%,#5b2d90 60%,#b10f74 100%);
         padding:2rem 2.5rem 1.8rem;border-radius:16px;margin-bottom:1rem;color:#fff;
         box-shadow:0 6px 28px rgba(91,45,144,0.22);">
        <div style="display:flex;align-items:center;gap:0.9rem;margin-bottom:0.8rem;">
            <span style="font-size:2.4rem;">🏛️</span>
            <div>
                <div style="font-size:0.72rem;font-weight:700;text-transform:uppercase;letter-spacing:0.12em;
                     color:rgba(255,255,255,0.65);margin-bottom:0.2rem;"></div>
                <h1 style="margin:0;font-size:2.1rem;font-weight:800;color:#fff;line-height:1.1;">
                    Enterprise Data Management Platform</h1>
            </div>
        </div>
        <p style="font-size:1rem;color:rgba(255,255,255,0.82);margin:0;line-height:1.6;">
            Empowering organizations to transform data governance from policy to practice through
            automated maturity assessment, quality monitoring, and intelligent issue resolution.</p>
    </div>""", unsafe_allow_html=True)

    st.markdown('<div class="dash-section-header"><div class="dash-section-dot"></div><h3>Our Solutions</h3><div class="dash-section-accent"></div></div>',unsafe_allow_html=True)

    # Custom CSS to make cards look like the design screenshot — no hover lift/glow
    st.markdown("""
    <style>
    .home-solution-card {
        background: #fff;
        border: 1.5px solid #e8e2f5;
        border-radius: 16px;
        padding: 1.4rem 1.5rem 1rem;
        min-height: 180px;
        position: relative;
        overflow: hidden;
        cursor: default;
        box-shadow: 0 2px 10px rgba(91,45,144,0.06);
        margin-bottom: 0.5rem;
    }
    .home-card-icon { font-size: 1.7rem; margin-bottom: 0.6rem; }
    .home-card-title { font-size: 0.97rem; font-weight: 700; color: #1a1a2e; margin-bottom: 0.4rem; }
    .home-card-desc  { font-size: 0.82rem; color: #555; line-height: 1.55; }
    .home-card-arrow { font-size: 0.9rem; color: #7a7a9a; margin-top: 0.6rem; display: block; }
    .home-card-accent {
        position: absolute; top: -28px; right: -28px;
        width: 90px; height: 90px; border-radius: 50%;
        opacity: 0.07;
    }
    .hca-magenta { background: #b10f74; }
    .hca-purple  { background: #5b2d90; }
    .hca-teal    { background: #0f766e; }
    .hca-amber   { background: #d97706; }
    /* Nav button below each card — styled as full-width outlined pill */
    div[data-testid="stHorizontalBlock"] .home-card-btn button {
        background: #fff !important;
        border: 1.5px solid #d9cef0 !important;
        color: #5b2d90 !important;
        font-weight: 600 !important;
        font-size: 0.85rem !important;
        border-radius: 12px !important;
        padding: 0.6rem 1rem !important;
        transition: background 0.15s, border-color 0.15s !important;
    }
    div[data-testid="stHorizontalBlock"] .home-card-btn button:hover {
        background: #f5f0fc !important;
        border-color: #9c6cd4 !important;
        color: #3b1f72 !important;
    }
    </style>
    """, unsafe_allow_html=True)

    c1,c2=st.columns(2,gap="large")
    with c1:
        st.markdown("""
        <div class="home-solution-card">
            <div class="home-card-accent hca-magenta"></div>
            <div class="home-card-icon">📈</div>
            <div class="home-card-title">Data Maturity Assessment</div>
            <div class="home-card-desc">Evaluate DAMA maturity dimensions across governance, quality, architecture, integration &amp; privacy.</div>
            <span class="home-card-arrow">→</span>
        </div>""", unsafe_allow_html=True)
        st.markdown('<div class="home-card-btn">', unsafe_allow_html=True)
        if st.button("Start Maturity Assessment →", use_container_width=True, key="home_mat"):
            st.session_state["page"] = "maturity"; st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)
    with c2:
        st.markdown("""
        <div class="home-solution-card">
            <div class="home-card-accent hca-purple"></div>
            <div class="home-card-icon">🔍</div>
            <div class="home-card-title">Data Quality Assessment</div>
            <div class="home-card-desc">Upload dataset, select dimensions &amp; rules, generate automated DQ scores with annexure reports.</div>
            <span class="home-card-arrow">→</span>
        </div>""", unsafe_allow_html=True)
        st.markdown('<div class="home-card-btn">', unsafe_allow_html=True)
        if st.button("Start DQ Assessment →", use_container_width=True, key="home_dq"):
            st.session_state["page"] = "dq"; st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    c3,c4=st.columns(2,gap="large")
    with c3:
        st.markdown("""
        <div class="home-solution-card">
            <div class="home-card-accent hca-teal"></div>
            <div class="home-card-icon">📋</div>
            <div class="home-card-title">Policy Hub</div>
            <div class="home-card-desc">Centralized governance repository for policy workflows, approval tracking, RBAC and compliance.</div>
            <span class="home-card-arrow">→</span>
        </div>""", unsafe_allow_html=True)
        st.markdown('<div class="home-card-btn">', unsafe_allow_html=True)
        if st.button("Open Policy Hub →", use_container_width=True, key="home_policy"):
            st.session_state["page"] = "policy"; st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)
    with c4:
        st.markdown("""
        <div class="home-solution-card">
            <div class="home-card-accent hca-amber"></div>
            <div class="home-card-icon">📁</div>
            <div class="home-card-title">Case Management</div>
            <div class="home-card-desc">Track, manage and resolve data quality issues and governance cases with full audit trail and escalation.</div>
            <span class="home-card-arrow">→</span>
        </div>""", unsafe_allow_html=True)
        st.markdown('<div class="home-card-btn">', unsafe_allow_html=True)
        if st.button("Open Case Management →", use_container_width=True, key="home_case"):
            st.session_state["page"] = "case"; st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    # ── About Us Section ───────────────────────────────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)
    _render_about_us()

    st.divider()
    _render_footer()


# ══════════════════════════════════════════════════════════════════════════════
#  SHARED: ABOUT US & FOOTER  (rendered on every page)
# ══════════════════════════════════════════════════════════════════════════════

def _render_about_us():
    """
    Uniqus Consultech – About Us section — white card style matching home solution cards.
    """
    st.markdown("""
    <style>
    /* ── About Us — white card shell matching home cards ── */
    .about-shell {
        background: #ffffff;
        border: 1.5px solid #e8e2f5;
        border-radius: 20px;
        padding: 1.8rem 2rem 1.5rem;
        margin-bottom: 1rem;
        box-shadow: 0 2px 12px rgba(91,45,144,0.07);
    }
    .about-header-row {
        display: flex;
        align-items: center;
        gap: 0.9rem;
        margin-bottom: 0.9rem;
        padding-bottom: 0.9rem;
        border-bottom: 2px solid #f0ebfa;
    }
    .about-logo-circle {
        width: 52px; height: 52px;
        border-radius: 14px;
        background: linear-gradient(135deg,#5b2d90,#b10f74);
        display: flex; align-items: center; justify-content: center;
        font-size: 1.7rem; flex-shrink: 0;
        box-shadow: 0 4px 14px rgba(91,45,144,0.25);
    }
    .about-header-text .about-name {
        font-size: 1.25rem;
        font-weight: 800;
        color: #1a1a2e;
        margin: 0 0 0.1rem;
    }
    .about-tagline {
        font-size: 0.74rem;
        font-style: italic;
        color: #9c6cd4;
        letter-spacing: 0.06em;
        text-transform: uppercase;
        font-weight: 600;
    }
    .about-intro {
        font-size: 0.88rem;
        color: #444;
        line-height: 1.75;
        margin-bottom: 1.2rem;
    }
    .about-intro strong { color: #3b1f72; }

    /* ── Stats strip ── */
    .about-stats {
        display: grid;
        grid-template-columns: repeat(6, 1fr);
        gap: 0.55rem;
        margin-bottom: 1.3rem;
    }
    @media(max-width:900px){ .about-stats { grid-template-columns: repeat(3,1fr); } }
    .about-stat {
        background: linear-gradient(135deg,#f5f0fc,#fdf2f8);
        border: 1.5px solid #e8e2f5;
        border-radius: 12px;
        padding: 0.7rem 0.4rem;
        text-align: center;
    }
    .about-stat-num {
        font-size: 1.35rem;
        font-weight: 800;
        background: linear-gradient(135deg,#5b2d90,#b10f74);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
        line-height: 1.1;
    }
    .about-stat-lbl {
        font-size: 0.63rem;
        color: #7a7a9a;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.04em;
        margin-top: 0.2rem;
    }

    /* ── Section label ── */
    .about-section-lbl {
        font-size: 0.68rem;
        font-weight: 700;
        text-transform: uppercase;
        letter-spacing: 0.09em;
        color: #b09dd6;
        margin-bottom: 0.55rem;
    }

    /* ── Practice pillars ── */
    .about-pillars {
        display: grid;
        grid-template-columns: repeat(4, 1fr);
        gap: 0.65rem;
        margin-bottom: 1.2rem;
    }
    @media(max-width:900px){ .about-pillars { grid-template-columns: repeat(2,1fr); } }
    .about-pillar {
        background: #faf8ff;
        border: 1.5px solid #e8e2f5;
        border-radius: 14px;
        padding: 0.9rem 0.85rem;
        transition: box-shadow 0.2s, border-color 0.2s;
    }
    .about-pillar:hover {
        box-shadow: 0 4px 16px rgba(91,45,144,0.12);
        border-color: #c4a8e8;
    }
    .about-pillar-icon { font-size: 1.3rem; margin-bottom: 0.3rem; }
    .about-pillar-title {
        font-size: 0.8rem;
        font-weight: 700;
        color: #3b1f72;
        margin-bottom: 0.25rem;
    }
    .about-pillar-desc {
        font-size: 0.72rem;
        color: #666;
        line-height: 1.5;
    }
    .about-pillar-badge {
        display: inline-block;
        background: linear-gradient(135deg,#f0e8fb,#fce7f5);
        border: 1px solid #d9cef0;
        border-radius: 999px;
        padding: 0.1rem 0.5rem;
        font-size: 0.62rem;
        font-weight: 700;
        color: #7c3aed;
        margin-top: 0.4rem;
        letter-spacing: 0.04em;
    }

    /* ── Tech Platforms row ── */
    .about-platforms {
        display: flex;
        flex-wrap: wrap;
        gap: 0.45rem;
        margin-bottom: 1.1rem;
    }
    .about-platform-chip {
        background: #f5f0fc;
        border: 1.5px solid #d9cef0;
        border-radius: 999px;
        padding: 0.25rem 0.8rem;
        font-size: 0.72rem;
        font-weight: 700;
        color: #5b2d90;
    }

    /* ── Investors row ── */
    .about-investors {
        display: flex;
        flex-wrap: wrap;
        gap: 0.45rem;
        margin-bottom: 0.9rem;
    }
    .about-inv-chip {
        background: #fff;
        border: 1.5px solid #e8e2f5;
        border-radius: 999px;
        padding: 0.2rem 0.75rem;
        font-size: 0.71rem;
        font-weight: 600;
        color: #555;
    }

    /* ── Bottom strip ── */
    .about-bottom {
        display: flex;
        align-items: center;
        justify-content: space-between;
        flex-wrap: wrap;
        gap: 0.5rem;
        padding-top: 0.85rem;
        border-top: 1.5px solid #f0ebfa;
        font-size: 0.77rem;
        color: #7a7a9a;
    }
    .about-bottom strong { color: #3b1f72; }
    .about-bottom a {
        color: #5b2d90;
        font-weight: 700;
        text-decoration: none;
    }
    .about-bottom a:hover { color: #b10f74; }
    </style>

    <div class="about-shell">

      <!-- Header row -->
      <div class="about-header-row">
        <div class="about-logo-circle">🏛️</div>
        <div class="about-header-text">
          <div class="about-name">Uniqus Consultech</div>
          <div class="about-tagline">Change the way consulting is done</div>
        </div>
      </div>

      <!-- Intro -->
      <p class="about-intro">
        <strong>Uniqus Consultech</strong> is a tech-enabled global consulting company founded in
        September 2022 by <strong>Jamil Khatri</strong> (former KPMG Partner &amp; Head of Accounting
        Advisory) and <strong>Sandip Khetan</strong> (former EY Partner). Uniqus specialises in
        Accounting &amp; Reporting, Governance Risk &amp; Compliance, Sustainability &amp; Climate,
        and Tech Consulting — delivering best-in-class solutions through proprietary technology
        platforms and a cloud-native global delivery model. Unlike traditional Big-4 firms, Uniqus
        operates exclusively as a consulting company (no audit), enabling it to raise capital, build
        technology products, and deploy an Employee Stock Ownership Plan — fundamentally redefining
        the consulting model.
      </p>

      <!-- Stats strip -->
      <div class="about-stats">
        <div class="about-stat"><div class="about-stat-num">700+</div><div class="about-stat-lbl">Professionals</div></div>
        <div class="about-stat"><div class="about-stat-num">400+</div><div class="about-stat-lbl">Clients Served</div></div>
        <div class="about-stat"><div class="about-stat-num">85+</div><div class="about-stat-lbl">Partners &amp; Directors</div></div>
        <div class="about-stat"><div class="about-stat-num">11</div><div class="about-stat-lbl">Global Offices</div></div>
        <div class="about-stat"><div class="about-stat-num">$250M</div><div class="about-stat-lbl">Valuation</div></div>
        <div class="about-stat"><div class="about-stat-num">2022</div><div class="about-stat-lbl">Year Founded</div></div>
      </div>

      <!-- Practice Pillars -->
      <div class="about-section-lbl">Our Practice Areas</div>
      <div class="about-pillars">
        <div class="about-pillar">
          <div class="about-pillar-icon">📊</div>
          <div class="about-pillar-title">Accounting &amp; Reporting (ARC)</div>
          <div class="about-pillar-desc">US GAAP, IFRS, SEC advisory, financial close automation,
            IPO readiness, and board-ready reporting for global and domestic markets.</div>
          <span class="about-pillar-badge">Reporting UniVerse</span>
        </div>
        <div class="about-pillar">
          <div class="about-pillar-icon">🛡️</div>
          <div class="about-pillar-title">Governance, Risk &amp; Compliance (GRC)</div>
          <div class="about-pillar-desc">SOX / ICOFR compliance, internal audit co-sourcing,
            ERM, internal controls design &amp; testing, and enterprise risk centralisation.</div>
          <span class="about-pillar-badge">Risk UniVerse</span>
        </div>
        <div class="about-pillar">
          <div class="about-pillar-icon">🌱</div>
          <div class="about-pillar-title">Sustainability &amp; Climate (SCC)</div>
          <div class="about-pillar-desc">ESG strategy, CSRD / BRSR / ISSB reporting, climate
            action, carbon accounting, and sustainability data integrity assurance.</div>
          <span class="about-pillar-badge">ESG UniVerse</span>
        </div>
        <div class="about-pillar">
          <div class="about-pillar-icon">🤖</div>
          <div class="about-pillar-title">Tech Consulting &amp; AI</div>
          <div class="about-pillar-desc">GenAI-powered solutions, finance automation, AI risk
            management, and digital transformation — combining domain depth with advanced technology.</div>
          <span class="about-pillar-badge">Uniqus AI</span>
        </div>
      </div>

      <!-- Platforms -->
      <div class="about-section-lbl">Proprietary Technology Platforms</div>
      <div class="about-platforms">
        <span class="about-platform-chip">🔵 Reporting UniVerse</span>
        <span class="about-platform-chip">🔴 Risk UniVerse</span>
        <span class="about-platform-chip">🟢 ESG UniVerse</span>
        <span class="about-platform-chip">🟣 UniQuest (AI Research)</span>
        <span class="about-platform-chip">✨ Uniqus AI (GenAI Engine)</span>
      </div>

      <!-- Investors -->
      <div class="about-section-lbl">Backed By</div>
      <div class="about-investors">
        <span class="about-inv-chip">🏦 Nexus Venture Partners</span>
        <span class="about-inv-chip">💼 Sorin Investments</span>
        <span class="about-inv-chip">🌐 UST Global</span>
        <span class="about-inv-chip">$42.5M Raised</span>
        <span class="about-inv-chip">Series C — April 2025</span>
      </div>

      <!-- Bottom strip -->
      <div class="about-bottom">
        <div>
          Clients include <strong>Bloom Energy, GAP, TaskUs, UST</strong>
          and 400+ enterprises across the US, India &amp; Middle East.
          Offices in <strong>San Jose · Mumbai · Dubai · Riyadh</strong> and 7 more cities globally.
        </div>
        <a href="https://uniqus.com/about-us/" target="_blank">🌐 uniqus.com/about-us →</a>
      </div>

    </div>
    """, unsafe_allow_html=True)


def _render_footer():
    """
    Uniqus Consultech — light-theme copyright footer matching the white card UI.
    Data copyright © Uniqus Consultech.
    """
    import datetime as _dt
    year = _dt.datetime.now().year
    st.markdown(f"""
    <style>
    .uq-footer {{
        background: #ffffff;
        border: 1.5px solid #e8e2f5;
        border-radius: 16px;
        padding: 1.2rem 1.8rem 1rem;
        margin-top: 1.5rem;
        box-shadow: 0 2px 12px rgba(91,45,144,0.07);
    }}
    .uq-footer-top {{
        display: flex;
        align-items: center;
        justify-content: space-between;
        flex-wrap: wrap;
        gap: 0.8rem;
        padding-bottom: 0.8rem;
        border-bottom: 1.5px solid #f0ebfa;
        margin-bottom: 0.7rem;
    }}
    .uq-footer-brand {{
        display: flex;
        align-items: center;
        gap: 0.65rem;
    }}
    .uq-footer-logo-circle {{
        width: 36px; height: 36px;
        border-radius: 10px;
        background: linear-gradient(135deg,#5b2d90,#b10f74);
        display: flex; align-items: center; justify-content: center;
        font-size: 1.1rem; flex-shrink: 0;
    }}
    .uq-footer-name {{
        font-size: 0.92rem;
        font-weight: 800;
        color: #1a1a2e;
        letter-spacing: 0.01em;
        font-family: 'Plus Jakarta Sans',sans-serif;
        margin: 0;
    }}
    .uq-footer-tagline {{
        font-size: 0.68rem;
        color: #9c6cd4;
        font-style: italic;
        font-weight: 600;
        font-family: 'Plus Jakarta Sans',sans-serif;
    }}
    .uq-footer-links {{
        display: flex;
        gap: 0.9rem;
        flex-wrap: wrap;
        align-items: center;
    }}
    .uq-footer-links a {{
        color: #5b2d90;
        font-size: 0.75rem;
        font-weight: 600;
        text-decoration: none;
        font-family: 'Plus Jakarta Sans',sans-serif;
        transition: color 0.15s;
    }}
    .uq-footer-links a:hover {{ color: #b10f74; }}
    .uq-footer-bottom {{
        display: flex;
        align-items: center;
        justify-content: space-between;
        flex-wrap: wrap;
        gap: 0.5rem;
    }}
    .uq-footer-copy {{
        font-size: 0.72rem;
        color: #7a7a9a;
        font-family: 'Plus Jakarta Sans',sans-serif;
        line-height: 1.6;
    }}
    .uq-footer-copy strong {{
        color: #3b1f72;
        font-weight: 700;
    }}
    .uq-footer-pills {{
        display: flex;
        gap: 0.4rem;
        flex-wrap: wrap;
    }}
    .uq-footer-pill {{
        background: linear-gradient(135deg,#f5f0fc,#fdf2f8);
        border: 1.5px solid #e0d5f5;
        border-radius: 999px;
        padding: 0.16rem 0.6rem;
        font-size: 0.65rem;
        font-weight: 700;
        color: #5b2d90;
        font-family: 'Plus Jakarta Sans',sans-serif;
    }}
    </style>

    <div class="uq-footer">
      <div class="uq-footer-top">
        <div class="uq-footer-brand">
          <div class="uq-footer-logo-circle">🏛️</div>
          <div>
            <div class="uq-footer-name">Uniqus Consultech</div>
            <div class="uq-footer-tagline">Change the way consulting is done</div>
          </div>
        </div>
        <div class="uq-footer-links">
          <a href="https://uniqus.com" target="_blank">🌐 uniqus.com</a>
          <a href="https://uniqus.com/about-us/" target="_blank">About Us</a>
          <a href="https://uniqus.com/our-services/" target="_blank">Services</a>
          <a href="https://uniqus.com/contact-us/" target="_blank">Contact</a>
          <a href="https://www.linkedin.com/company/uniqus-consultech" target="_blank">LinkedIn</a>
        </div>
      </div>
      <div class="uq-footer-bottom">
        <div class="uq-footer-copy">
          <strong>© {year} Uniqus Consultech Inc.</strong> All rights reserved.
          &nbsp;·&nbsp; Enterprise Data Management Platform
          &nbsp;·&nbsp; <strong>Data copyright © Uniqus Consultech</strong>
          &nbsp;·&nbsp; copyright@uniqus.com
        </div>
        <div class="uq-footer-pills">
          <span class="uq-footer-pill">🇺🇸 USA</span>
          <span class="uq-footer-pill">🇮🇳 India</span>
          <span class="uq-footer-pill">🌍 Middle East</span>
          <span class="uq-footer-pill">700+ Professionals</span>
          <span class="uq-footer-pill">400+ Clients</span>
        </div>
      </div>
    </div>
    """, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
#  PAGE: DQ ASSESSMENT — DYNAMIC BUSINESS RULE CRITERIA BUILDER
# ══════════════════════════════════════════════════════════════════════════════

# ── Dataset-type rule library ─────────────────────────────────────────────────
_DATASET_RULE_LIBRARY = {
    "Customer": {
        "completeness": ["Not Null","Not Empty","Mandatory Column","Minimum Length"],
        "validity": [
            "Email Format","Phone Format","Numeric Range",
            "Allowed Values","Custom Regex","Data Type Validation",
            "Length Check","Format Check",
        ],
        "standardization": [
            "Special Characters Not Allowed","Convert to Uppercase","Convert to Lowercase","Date Format",
        ],
        "suggested_mappings": {
            "email":["Email Format"],"phone":["Phone Format"],"mobile":["Phone Format"],
            "dob":["Date Format"],"date_of_birth":["Date Format"],"birthdate":["Date Format"],
            "age":["Numeric Range"],"gender":["Allowed Values"],"sex":["Allowed Values"],
            "country":["Allowed Values"],"status":["Allowed Values"],"name":["Minimum Length"],
            "first_name":["Minimum Length"],"last_name":["Minimum Length"],
            "customer_id":["Not Null","Not Empty"],"id":["Not Null"],
        },
    },
    "Vendor": {
        "completeness": ["Not Null","Not Empty","Mandatory Column","Minimum Length"],
        "validity": [
            "Email Format","Phone Format","Numeric Range","Allowed Values",
            "Custom Regex","Data Type Validation","PAN Format",
            "Length Check","Format Check",
        ],
        "standardization": [
            "Special Characters Not Allowed","Convert to Uppercase","Convert to Lowercase","Date Format",
        ],
        "suggested_mappings": {
            "vendor_email":["Email Format"],"contact_email":["Email Format"],
            "phone":["Phone Format"],"pan":["PAN Format"],"pan_number":["PAN Format"],
            "gstin":["Custom Regex"],"gst":["Custom Regex"],
            "payment_terms":["Allowed Values"],"status":["Allowed Values"],
            "vendor_id":["Not Null","Not Empty"],"vendor_name":["Minimum Length"],
            "credit_limit":["Numeric Range"],"rating":["Numeric Range"],
        },
    },
    "Product": {
        "completeness": ["Not Null","Not Empty","Mandatory Column","Minimum Length"],
        "validity": [
            "Numeric Range","Allowed Values","Custom Regex","Data Type Validation",
            "Length Check","Format Check",
        ],
        "standardization": [
            "Special Characters Not Allowed","Convert to Uppercase","Convert to Lowercase","Date Format",
        ],
        "suggested_mappings": {
            "sku":["Not Null","Not Empty"],"product_id":["Not Null"],
            "price":["Numeric Range"],"cost":["Numeric Range"],"mrp":["Numeric Range"],
            "quantity":["Numeric Range"],"stock":["Numeric Range"],
            "category":["Allowed Values"],"status":["Allowed Values"],
            "launch_date":["Date Format"],"expiry_date":["Date Format"],
            "product_name":["Minimum Length"],"description":["Minimum Length"],
        },
    },
    "Finance": {
        "completeness": ["Not Null","Not Empty","Mandatory Column"],
        "validity": [
            "Numeric Range","Allowed Values","Data Type Validation",
            "Custom Regex","Length Check","Format Check",
        ],
        "standardization": [
            "Special Characters Not Allowed","Convert to Uppercase","Convert to Lowercase","Date Format",
        ],
        "suggested_mappings": {
            "amount":["Numeric Range"],"debit":["Numeric Range"],"credit":["Numeric Range"],
            "transaction_date":["Date Format"],"posting_date":["Date Format"],
            "account_no":["Custom Regex"],"ifsc":["Custom Regex"],
            "currency":["Allowed Values"],"transaction_type":["Allowed Values"],
            "gl_code":["Not Null","Not Empty"],"cost_center":["Not Null"],
        },
    },
    "Other": {
        "completeness": ["Not Null","Not Empty","Mandatory Column","Minimum Length","Whitespace Only"],
        "validity": [
            "Email Format","Phone Format","Numeric Range",
            "Allowed Values","Custom Regex","Data Type Validation","PAN Format",
            "Length Check","Format Check",
        ],
        "standardization": [
            "Special Characters Not Allowed","Convert to Uppercase","Convert to Lowercase","Date Format",
        ],
        "suggested_mappings": {},
    },
}

_RULE_TOOLTIPS = {
    # ── Completeness Rules ──
    "Not Null": "Flags rows where the column value is NULL or missing.",
    "Not Empty": "Flags rows where the column value is an empty string.",
    "Whitespace Only": "Flags rows containing only spaces or blank characters.",
    "Minimum Length": "Flags values shorter than the specified minimum character count.",
    "Mandatory Column": "Treats column as required — all rows must have a valid value.",
    # ── Validation Rules (check only, no modification) ──
    "Email Format": "Validates email address format (user@domain.tld).",
    "Phone Format": "Validates phone numbers (7–15 digits, allows +, spaces, dashes).",
    "PAN Format": "Validates Indian PAN card format (AAAAA9999A).",
    "Numeric Range": "Flags numeric values outside the configured Min / Max range.",
    "Allowed Values": "Flags values not in the specified allowed list.",
    "Custom Regex": "Validates values against a custom regular expression pattern.",
    "Data Type Validation": "Checks that column values match the expected data type.",
    "Length Check": "Flags values that exceed the configured maximum character length.",
    "Format Check": "Validates values against expected format patterns (e.g. alphanumeric only).",
    # ── Standardisation Rules (flag non-standard, no auto-correction) ──
    "Special Characters Not Allowed": "Flags values containing special characters (non-alphanumeric, non-space).",
    "Convert to Uppercase": "Flags text values that are not in UPPERCASE.",
    "Convert to Lowercase": "Flags text values that are not in lowercase.",
    "Date Format": "Validates date values against expected format (e.g. %Y-%m-%d).",
}

_COMP_DIMENSION   = "Completeness"
_VAL_DIMENSION    = "Validity"
_STD_DIMENSION    = "Standardization"

# ── Change 8: Clear separation of Validation vs Standardisation rules ──
# VALIDATION RULES: Only check/flag, NO modification of data
_VALIDATION_ONLY_RULES = {
    "Not Null", "Not Empty", "Whitespace Only",
    "Minimum Length", "Mandatory Column",
    "Email Format", "Phone Format", "PAN Format",
    "Numeric Range", "Allowed Values",
    "Custom Regex", "Data Type Validation",
    "Format Check", "Length Check",
}

# STANDARDISATION RULES: Transformations that modify data
_STANDARDISATION_RULES = {
    "Special Characters Not Allowed",
    "Convert to Uppercase",
    "Convert to Lowercase",
    "Date Format",
}

# NOTE: The system does NOT auto-correct data.
# Standardisation rules only HIGHLIGHT records that violate the expected standard.
# The DQ engine flags non-standard values but preserves the original data.

def _auto_detect_dataset_type(columns):
    """Suggest a dataset type based on column name heuristics."""
    col_set = {c.lower().replace(" ","_").replace("-","_") for c in columns}
    scores = {"Customer":0,"Vendor":0,"Product":0,"Finance":0}
    customer_hints = {"email","phone","mobile","age","gender","dob","customer_id","first_name","last_name","address","zip","postal"}
    vendor_hints   = {"vendor","supplier","gstin","pan","ifsc","vendor_id","vendor_name","payment_terms"}
    product_hints  = {"sku","product_id","mrp","price","category","brand","uom","unit_of_measure","barcode"}
    finance_hints  = {"amount","debit","credit","gl_code","cost_center","ledger","account_no","transaction","voucher"}
    for c in col_set:
        for h in customer_hints:
            if h in c: scores["Customer"]+=1
        for h in vendor_hints:
            if h in c: scores["Vendor"]+=1
        for h in product_hints:
            if h in c: scores["Product"]+=1
        for h in finance_hints:
            if h in c: scores["Finance"]+=1
    best = max(scores, key=scores.get)
    return best if scores[best] > 0 else "Other"


def _build_dq_criteria_builder(all_columns, df_sample=None):
    """
    Renders the full Dynamic Business Rule Criteria Builder.
    Returns (rules_by_dim, merged_cfg) compatible with _run_dq_assessment().
    """
    # ── Session-state boot ────────────────────────────────────────────
    if "dq_rule_entries" not in st.session_state:
        st.session_state["dq_rule_entries"] = []          # unified rule list
    if "dq_dataset_type" not in st.session_state:
        st.session_state["dq_dataset_type"] = "Customer"
    if "dq_cb_search" not in st.session_state:
        st.session_state["dq_cb_search"] = ""

    detected_type = _auto_detect_dataset_type(all_columns)
    dataset_type = st.session_state.get("dq_dataset_type", detected_type)
    st.session_state["dq_dataset_type"] = dataset_type

    library = _DATASET_RULE_LIBRARY[dataset_type]
    rule_entries = st.session_state["dq_rule_entries"]

    # ── ② Criteria Builder ────────────────────────────────────────────
    st.markdown("""
    <div style="background:linear-gradient(135deg,#f5f0fc,#fdf2f8);border:1.5px solid #d9cef0;
         border-radius:14px;padding:1.1rem 1.4rem 0.5rem;margin-bottom:0.6rem;margin-top:0.5rem;">
        <div style="font-size:0.75rem;font-weight:700;text-transform:uppercase;letter-spacing:0.08em;
             color:#7a4dbb;margin-bottom:0.3rem;">② Rule Criteria Builder</div>
        <div style="font-size:0.82rem;color:#5b2d90;font-weight:500;">
            Configure rules per column across Completeness, Validity and Standardization dimensions.
            Use <strong>Bulk Apply</strong> to map one rule to multiple columns at once.
        </div>
    </div>""", unsafe_allow_html=True)

    tab_single, tab_suggest, tab_bulk, tab_view = st.tabs([
        "➕ Add Single Rule", "🤖 Smart Suggestions", "⚡ Bulk Apply", "📋 Rule Summary"
    ])

    # ── Tab A: Add Single Rule ────────────────────────────────────────
    with tab_single:
        _render_add_single_rule(all_columns, library, rule_entries)

    # ── Tab B: Smart Suggestions ─────────────────────────────────────
    with tab_suggest:
        _render_smart_suggestions(all_columns, library, rule_entries)

    # ── Tab C: Bulk Apply ────────────────────────────────────────────
    with tab_bulk:
        _render_bulk_apply(all_columns, library, rule_entries)

    # ── Tab D: Rule Summary ──────────────────────────────────────────
    with tab_view:
        _render_rule_summary(rule_entries)

    # ── Inline mini-summary below builder ────────────────────────────
    if rule_entries:
        _render_inline_rule_grid(rule_entries)

    # ── Extract rules_by_dim + merged_cfg ─────────────────────────────
    return _extract_rules_cfg(rule_entries, all_columns)


def _render_add_single_rule(all_columns, library, rule_entries):
    """Tab A — Add a single column→dimension→rule mapping with full config."""
    st.markdown(
        '<p style="font-size:0.83rem;color:#555;margin-bottom:0.7rem;">'
        'Map one rule to one column with full configuration. '
        'Rules with complex config (Numeric Range, Allowed Values, etc.) show inline inputs.</p>',
        unsafe_allow_html=True,
    )

    all_dims      = [_COMP_DIMENSION, _VAL_DIMENSION, _STD_DIMENSION]
    dim_rule_map  = {
        _COMP_DIMENSION: library["completeness"],
        _VAL_DIMENSION:  library["validity"],
        _STD_DIMENSION:  library["standardization"],
    }

    sr_c1, sr_c2, sr_c3 = st.columns([2, 1.5, 2])
    with sr_c1:
        search_col = st.text_input(
            "🔍 Search columns",
            value=st.session_state.get("dq_cb_search",""),
            key="dq_cb_search_input",
            placeholder="Type to filter columns…",
        )
        st.session_state["dq_cb_search"] = search_col
        filtered_cols = [c for c in all_columns
                         if search_col.lower() in c.lower()] if search_col else all_columns
        sel_col = st.selectbox(
            "Column",
            options=filtered_cols,
            key="dq_sr_col",
            help="Column to apply the rule to.",
        )
    with sr_c2:
        sel_dim = st.selectbox(
            "Dimension",
            options=all_dims,
            key="dq_sr_dim",
        )
    with sr_c3:
        available_rules = dim_rule_map.get(sel_dim, [])
        sel_rule = st.selectbox(
            "Rule",
            options=available_rules,
            key="dq_sr_rule",
            help=_RULE_TOOLTIPS.get(available_rules[0] if available_rules else "", ""),
        )
        if sel_rule:
            tip = _RULE_TOOLTIPS.get(sel_rule, "")
            if tip:
                st.caption(f"ℹ️ {tip}")

    # Config inputs
    rule_cfg = _render_rule_config_inputs(sel_rule, key_prefix="sr")

    # Mandatory flag for Completeness
    is_mandatory = False
    if sel_dim == _COMP_DIMENSION:
        is_mandatory = st.checkbox(
            "Mark as Mandatory (highlights column in summary)",
            key="dq_sr_mandatory",
        )
        if is_mandatory:
            rule_cfg["mandatory"] = True

    _add_btn_col, _ = st.columns([1, 3])
    with _add_btn_col:
        if st.button("➕ Add Rule Mapping", key="dq_sr_add", use_container_width=True, type="primary"):
            if not sel_col or not sel_rule:
                st.warning("⚠️ Select a column and a rule.")
            else:
                rule_entries.append({
                    "column":    sel_col,
                    "dimension": sel_dim,
                    "rule":      sel_rule,
                    "config":    rule_cfg,
                    "mandatory": is_mandatory,
                })
                st.success(f"✅ **{sel_rule}** → **{sel_col}** [{sel_dim}]")
                st.rerun()


def _render_rule_config_inputs(rule_name, key_prefix="cfg"):
    """Render inline config widgets for parameterised rules. Returns config dict.
    Enhanced: business-context defaults, flexible formats, expanded domain options."""
    cfg = {}
    if rule_name == "Numeric Range":
        r1, r2 = st.columns(2)
        with r1:
            cfg["range_min"] = st.number_input("Min Value", value=0.0, key=f"{key_prefix}_rmin",
                                                help="Minimum acceptable value (inclusive). Set based on your business domain.")
        with r2:
            cfg["range_max"] = st.number_input("Max Value", value=999999.0, key=f"{key_prefix}_rmax",
                                                help="Maximum acceptable value (inclusive). Use realistic thresholds for your data.")
    elif rule_name == "Allowed Values":
        cfg["allowed_values_str"] = st.text_input(
            "Comma-separated allowed values",
            key=f"{key_prefix}_allowed",
            placeholder="e.g. Male,Female,Other  or  Active,Inactive,Pending",
            help="Case-insensitive matching. Values are trimmed of whitespace before comparison.",
        )
        cfg["case_insensitive"] = st.checkbox(
            "Case-insensitive matching",
            value=True,
            key=f"{key_prefix}_allowed_ci",
            help="When enabled, 'male' matches 'Male', 'MALE', etc.",
        )
    elif rule_name == "Custom Regex":
        cfg["custom_regex"] = st.text_input(
            "Regular Expression Pattern",
            key=f"{key_prefix}_regex",
            placeholder="e.g. ^[A-Z]{2}[0-9]{6}$",
            help="Python regex pattern. Use ^ and $ for full-match validation.",
        )
    elif rule_name == "Date Format":
        cfg["date_fmt"] = st.text_input(
            "Date format (leave empty for auto-detect)",
            value="",
            key=f"{key_prefix}_datefmt",
            placeholder="e.g. %Y-%m-%d, %d/%m/%Y, %m-%d-%Y",
            help="Python strftime format. Empty = auto-detect common formats.",
        )
    elif rule_name == "Minimum Length":
        cfg["min_length_val"] = st.number_input(
            "Minimum length (characters)",
            min_value=1, value=2, key=f"{key_prefix}_minlen",
            help="Minimum character count. Set to 2 for names, 3+ for descriptions.",
        )
    elif rule_name == "Phone Format":
        cfg["phone_flexible"] = st.checkbox(
            "Flexible phone format (allow +, spaces, dashes, parentheses, 7-15 digits)",
            value=True,
            key=f"{key_prefix}_phone_flex",
            help="Accepts international formats: +91 9876543210, (555) 123-4567, etc.",
        )
    elif rule_name == "Email Format":
        cfg["email_flexible"] = st.checkbox(
            "Flexible email validation (allow subdomains, + aliases)",
            value=True,
            key=f"{key_prefix}_email_flex",
            help="Accepts user+tag@subdomain.example.com patterns.",
        )
    elif rule_name == "Data Type Validation":
        cfg["data_type"] = st.selectbox(
            "Expected data type",
            options=["string","numeric","integer","float","date"],
            key=f"{key_prefix}_dtype",
        )
    elif rule_name == "PAN Format":
        cfg["pan_pattern"] = st.text_input(
            "PAN regex (pre-filled for Indian PAN)",
            value="^[A-Z]{5}[0-9]{4}[A-Z]$",
            key=f"{key_prefix}_pan",
            help="Default: Indian PAN format AAAAA9999A. Modify for other ID formats.",
        )
    elif rule_name == "Special Characters Not Allowed":
        cfg["allowed_chars_pattern"] = st.text_input(
            "Allowed characters regex (values matching this are VALID)",
            value="^[a-zA-Z0-9\\s]+$",
            key=f"{key_prefix}_specchar",
            help="Default: alphanumeric + spaces only. Modify to allow specific characters.",
        )
    elif rule_name == "Length Check":
        cfg["max_length_val"] = st.number_input(
            "Maximum length (characters)",
            min_value=1, value=255, key=f"{key_prefix}_maxlen",
            help="Values exceeding this length will be flagged.",
        )
    elif rule_name == "Format Check":
        cfg["format_pattern"] = st.text_input(
            "Expected format regex pattern",
            value="^[a-zA-Z0-9]+$",
            key=f"{key_prefix}_fmtcheck",
            placeholder="e.g. ^[A-Z]{2}[0-9]{4}$",
            help="Python regex. Values NOT matching this pattern will be flagged.",
        )
    elif rule_name == "Convert to Uppercase":
        st.caption("ℹ️ Flags text values that are not in UPPERCASE.")
    elif rule_name == "Convert to Lowercase":
        st.caption("ℹ️ Flags text values that are not in lowercase.")
    return cfg


def _render_bulk_apply(all_columns, library, rule_entries):
    """Tab B — Apply one rule to many columns simultaneously."""
    st.markdown(
        '<p style="font-size:0.83rem;color:#555;margin-bottom:0.7rem;">'
        'Select multiple columns and apply the same rule to all of them at once. '
        'Great for applying Not Null or Trim Spaces across a whole group of columns.</p>',
        unsafe_allow_html=True,
    )
    all_dims = [_COMP_DIMENSION, _VAL_DIMENSION, _STD_DIMENSION]
    dim_rule_map = {
        _COMP_DIMENSION: library["completeness"],
        _VAL_DIMENSION:  library["validity"],
        _STD_DIMENSION:  library["standardization"],
    }
    b1, b2 = st.columns([1.5, 1])
    with b1:
        bulk_cols = st.multiselect(
            "Select Columns (bulk)",
            options=all_columns,
            key="dq_bulk_cols",
            placeholder="Choose columns…",
        )
    with b2:
        bulk_dim = st.selectbox("Dimension", options=all_dims, key="dq_bulk_dim")

    bulk_rule = st.selectbox(
        "Rule to apply",
        options=dim_rule_map.get(bulk_dim, []),
        key="dq_bulk_rule",
        help=_RULE_TOOLTIPS.get(dim_rule_map.get(bulk_dim,[""])[0] if dim_rule_map.get(bulk_dim) else "", ""),
    )
    if bulk_rule:
        st.caption(f"ℹ️ {_RULE_TOOLTIPS.get(bulk_rule,'')}")

    bulk_cfg = _render_rule_config_inputs(bulk_rule, key_prefix="bulk")

    bulk_mandatory = False
    if bulk_dim == _COMP_DIMENSION:
        bulk_mandatory = st.checkbox("Mark all as Mandatory", key="dq_bulk_mandatory")

    _bb, _ = st.columns([1, 3])
    with _bb:
        if st.button("⚡ Bulk Add Rules", key="dq_bulk_add", use_container_width=True, type="primary"):
            if not bulk_cols:
                st.warning("⚠️ Select at least one column.")
            elif not bulk_rule:
                st.warning("⚠️ Select a rule.")
            else:
                added = 0
                for col in bulk_cols:
                    rule_entries.append({
                        "column":    col,
                        "dimension": bulk_dim,
                        "rule":      bulk_rule,
                        "config":    dict(bulk_cfg),
                        "mandatory": bulk_mandatory,
                    })
                    added += 1
                st.success(f"✅ Added **{bulk_rule}** to {added} column(s) [{bulk_dim}]")
                st.rerun()


def _render_smart_suggestions(all_columns, library, rule_entries):
    """Tab C — Auto-suggest rules based on column names and dataset type."""
    st.markdown(
        '<p style="font-size:0.83rem;color:#555;margin-bottom:0.7rem;">'
        'Rules are suggested by matching column names to known patterns for this dataset type. '
        'Click <strong>Apply All Suggestions</strong> or cherry-pick individual ones.</p>',
        unsafe_allow_html=True,
    )
    suggestions = library.get("suggested_mappings", {})
    matches = []
    for col in all_columns:
        col_key = col.lower().replace(" ","_").replace("-","_")
        # exact match
        if col_key in suggestions:
            for rule in suggestions[col_key]:
                dim = _rule_to_dim(rule, library)
                if dim:
                    matches.append({"column":col,"dimension":dim,"rule":rule,"config":{},"mandatory":False})
        else:
            # partial match
            for key, rules in suggestions.items():
                if key in col_key or col_key in key:
                    for rule in rules:
                        dim = _rule_to_dim(rule, library)
                        if dim:
                            matches.append({"column":col,"dimension":dim,"rule":rule,"config":{},"mandatory":False})
                    break

    if not matches:
        st.info(f"ℹ️ No automatic suggestions found for the **{st.session_state.get('dq_dataset_type','Customer')}** dataset type and current columns. Try a different dataset type or add rules manually.")
    else:
        st.markdown(
            f'<div style="background:#f0fdf4;border:1px solid #bbf7d0;border-radius:8px;'
            f'padding:0.6rem 1rem;font-size:0.83rem;color:#166534;margin-bottom:0.6rem;">'
            f'🤖 <strong>{len(matches)}</strong> rule suggestion(s) found based on column name patterns.</div>',
            unsafe_allow_html=True,
        )
        # Preview table
        preview_df = pd.DataFrame([{
            "Column":    m["column"],
            "Dimension": m["dimension"],
            "Rule":      m["rule"],
        } for m in matches])
        st.dataframe(preview_df, use_container_width=True, hide_index=True, height=min(len(matches)*38+50, 320))

        sug_c1, sug_c2 = st.columns([1, 3])
        with sug_c1:
            if st.button("✨ Apply All Suggestions", key="dq_apply_all_sug", use_container_width=True, type="primary"):
                added = 0
                existing_keys = {(r["column"],r["dimension"],r["rule"]) for r in rule_entries}
                for m in matches:
                    key = (m["column"],m["dimension"],m["rule"])
                    if key not in existing_keys:
                        rule_entries.append(dict(m))
                        existing_keys.add(key)
                        added += 1
                st.success(f"✅ Applied {added} new suggestion(s).")
                st.rerun()


def _rule_to_dim(rule_name, library):
    """Map a rule name back to its dimension based on the current library."""
    if rule_name in library["completeness"]: return _COMP_DIMENSION
    if rule_name in library["validity"]:     return _VAL_DIMENSION
    if rule_name in library["standardization"]: return _STD_DIMENSION
    return None


# ═══════════════════════════════════════════════════════════════════════
#  RULE SUMMARY — EDITABLE INLINE CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════

# Rules whose config inputs are mandatory (saving with empty value blocked)
_RULES_REQUIRING_CONFIG = {"Allowed Values", "Custom Regex"}

# Rules that have user-editable parameters
_CONFIGURABLE_RULES = {
    "Numeric Range", "Allowed Values", "Custom Regex", "Date Format",
    "Minimum Length",
    "Data Type Validation", "PAN Format", "Phone Format", "Email Format",
    "Special Characters Not Allowed", "Length Check", "Format Check",
    "Convert to Uppercase", "Convert to Lowercase",
}


def _format_cfg_display(rule_name, cfg):
    """Return a concise, human-readable string for a rule's configuration."""
    if not cfg:
        return "Default"
    parts = []
    rn = rule_name

    if rn == "Numeric Range":
        lo = cfg.get("range_min")
        hi = cfg.get("range_max")
        if lo is not None and hi is not None:
            parts.append(f"{lo}–{hi}")
        elif lo is not None:
            parts.append(f"Min {lo}")
        elif hi is not None:
            parts.append(f"Max {hi}")
    elif rn == "Allowed Values":
        raw = cfg.get("allowed_values_str", "")
        if raw:
            items = [v.strip() for v in raw.split(",") if v.strip()]
            preview = ", ".join(items[:4])
            if len(items) > 4:
                preview += f" (+{len(items)-4} more)"
            parts.append(preview)
        if cfg.get("case_insensitive") is False:
            parts.append("case-sensitive")
    elif rn == "Minimum Length":
        ml = cfg.get("min_length_val")
        if ml is not None:
            parts.append(f"{ml} characters")
    elif rn == "Custom Regex":
        p = cfg.get("custom_regex", "")
        if p:
            parts.append(f"/{p[:35]}{'…' if len(p)>35 else ''}/")
    elif rn == "Date Format":
        f = cfg.get("date_fmt", "")
        parts.append(f if f else "auto-detect")
    elif rn == "Normalize Date Format":
        f = cfg.get("date_target_fmt", "")
        if f:
            parts.append(f"→ {f}")
    elif rn == "Replace Null with Default":
        d = cfg.get("null_default", "")
        if d:
            parts.append(f'→ "{d}"')
    elif rn == "Data Type Validation":
        dt = cfg.get("data_type", "")
        if dt:
            parts.append(dt)
    elif rn == "PAN Format":
        p = cfg.get("pan_pattern", "")
        if p:
            parts.append(f"/{p[:30]}/")
    elif rn == "Phone Format":
        if cfg.get("phone_flexible"):
            parts.append("flexible")
    elif rn == "Email Format":
        if cfg.get("email_flexible"):
            parts.append("flexible")
    else:
        for k, v in cfg.items():
            if v not in ("", None, True, False):
                parts.append(f"{k}={v}")

    return "; ".join(parts) if parts else "Default"


def _render_inline_config_editor(entry_idx, rule_name, col_name, current_cfg):
    """
    Render dynamic input widgets for *one* rule entry.
    On 💾 Save the new config is written directly to
    ``st.session_state["dq_rule_entries"][entry_idx]["config"]``.
    """
    kp = f"rcfg_{entry_idx}"          # unique key prefix

    st.markdown(
        f'<div style="font-size:0.78rem;color:#3b1f72;font-weight:600;'
        f'margin-bottom:0.35rem;">Configure: <strong>{rule_name}</strong> '
        f'on column <strong>{col_name}</strong></div>',
        unsafe_allow_html=True,
    )

    new_cfg = dict(current_cfg)       # mutable copy
    validation_error = None           # set to a string to block Save

    # ── Rule-type specific widgets ────────────────────────────────
    if rule_name == "Numeric Range":
        c1, c2 = st.columns(2)
        with c1:
            new_cfg["range_min"] = st.number_input(
                "Min Value", value=float(current_cfg.get("range_min", 0.0)),
                key=f"{kp}_rmin",
                help="Minimum acceptable value (inclusive).",
            )
        with c2:
            new_cfg["range_max"] = st.number_input(
                "Max Value", value=float(current_cfg.get("range_max", 999999.0)),
                key=f"{kp}_rmax",
                help="Maximum acceptable value (inclusive).",
            )
        if new_cfg["range_min"] > new_cfg["range_max"]:
            validation_error = "Min value cannot exceed Max value."

    elif rule_name == "Allowed Values":
        new_cfg["allowed_values_str"] = st.text_area(
            "Allowed values (comma-separated)",
            value=current_cfg.get("allowed_values_str", ""),
            key=f"{kp}_allowed", height=80,
            placeholder="Active, Inactive, Suspended, Pending",
            help="Enter each allowed value separated by commas.",
        )
        new_cfg["case_insensitive"] = st.checkbox(
            "Case-insensitive matching",
            value=current_cfg.get("case_insensitive", True),
            key=f"{kp}_ci",
        )
        raw = new_cfg["allowed_values_str"]
        if raw:
            parsed = [v.strip() for v in raw.split(",") if v.strip()]
            st.markdown(
                f'<div style="font-size:0.76rem;color:#166534;background:#f0fdf4;'
                f'border:1px solid #bbf7d0;border-radius:6px;'
                f'padding:0.3rem 0.6rem;margin-top:0.2rem;">'
                f'✅ <strong>{len(parsed)}</strong> value(s): '
                f'{", ".join(parsed)}</div>',
                unsafe_allow_html=True,
            )
        else:
            validation_error = "Allowed Values requires at least one value."

    elif rule_name == "Minimum Length":
        new_cfg["min_length_val"] = st.number_input(
            "Minimum length (characters)",
            min_value=1, max_value=1000,
            value=int(current_cfg.get("min_length_val", 2)),
            step=1, key=f"{kp}_minlen",
            help="Values shorter than this will be flagged.",
        )

    elif rule_name == "Custom Regex":
        new_cfg["custom_regex"] = st.text_input(
            "Regular Expression Pattern",
            value=current_cfg.get("custom_regex", ""),
            key=f"{kp}_regex",
            placeholder="e.g. ^[A-Z]{2}[0-9]{6}$",
            help="Python regex. Use ^ and $ for full-match.",
        )
        pat = new_cfg["custom_regex"]
        if pat:
            import re as _re
            try:
                _re.compile(pat)
                st.markdown(
                    '<span style="font-size:0.76rem;color:#166534;">'
                    '✅ Valid regex pattern</span>',
                    unsafe_allow_html=True,
                )
            except _re.error as e:
                validation_error = f"Invalid regex: {e}"
        else:
            validation_error = "Custom Regex requires a pattern."

    elif rule_name == "Date Format":
        new_cfg["date_fmt"] = st.text_input(
            "Date format string",
            value=current_cfg.get("date_fmt", ""),
            key=f"{kp}_datefmt",
            placeholder="%Y-%m-%d  or  %d/%m/%Y  (empty = auto-detect)",
        )

    elif rule_name == "Normalize Date Format":
        new_cfg["date_target_fmt"] = st.text_input(
            "Target date format",
            value=current_cfg.get("date_target_fmt", "%Y-%m-%d"),
            key=f"{kp}_datetgt",
        )

    elif rule_name == "Replace Null with Default":
        new_cfg["null_default"] = st.text_input(
            "Default replacement value",
            value=current_cfg.get("null_default", "N/A"),
            key=f"{kp}_nulldef",
        )
        if not new_cfg["null_default"].strip():
            validation_error = "Default value cannot be empty."

    elif rule_name == "Data Type Validation":
        opts = ["string", "numeric", "integer", "float", "date"]
        cur  = current_cfg.get("data_type", "string")
        new_cfg["data_type"] = st.selectbox(
            "Expected data type", options=opts,
            index=opts.index(cur) if cur in opts else 0,
            key=f"{kp}_dtype",
        )

    elif rule_name == "PAN Format":
        new_cfg["pan_pattern"] = st.text_input(
            "PAN regex pattern",
            value=current_cfg.get("pan_pattern",
                                  "^[A-Z]{5}[0-9]{4}[A-Z]$"),
            key=f"{kp}_pan",
        )

    elif rule_name == "Phone Format":
        new_cfg["phone_flexible"] = st.checkbox(
            "Flexible phone format (allow +, spaces, dashes, "
            "parentheses, 7–15 digits)",
            value=current_cfg.get("phone_flexible", True),
            key=f"{kp}_phone_flex",
        )

    elif rule_name == "Email Format":
        new_cfg["email_flexible"] = st.checkbox(
            "Flexible email (allow subdomains, + aliases)",
            value=current_cfg.get("email_flexible", True),
            key=f"{kp}_email_flex",
        )

    # ── Validation banner ─────────────────────────────────────────
    if validation_error:
        st.markdown(
            f'<div style="font-size:0.78rem;color:#dc2626;background:#fef2f2;'
            f'border:1px solid #fecaca;border-radius:6px;'
            f'padding:0.35rem 0.7rem;margin-top:0.3rem;">'
            f'⚠️ {validation_error}</div>',
            unsafe_allow_html=True,
        )

    # ── Save / Reset ──────────────────────────────────────────────
    b1, b2, _ = st.columns([1, 1, 2])
    with b1:
        if st.button(
            "💾 Save", key=f"{kp}_save",
            use_container_width=True, type="primary",
            disabled=validation_error is not None,
        ):
            st.session_state["dq_rule_entries"][entry_idx]["config"] = new_cfg
            st.success(f"✅ Saved — {rule_name} → {col_name}")
            st.rerun()
    with b2:
        if st.button("↩ Reset to Default", key=f"{kp}_reset",
                      use_container_width=True):
            st.session_state["dq_rule_entries"][entry_idx]["config"] = {}
            st.rerun()


# ──────────────────────────────────────────────────────────────────────

def _render_rule_summary(rule_entries):
    """
    Tab D — Editable Rule Summary.

    Upper half:  read-only summary grid with formatted Configuration column.
    Lower half:  per-rule expandable inline editors (dynamic widgets based on
                 rule type).  Save writes back to session_state and persists
                 through assessment execution and Excel report generation.
    """
    if not rule_entries:
        st.info("No rule mappings configured yet. Use the other tabs to add rules.")
        return

    # ── Search / filter ──────────────────────────────────────────
    filt = st.text_input(
        "🔍 Filter by column or rule", key="dq_summary_filter",
        placeholder="Type to filter…",
    )
    filtered_indices = []
    for i, r in enumerate(rule_entries):
        if filt:
            if (filt.lower() not in r["column"].lower()
                    and filt.lower() not in r["rule"].lower()):
                continue
        filtered_indices.append(i)

    if not filtered_indices:
        st.info("No rules match the current filter.")
        return

    # ── Summary grid (read-only) ─────────────────────────────────
    display_rows = []
    for seq, idx in enumerate(filtered_indices):
        r = rule_entries[idx]
        cfg_display = _format_cfg_display(r["rule"], r.get("config", {}))
        display_rows.append({
            "#":             seq + 1,
            "Column":        r["column"],
            "Dimension":     r["dimension"],
            "Rule":          r["rule"],
            "Configuration": cfg_display,
            "Type":          "⭐ Mandatory" if r.get("mandatory") else "Optional",
        })

    st.dataframe(
        pd.DataFrame(display_rows),
        use_container_width=True, hide_index=True,
        height=min(len(display_rows) * 38 + 60, 420),
    )

    # ── Quick stats banner ───────────────────────────────────────
    n_configured = sum(
        1 for idx in filtered_indices
        if rule_entries[idx].get("config")
        and any(v not in ("", None) for v in rule_entries[idx]["config"].values())
    )
    n_default = len(filtered_indices) - n_configured
    st.markdown(
        f'<div style="background:linear-gradient(135deg,#f5f0fc,#fdf2f8);'
        f'border:1.5px solid #d9cef0;border-radius:10px;'
        f'padding:0.5rem 1rem;margin:0.5rem 0;font-size:0.82rem;color:#3b1f72;">'
        f'⚙️ <strong>{n_configured}</strong> rule(s) configured &nbsp;·&nbsp; '
        f'<span style="color:#7a7a9a;">{n_default} using defaults</span> &nbsp;·&nbsp; '
        f'Click <strong>✏️ Edit</strong> below to modify parameters'
        f'</div>',
        unsafe_allow_html=True,
    )

    # ── Inline config editors ────────────────────────────────────
    st.markdown(
        '<div style="font-size:0.85rem;font-weight:700;color:#5b2d90;'
        'margin:0.6rem 0 0.3rem;">✏️ Edit Rule Configuration</div>',
        unsafe_allow_html=True,
    )

    for seq, idx in enumerate(filtered_indices):
        r = rule_entries[idx]
        rule_name = r["rule"]
        col_name  = r["column"]
        cfg       = r.get("config", {})
        cfg_text  = _format_cfg_display(rule_name, cfg)

        is_configurable = rule_name in _CONFIGURABLE_RULES

        if is_configurable:
            # Determine icon: ⚠️ if mandatory-config rule has no value
            needs_attention = (
                rule_name in _RULES_REQUIRING_CONFIG
                and cfg_text == "Default"
            )
            icon = "⚠️" if needs_attention else "⚙️"
            label = (
                f"{icon} #{seq+1}  {rule_name} → {col_name}   "
                f"〔{cfg_text}〕"
            )
            with st.expander(label, expanded=needs_attention):
                _render_inline_config_editor(idx, rule_name, col_name, cfg)
        else:
            st.markdown(
                f'<div style="background:#f9f8fc;border:1px solid #e8e2f5;'
                f'border-radius:8px;padding:0.35rem 0.8rem;font-size:0.8rem;'
                f'color:#5b2d90;margin-bottom:0.25rem;">'
                f'<strong>#{seq+1}</strong> &nbsp; {rule_name} → {col_name}'
                f' &nbsp;<span style="color:#7a7a9a;">— no configurable '
                f'parameters</span></div>',
                unsafe_allow_html=True,
            )

    # ── Delete controls ──────────────────────────────────────────
    st.markdown(
        '<div style="font-size:0.8rem;color:#7a7a9a;'
        'margin:0.7rem 0 0.3rem;">Remove individual mappings:</div>',
        unsafe_allow_html=True,
    )
    n_del_cols = min(len(rule_entries), 8)
    del_cols = st.columns(n_del_cols)
    for ci, r in enumerate(rule_entries):
        with del_cols[ci % n_del_cols]:
            tip = f"Remove: {r['rule']} → {r['column']}"
            if st.button(f"🗑 #{ci+1}", key=f"dq_del_{ci}",
                         use_container_width=True, help=tip):
                st.session_state["dq_rule_entries"].pop(ci)
                st.rerun()

    _, cc = st.columns([4, 1])
    with cc:
        if st.button("🗑 Clear All Rules", key="dq_clear_all_rules",
                      use_container_width=True):
            st.session_state["dq_rule_entries"] = []
            st.rerun()


def _render_inline_rule_grid(rule_entries):
    """Compact active-rules count bar shown below the builder."""
    from collections import Counter
    dim_counts = Counter(r["dimension"] for r in rule_entries)
    cols_count = len(set(r["column"] for r in rule_entries))
    total = len(rule_entries)
    mand = sum(1 for r in rule_entries if r.get("mandatory"))
    parts = " &nbsp;·&nbsp; ".join(
        f"<span style='color:#5b2d90;font-weight:700;'>{v}</span> {k}"
        for k, v in dim_counts.items()
    )
    st.markdown(f"""
    <div style="background:linear-gradient(135deg,#f5f0fc,#fdf2f8);
         border:1.5px solid #d9cef0;border-radius:10px;
         padding:0.55rem 1rem;margin-top:0.6rem;font-size:0.82rem;color:#3b1f72;">
        <strong>Active rules:</strong> {total} across {cols_count} column(s) &nbsp;·&nbsp;
        {parts} &nbsp;·&nbsp;
        <span style="color:#b10f74;font-weight:700;">{mand}</span> mandatory
    </div>""", unsafe_allow_html=True)


def _extract_rules_cfg(rule_entries, all_columns):
    """
    Convert the unified rule_entries list into (rules_by_dim, merged_cfg)
    that is compatible with _run_dq_assessment().
    
    Enhanced: stores per-column configs in rule_column_configs for traceability
    and correctly merges configs without losing column-specific overrides.
    """
    rules_by_dim: dict = {}
    merged_cfg:   dict = {}

    # Per-column config registry: {(column, rule): config_dict}
    rule_column_configs: dict = {}

    # Collect per-dimension entries
    comp_entries = [r for r in rule_entries if r["dimension"] == _COMP_DIMENSION]
    val_entries  = [r for r in rule_entries if r["dimension"] == _VAL_DIMENSION]
    std_entries  = [r for r in rule_entries if r["dimension"] == _STD_DIMENSION]

    # ── Completeness ─────────────────────────────────────────────────
    if comp_entries:
        rules_by_dim[_COMP_DIMENSION] = list(dict.fromkeys(r["rule"] for r in comp_entries))
        comp_cols = list(dict.fromkeys(r["column"] for r in comp_entries))
        merged_cfg["comp_columns"] = comp_cols
        # Pick first min_length config if any
        for r in comp_entries:
            if "min_length_val" in r["config"]:
                merged_cfg["min_length_val"] = r["config"]["min_length_val"]
                break
        # Mandatory cols
        mandatory_cols = [r["column"] for r in comp_entries if r.get("mandatory") or r["rule"] == "Mandatory Column"]
        if mandatory_cols:
            merged_cfg["mandatory_cols"] = mandatory_cols
        # Store per-column configs
        for r in comp_entries:
            if r["config"]:
                rule_column_configs[(r["column"], r["rule"])] = dict(r["config"])

    # ── Validity ─────────────────────────────────────────────────────
    if val_entries:
        rules_by_dim[_VAL_DIMENSION] = list(dict.fromkeys(r["rule"] for r in val_entries))
        val_cols = list(dict.fromkeys(r["column"] for r in val_entries))
        merged_cfg["val_columns"] = val_cols
        # Per-column config: collect ALL configs, use last as global fallback
        # but store per-column overrides for traceability
        for r in val_entries:
            cfg = r["config"]
            if "range_min" in cfg:
                merged_cfg["range_min"] = cfg["range_min"]
            if "range_max" in cfg:
                merged_cfg["range_max"] = cfg["range_max"]
            if "allowed_values_str" in cfg:
                merged_cfg["allowed_values_str"] = cfg["allowed_values_str"]
            if "custom_regex" in cfg:
                merged_cfg["custom_regex"] = cfg["custom_regex"]
            if "date_fmt" in cfg:
                merged_cfg["date_fmt"] = cfg["date_fmt"]
            # Store per-column configs
            if cfg:
                rule_column_configs[(r["column"], r["rule"])] = dict(cfg)

    # ── Standardization ──────────────────────────────────────────────
    if std_entries:
        rules_by_dim[_STD_DIMENSION] = list(dict.fromkeys(r["rule"] for r in std_entries))
        std_cols = list(dict.fromkeys(r["column"] for r in std_entries))
        merged_cfg["std_columns"] = std_cols
        for r in std_entries:
            cfg = r["config"]
            if "date_target_fmt" in cfg:
                merged_cfg["date_target_fmt"] = cfg["date_target_fmt"]
            if "null_default" in cfg:
                merged_cfg["null_default"] = cfg["null_default"]
            if cfg:
                rule_column_configs[(r["column"], r["rule"])] = dict(cfg)

    # Store full per-column config registry
    merged_cfg["rule_column_configs"] = rule_column_configs

    return rules_by_dim, merged_cfg

def _ui_uniqueness(all_columns):
    st.markdown('<div class="dim-header">🔑 Uniqueness — Duplicate Criteria Builder</div>', unsafe_allow_html=True)

    # ── Session state initialisation ────────────────────────────────────
    if "dup_exact_rules" not in st.session_state:
        st.session_state["dup_exact_rules"] = []
    if "dup_fuzzy_rules" not in st.session_state:
        st.session_state["dup_fuzzy_rules"] = []

    exact_rules = st.session_state["dup_exact_rules"]
    fuzzy_rules = st.session_state["dup_fuzzy_rules"]

    # ── Tabs: Add Exact Rule | Add Fuzzy Rule ───────────────────────────
    tab_exact, tab_fuzzy = st.tabs(["Add Exact Rule", "Add Fuzzy Rule"])

    with tab_exact:
        st.markdown(
            '<p style="font-size:0.83rem;color:#555;margin-bottom:0.7rem;">'
            'Exact match on one or more columns (combination). '
            'Example: <code>VendorName + City + Country</code></p>',
            unsafe_allow_html=True,
        )
        ex_col_a, ex_col_b = st.columns([3, 2])
        with ex_col_a:
            exact_cols_sel = st.multiselect(
                "Exact rule columns",
                options=all_columns,
                key="dcb_exact_cols",
                placeholder="Choose options",
            )
        with ex_col_b:
            default_exact_name = ("EXACT: " + " + ".join(exact_cols_sel)) if exact_cols_sel else "EXACT: Rule"
            exact_rule_name = st.text_input(
                "Exact rule name",
                value=default_exact_name,
                key="dcb_exact_name",
            )
        exact_ignore_nulls = st.checkbox(
            "Ignore rows where ANY selected field is NULL/blank (exact)",
            value=True,
            key="dcb_exact_ignore_nulls",
        )
        _ebtn, _ = st.columns([1, 3])
        with _ebtn:
            if st.button("Add Exact Rule", key="dcb_add_exact", use_container_width=True, type="primary"):
                if not exact_cols_sel:
                    st.warning("⚠️ Select at least one column before adding a rule.")
                else:
                    rule_type = "Single Column Exact Match" if len(exact_cols_sel) == 1 else "Combination Column Exact Match"
                    st.session_state["dup_exact_rules"].append({
                        "name":         exact_rule_name.strip() or default_exact_name,
                        "cols":         list(exact_cols_sel),
                        "type":         rule_type,
                        "ignore_nulls": exact_ignore_nulls,
                    })
                    st.success(f"✅ Rule **{exact_rule_name}** added.")
                    st.rerun()

    with tab_fuzzy:
        st.markdown(
            '<p style="font-size:0.83rem;color:#555;margin-bottom:0.7rem;">'
            'Fuzzy (approximate) match across one or more columns using weighted similarity scoring.</p>',
            unsafe_allow_html=True,
        )
        fz_col_a, fz_col_b = st.columns([3, 2])
        with fz_col_a:
            fuzzy_cols_sel = st.multiselect(
                "Fuzzy rule columns",
                options=all_columns,
                key="dcb_fuzzy_cols",
                placeholder="Choose options",
            )
        with fz_col_b:
            default_fuzzy_name = ("FUZZY: " + " + ".join(fuzzy_cols_sel)) if fuzzy_cols_sel else "FUZZY: Rule"
            fuzzy_rule_name = st.text_input(
                "Fuzzy rule name",
                value=default_fuzzy_name,
                key="dcb_fuzzy_name",
            )

        # Per-column weights
        fuzzy_weights_input: List[float] = []
        if fuzzy_cols_sel:
            st.markdown(
                '<div style="font-size:0.82rem;font-weight:600;color:#5b2d90;margin-bottom:0.3rem;">'
                "Per-column weights (higher = more important)</div>",
                unsafe_allow_html=True,
            )
            w_cols = st.columns(min(len(fuzzy_cols_sel), 4))
            for wi, col_name in enumerate(fuzzy_cols_sel):
                with w_cols[wi % 4]:
                    w = st.number_input(
                        f"Weight: {col_name[:18]}",
                        min_value=0.0, max_value=10.0, value=1.0, step=0.5,
                        key=f"dcb_fzw_{wi}_{col_name}",
                    )
                    fuzzy_weights_input.append(w)

        fz_adv1, fz_adv2, fz_adv3 = st.columns(3)
        with fz_adv1:
            fuzzy_threshold_val = st.slider(
                "Similarity threshold (%)", 60, 99, 85, 1, key="dcb_fuzzy_threshold",
                help="Minimum weighted score to flag two records as duplicates.",
            )
        with fz_adv2:
            fuzzy_max_pairs_val = st.number_input(
                "Max pairs per block", min_value=1_000, max_value=500_000, value=20_000, step=1_000,
                key="dcb_fuzzy_max_pairs",
                help="Blocks with more pairwise comparisons than this are skipped.",
            )
        with fz_adv3:
            fuzzy_ignore_nulls_val = st.checkbox(
                "Ignore rows where ANY selected field is NULL/blank (fuzzy)",
                value=True, key="dcb_fuzzy_ignore_nulls",
            )
        _fbtn, _ = st.columns([1, 3])
        with _fbtn:
            if st.button("Add Fuzzy Rule", key="dcb_add_fuzzy", use_container_width=True, type="primary"):
                if not fuzzy_cols_sel:
                    st.warning("⚠️ Select at least one column before adding a rule.")
                else:
                    st.session_state["dup_fuzzy_rules"].append({
                        "name":         fuzzy_rule_name.strip() or default_fuzzy_name,
                        "cols":         list(fuzzy_cols_sel),
                        "weights":      fuzzy_weights_input if fuzzy_weights_input else [1.0] * len(fuzzy_cols_sel),
                        "threshold":    int(fuzzy_threshold_val),
                        "max_pairs":    int(fuzzy_max_pairs_val),
                        "ignore_nulls": fuzzy_ignore_nulls_val,
                    })
                    st.success(f"✅ Rule **{fuzzy_rule_name}** added.")
                    st.rerun()

    # ── Current duplicate rules display ─────────────────────────────────
    st.markdown("#### Current duplicate rules")
    all_added = exact_rules + fuzzy_rules
    if not all_added:
        st.markdown(
            '<div style="background:#eaf3fb;border:1px solid #b8d8f0;border-radius:8px;'
            'padding:0.7rem 1.1rem;font-size:0.88rem;color:#2c5f8a;">'
            "No rules added yet. Add at least one rule to run duplicate detection.</div>",
            unsafe_allow_html=True,
        )
    else:
        # Build display table
        display_rows = []
        for i, r in enumerate(exact_rules):
            display_rows.append({
                "#":          f"E{i+1}",
                "Rule Name":  r["name"],
                "Type":       r["type"],
                "Columns":    " + ".join(r["cols"]),
                "Threshold":  "Exact",
                "Skip Nulls": "Yes" if r["ignore_nulls"] else "No",
            })
        for i, r in enumerate(fuzzy_rules):
            display_rows.append({
                "#":          f"F{i+1}",
                "Rule Name":  r["name"],
                "Type":       "Hybrid Fuzzy Match",
                "Columns":    " + ".join(r["cols"]),
                "Threshold":  f"{r['threshold']}%",
                "Skip Nulls": "Yes" if r["ignore_nulls"] else "No",
            })
        st.dataframe(pd.DataFrame(display_rows), use_container_width=True, hide_index=True)

        # Delete buttons
        st.markdown('<div style="font-size:0.8rem;color:#7a7a9a;margin-bottom:0.35rem;">Remove a rule:</div>', unsafe_allow_html=True)
        all_keys = [(i, "exact") for i in range(len(exact_rules))] + [(i, "fuzzy") for i in range(len(fuzzy_rules))]
        del_cols = st.columns(min(len(all_keys), 6))
        for ci, (idx, rtype) in enumerate(all_keys):
            rname = (exact_rules if rtype == "exact" else fuzzy_rules)[idx]["name"]
            prefix = "E" if rtype == "exact" else "F"
            with del_cols[ci % 6]:
                if st.button(f"🗑 {prefix}{idx+1}", key=f"dcb_del_{rtype}_{idx}", use_container_width=True, help=f"Remove: {rname}"):
                    st.session_state[f"dup_{'exact' if rtype == 'exact' else 'fuzzy'}_rules"].pop(idx)
                    st.rerun()

        _, clear_col = st.columns([4, 1])
        with clear_col:
            if st.button("🗑 Clear All", key="dcb_clear_all", use_container_width=True):
                st.session_state["dup_exact_rules"] = []
                st.session_state["dup_fuzzy_rules"] = []
                st.rerun()

    # ── Build rules list + cfg for the engine ───────────────────────────
    rules: List[str] = []
    cfg: dict = {}

    # Collect exact rules
    single_cols_all: List[str] = []
    combo_rules: List[List[str]] = []
    for r in exact_rules:
        if r["type"] == "Single Column Exact Match":
            for c in r["cols"]:
                if c not in single_cols_all:
                    single_cols_all.append(c)
            if "Single Column Exact Match" not in rules:
                rules.append("Single Column Exact Match")
        else:
            combo_rules.append(r["cols"])
            if "Combination Column Exact Match" not in rules:
                rules.append("Combination Column Exact Match")

    if single_cols_all:
        cfg["single_dup_cols"] = single_cols_all
    if combo_rules:
        # Pass first combo rule's columns (engine currently supports one combo set)
        cfg["combo_dup_cols"] = combo_rules[0]

    # Collect fuzzy rules (pass first fuzzy rule to engine for now)
    if fuzzy_rules:
        rules.append("Hybrid Fuzzy Match")
        fr = fuzzy_rules[0]
        cfg["fuzzy_cols"]         = fr["cols"]
        cfg["fuzzy_threshold"]    = fr["threshold"]
        cfg["fuzzy_weights"]      = fr["weights"]
        cfg["fuzzy_max_pairs"]    = fr["max_pairs"]
        cfg["fuzzy_ignore_nulls"] = fr["ignore_nulls"]

    return rules, cfg

def _ui_standardization(all_columns):
    # Legacy stub — replaced by _build_dq_criteria_builder; kept for compatibility
    return [], {}


def _count_unique_duplicate_rows(dup_records, total_records):
    """
    Count unique row indices flagged as duplicates across all rules.
    Merges exact and fuzzy results, avoids double-counting rows flagged
    by multiple rules, and guarantees the count never exceeds total_records.
    """
    if dup_records is None or (hasattr(dup_records, 'empty') and dup_records.empty):
        return 0

    duplicate_row_ids = set()

    # Try to extract unique original row indices from common column patterns
    # produced by the duplicate detection engine (match-pair formats)
    idx_columns = [c for c in dup_records.columns if c.lower() in (
        'row_index', 'row_idx', 'row_number', 'row_num',
        'row_index_1', 'row_idx_1', 'index_1', 'idx_1',
        'row_index_2', 'row_idx_2', 'index_2', 'idx_2',
        'original_index', 'source_row', 'record_index',
    )]

    if idx_columns:
        for col in idx_columns:
            duplicate_row_ids.update(
                dup_records[col].dropna().astype(int).tolist()
            )
    else:
        duplicate_row_ids.update(dup_records.index.tolist())

    unique_count = len(duplicate_row_ids)
    return min(unique_count, total_records)


def _count_unique_invalid_rows(all_annexure, total_records):
    """
    Count unique rows with at least one issue across ALL dimensions.
    Each row is counted once regardless of how many rules it fails.
    """
    if not all_annexure:
        return 0
    invalid_row_ids = set()
    for r in all_annexure:
        try:
            invalid_row_ids.add(int(r.get("Row_Number", -1)))
        except (ValueError, TypeError):
            pass
    invalid_row_ids.discard(-1)
    return min(len(invalid_row_ids), total_records)


def _classify_issues(all_annexure):
    """
    Classify detected issues into user-friendly categories with severity.
    Returns dict: {category: {"count": int, "severity": str, "rows": set}}
    """
    categories = {
        "Missing Values":      {"count": 0, "rows": set(), "rules": set()},
        "Invalid Format":      {"count": 0, "rows": set(), "rules": set()},
        "Domain Violations":   {"count": 0, "rows": set(), "rules": set()},
        "Duplicate Records":   {"count": 0, "rows": set(), "rules": set()},
        "Non-Standard Values": {"count": 0, "rows": set(), "rules": set()},
    }
    rule_category_map = {
        "Not Null": "Missing Values", "Not Empty": "Missing Values",
        "Mandatory Column": "Missing Values", "Whitespace Only": "Missing Values",
        "Email Format": "Invalid Format", "Phone Format": "Invalid Format",
        "PAN Format": "Invalid Format",
        "Custom Regex": "Invalid Format", "Data Type Validation": "Invalid Format",
        "Length Check": "Invalid Format", "Format Check": "Invalid Format",
        "Numeric Range": "Domain Violations", "Allowed Values": "Domain Violations",
        "Minimum Length": "Domain Violations",
        "Single Column Exact Match": "Duplicate Records",
        "Combination Column Exact Match": "Duplicate Records",
        "Hybrid Fuzzy Match": "Duplicate Records",
        "Convert to Lowercase": "Non-Standard Values",
        "Convert to Uppercase": "Non-Standard Values",
        "Special Characters Not Allowed": "Non-Standard Values",
        "Date Format": "Non-Standard Values",
    }
    for r in all_annexure:
        rule = r.get("Rule", r.get("Check", ""))
        cat = rule_category_map.get(rule, "Invalid Format")
        try:
            row_id = int(r.get("Row_Number", -1))
        except (ValueError, TypeError):
            row_id = -1
        categories[cat]["count"] += 1
        categories[cat]["rules"].add(rule)
        if row_id >= 0:
            categories[cat]["rows"].add(row_id)

    # Assign severity based on failure ratio
    result = {}
    for cat, data in categories.items():
        if data["count"] == 0:
            continue
        unique_rows = len(data["rows"])
        if unique_rows > 500 or data["count"] > 1000:
            severity = "HIGH"
        elif unique_rows > 100 or data["count"] > 200:
            severity = "MEDIUM"
        else:
            severity = "LOW"
        result[cat] = {
            "count": data["count"],
            "unique_rows": unique_rows,
            "severity": severity,
            "rules": list(data["rules"]),
        }
    return result


def _build_rule_execution_log(all_annexure, dim_scores, rule_entries=None):
    """
    Build a rule-level execution summary: failures per rule, dimension, config used.
    """
    from collections import Counter
    rule_counter = Counter()
    rule_dim_map = {}
    for r in all_annexure:
        rule = r.get("Rule", r.get("Check", "Unknown"))
        dim = r.get("Dimension", "Unknown")
        rule_counter[rule] += 1
        rule_dim_map[rule] = dim

    log_rows = []
    for rule, count in rule_counter.most_common():
        cfg_str = "—"
        if rule_entries:
            for entry in rule_entries:
                if entry.get("rule") == rule and entry.get("config"):
                    cfg_str = "; ".join(f"{k}={v}" for k, v in entry["config"].items() if v not in ("", None))
                    if cfg_str:
                        break
            if not cfg_str:
                cfg_str = "—"
        log_rows.append({
            "Rule": rule,
            "Dimension": rule_dim_map.get(rule, "—"),
            "Failures": count,
            "Config": cfg_str,
        })
    return log_rows


def _render_dq_score_cards(overall, dim_scores):
    cols=st.columns(len(dim_scores)+1)
    with cols[0]:
        st.markdown(f'<div class="score-card"><div class="val {_score_cls(overall)}">{overall:.1f}%</div><div class="lbl">Overall DQ Score</div></div>',unsafe_allow_html=True)
    for i,(dim,score) in enumerate(dim_scores.items()):
        with cols[i+1]:
            st.markdown(f'<div class="score-card"><div class="val {_score_cls(score)}">{score:.1f}%</div><div class="lbl">{dim}</div></div>',unsafe_allow_html=True)

def _render_dq_results(overall, dim_scores, all_annexure, dup_records, total_records, excel_bytes, obj_name, rule_entries=None):
    _render_dq_score_cards(overall, dim_scores)

    # ── Dataset Overview — accurate counts ──────────────────────────────
    st.markdown("<br>",unsafe_allow_html=True)
    dc = _count_unique_duplicate_rows(dup_records, total_records)
    ic = _count_unique_invalid_rows(all_annexure, total_records)
    qs1,qs2,qs3,qs4=st.columns(4)
    with qs1: st.metric("Total Records",f"{total_records:,}")
    with qs2: st.metric("Records with Issues",f"{ic:,}")
    with qs3: st.metric("Duplicate Records",f"{dc:,}")
    with qs4: st.metric("Total Issues Found",f"{len(all_annexure):,}")

    # ── Validation check banner ─────────────────────────────────────────
    if dc > total_records:
        st.error(f"⚠️ Data integrity warning: Duplicate count ({dc:,}) exceeds total records ({total_records:,}). Please review uniqueness rules.")
    if ic > total_records:
        st.error(f"⚠️ Data integrity warning: Invalid count ({ic:,}) exceeds total records ({total_records:,}). Please review rule configuration.")

    # ── Issue Classification ────────────────────────────────────────────
    issue_classes = _classify_issues(all_annexure)
    if issue_classes:
        st.markdown("### 🏷️ Issue Classification")
        severity_colors = {"HIGH": "#dc2626", "MEDIUM": "#d97706", "LOW": "#059669"}
        ic_cols = st.columns(min(len(issue_classes), 5))
        for i, (cat, info) in enumerate(issue_classes.items()):
            sev_col = severity_colors.get(info["severity"], "#7a7a9a")
            with ic_cols[i % min(len(issue_classes), 5)]:
                st.markdown(f"""
                <div style="background:#fff;border:1.5px solid #d9cef0;border-radius:10px;
                     padding:0.7rem 0.9rem;text-align:center;margin-bottom:0.5rem;">
                    <div style="font-size:1.3rem;font-weight:800;color:#5b2d90;">{info['count']:,}</div>
                    <div style="font-size:0.78rem;font-weight:600;color:#3b1f72;">{cat}</div>
                    <div style="font-size:0.7rem;font-weight:700;color:{sev_col};
                         text-transform:uppercase;letter-spacing:0.04em;">{info['severity']}</div>
                    <div style="font-size:0.68rem;color:#7a7a9a;">{info['unique_rows']:,} unique rows</div>
                </div>""", unsafe_allow_html=True)

    # ── Dimension-wise Invalid Counts ───────────────────────────────────
    if all_annexure and dim_scores:
        annex_df = pd.DataFrame(all_annexure)
        st.markdown("### 📊 Dimension-wise Breakdown")
        dim_summary_rows = []
        for dim in dim_scores:
            ddf = annex_df[annex_df["Dimension"] == dim] if "Dimension" in annex_df.columns else pd.DataFrame()
            dim_issues = len(ddf)
            dim_unique_rows = len(set(int(r) for r in ddf["Row_Number"].dropna())) if not ddf.empty and "Row_Number" in ddf.columns else 0
            dim_summary_rows.append({
                "Dimension": dim,
                "Score": f"{dim_scores[dim]:.1f}%",
                "Total Issues": dim_issues,
                "Unique Rows Affected": min(dim_unique_rows, total_records),
                "Rules Applied": len(ddf["Rule_Applied"].unique()) if not ddf.empty and "Rule_Applied" in ddf.columns else 0,
            })
        st.dataframe(pd.DataFrame(dim_summary_rows), use_container_width=True, hide_index=True)

    # ── Issue Annexures (tabs per dimension + All) ──────────────────────
    if all_annexure or not dup_records.empty:
        st.markdown("### 📋 Issue Annexures")
        annex_df = pd.DataFrame(all_annexure) if all_annexure else pd.DataFrame()
        tab_names = list(dim_scores.keys()) + ["All Issues"]
        tabs = st.tabs(tab_names)
        for i, dim in enumerate(dim_scores.keys()):
            with tabs[i]:
                ddf = annex_df[annex_df["Dimension"] == dim] if not annex_df.empty and "Dimension" in annex_df.columns else pd.DataFrame()
                dim_unique = len(set(int(r) for r in ddf["Row_Number"].dropna())) if not ddf.empty and "Row_Number" in ddf.columns else 0
                st.write(f"**{len(ddf):,}** issues across **{dim_unique:,}** unique rows in {dim}")
                if not ddf.empty:
                    st.dataframe(ddf.head(500), use_container_width=True, height=300)
                # Show duplicates inside the Uniqueness tab
                if dim == "Uniqueness" and not dup_records.empty:
                    st.markdown("#### 🔑 Duplicate Records")
                    st.markdown(f'<p style="font-size:0.84rem;color:#5b2d90;font-weight:600;">'
                                f'{dc:,} unique rows flagged as duplicates out of {total_records:,} total records</p>',
                                unsafe_allow_html=True)
                    st.dataframe(dup_records.head(500), use_container_width=True, height=300)
        with tabs[-1]:
            all_unique = _count_unique_invalid_rows(all_annexure, total_records)
            st.write(f"**{len(annex_df):,}** total issues across **{all_unique:,}** unique rows")
            if not annex_df.empty:
                st.dataframe(annex_df.head(1000), use_container_width=True, height=400)
            if not dup_records.empty:
                st.markdown("#### 🔑 Duplicate Records")
                st.dataframe(dup_records.head(500), use_container_width=True, height=300)

    # ── Rule Execution Log ──────────────────────────────────────────────
    if all_annexure:
        with st.expander("📝 Rule Execution Log — Failures per Rule", expanded=False):
            log_rows = _build_rule_execution_log(all_annexure, dim_scores, rule_entries)
            if log_rows:
                st.dataframe(pd.DataFrame(log_rows), use_container_width=True, hide_index=True)
            else:
                st.info("No rule execution data available.")

    # ── Rule Configuration Summary ──────────────────────────────────────
    if rule_entries:
        with st.expander("⚙️ Rule Configuration Summary", expanded=False):
            cfg_rows = []
            for i, r in enumerate(rule_entries):
                cfg_str = _format_cfg_display(r.get("rule", ""), r.get("config", {}))
                cfg_rows.append({
                    "#": i + 1,
                    "Column": r.get("column", "—"),
                    "Dimension": r.get("dimension", "—"),
                    "Rule": r.get("rule", "—"),
                    "Configuration": cfg_str,
                    "Mandatory": "Yes" if r.get("mandatory") else "No",
                })
            st.dataframe(pd.DataFrame(cfg_rows), use_container_width=True, hide_index=True)

    # ── Download Section ────────────────────────────────────────────────
    st.markdown('<div class="dash-section-header" style="margin-top:1.5rem;"><div class="dash-section-dot"></div><h3>Download Reports</h3><div class="dash-section-accent"></div></div>',unsafe_allow_html=True)
    st.markdown('<div class="dl-card"><div class="dl-card-icon">📊</div><div class="dl-card-title">Excel DQ Report</div><div class="dl-card-desc">Multi-sheet enterprise DQ report with executive summary, dimension scorecard, column profiling, rule failures, row annexures, and standardization details.</div></div>',unsafe_allow_html=True)
    ts=datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    st.download_button("⬇ Download Excel Report",data=excel_bytes,
        file_name=f"DQ_Report_{obj_name}_{ts}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,key="dq_xl_dl")
    st.markdown('<p style="font-size:0.76rem;color:#7a7a9a;margin-top:0.25rem;line-height:1.45;">Includes DQ score, summary dashboard, column analysis, dimension scores, and annexures.</p>',unsafe_allow_html=True)

    st.markdown("<br>",unsafe_allow_html=True)
    _,nc,_=st.columns([1,1.2,1])
    with nc:
        if st.button("📈 Continue to Maturity Assessment →",type="primary",use_container_width=True,key="dq_to_mat"):
            st.session_state["page"]="maturity"; st.rerun()



def _status_banner(msg, pct):
    """Small inline status banner replacing the wide purple progress bar."""
    st.markdown(f"""
    <div style="display:flex;align-items:center;gap:0.9rem;
         background:#f5f0fc;border:1.5px solid #d9cef0;border-radius:12px;
         padding:0.55rem 1rem;margin-bottom:0.6rem;">
        <div style="font-size:0.84rem;font-weight:600;color:#5b2d90;white-space:nowrap;">{msg}</div>
        <div style="flex:1;background:#e8e2f5;border-radius:999px;height:6px;overflow:hidden;">
            <div style="width:{pct}%;height:100%;border-radius:999px;
                 background:linear-gradient(90deg,#5b2d90,#b10f74);
                 transition:width 0.3s ease;"></div>
        </div>
        <div style="font-size:0.78rem;font-weight:700;color:#b10f74;white-space:nowrap;">{pct}%</div>
    </div>""", unsafe_allow_html=True)

def _run_dq_assessment(df, all_columns, selected_dims, rules_by_dim, cfg, obj_name, rule_entries=None):
    status = st.empty()
    def upd(msg, pct):
        with status:
            _status_banner(msg, pct)

    upd("⚙️ Initializing...", 0)
    all_annexure=[]; dim_scores={}; dup_records=pd.DataFrame(); standardized_df=None
    try:
        if selected_dims.get("Completeness") and rules_by_dim.get("Completeness"):
            upd("🔍 Running Completeness checks...", 10)
            ca=execute_completeness_rules(df,rules_by_dim["Completeness"],cfg.get("comp_columns",all_columns),min_length_val=cfg.get("min_length_val",3),mandatory_cols=cfg.get("mandatory_cols"))
            all_annexure.extend(ca)
            dim_scores["Completeness"]=compute_completeness_score(df,ca,cfg.get("comp_columns",all_columns),rules_by_dim["Completeness"])

        if selected_dims.get("Validity") and rules_by_dim.get("Validity"):
            upd("✅ Running Validity checks...", 30)
            # Build column_rule_map from rule_entries so per-column configs (e.g. Allowed Values) are used
            val_column_rule_map = None
            if rule_entries:
                val_column_rule_map = [
                    {"column": r["column"], "rule": r["rule"], "config": r.get("config", {})}
                    for r in rule_entries
                    if r.get("dimension") == "Validity"
                ]
            va=execute_validity_rules(
                df,
                rules_by_dim["Validity"],
                cfg.get("val_columns", all_columns),
                range_min=cfg.get("range_min", 0),
                range_max=cfg.get("range_max", 100),
                allowed_values_str=cfg.get("allowed_values_str", ""),
                custom_regex=cfg.get("custom_regex", ""),
                date_fmt=cfg.get("date_fmt", ""),
                column_rule_map=val_column_rule_map if val_column_rule_map else None,
            )
            all_annexure.extend(va)
            dim_scores["Validity"]=compute_validity_score(df,va,cfg.get("val_columns",all_columns),rules_by_dim["Validity"])

        if selected_dims.get("Uniqueness") and rules_by_dim.get("Uniqueness"):
            upd("🔑 Running Uniqueness checks...", 50)
            dup_records,ua,warnings=execute_uniqueness_rules(
                df, rules_by_dim["Uniqueness"],
                single_cols=cfg.get("single_dup_cols") or ([cfg["single_dup_col"]] if cfg.get("single_dup_col") else []),
                combo_cols=cfg.get("combo_dup_cols"),
                fuzzy_cols=cfg.get("fuzzy_cols"),
                fuzzy_threshold=cfg.get("fuzzy_threshold",80),
                fuzzy_weights=cfg.get("fuzzy_weights"),
                fuzzy_max_pairs=cfg.get("fuzzy_max_pairs", 20_000),
                fuzzy_ignore_nulls=cfg.get("fuzzy_ignore_nulls", True),
            )
            # ── Deduplicate dup_records: ensure unique rows only ──────────
            if not dup_records.empty:
                dup_records = dup_records.loc[~dup_records.index.duplicated(keep='first')]
                # Safety cap: never exceed total records
                if len(dup_records) > len(df):
                    dup_records = dup_records.iloc[:len(df)]
            all_annexure.extend(ua)
            dim_scores["Uniqueness"]=compute_uniqueness_score(df,dup_records)

        if selected_dims.get("Standardization") and rules_by_dim.get("Standardization"):
            upd("🔧 Running Standardization checks...", 70)
            std_column_rule_map = None
            if rule_entries:
                std_column_rule_map = [
                    {"column": r["column"], "rule": r["rule"], "config": r.get("config", {})}
                    for r in rule_entries
                    if r.get("dimension") == "Standardization"
                ]
            standardized_df,sa=execute_standardization_rules(
                df,
                rules_by_dim["Standardization"],
                cfg.get("std_columns", all_columns),
                date_target_fmt=cfg.get("date_target_fmt", "%Y-%m-%d"),
                null_default=cfg.get("null_default", "N/A"),
                column_rule_map=std_column_rule_map if std_column_rule_map else None,
            )
            all_annexure.extend(sa)
            dim_scores["Standardization"]=compute_standardization_score(df,sa,cfg.get("std_columns",all_columns),rules_by_dim["Standardization"])

        upd("📊 Computing scores...", 85)
        overall=compute_overall_score(dim_scores)
        upd("🧹 Building clean dataset...", 90)
        clean_df=build_clean_dataset(df,standardized_df,dup_records,all_annexure)
        upd("📥 Generating Excel report...", 95)
        ucfg={"fuzzy_threshold":cfg.get("fuzzy_threshold","N/A")}
        # Pass per-column configs for report traceability
        if rule_entries:
            ucfg["rule_entries"] = rule_entries
            ucfg["per_column_configs"] = {}
            for entry in rule_entries:
                col = entry.get("column", "")
                rule = entry.get("rule", "")
                entry_cfg = entry.get("config", {})
                if entry_cfg and any(v not in ("", None) for v in entry_cfg.values()):
                    key = f"{col}|{rule}"
                    ucfg["per_column_configs"][key] = entry_cfg
        excel_bytes=generate_excel_report(df,clean_df,dup_records,all_annexure,dim_scores,overall,[d for d in selected_dims if selected_dims[d]],ucfg)
        upd("✅ Assessment complete!", 100)
        status.empty()

        st.session_state.update({"dq_score":overall,"dq_dim_scores":dim_scores,
            "dq_object_name":obj_name or "Customer","dq_excel_bytes":excel_bytes,
            "dq_annexure":all_annexure,"dq_dup_records":dup_records,"dq_total_records":len(df),
            "dq_unique_dup_count":_count_unique_duplicate_rows(dup_records, len(df)),
            "dq_unique_invalid_count":_count_unique_invalid_rows(all_annexure, len(df)),
            "dq_rule_entries_snapshot": rule_entries or []})
        st.session_state["mat_objects"]=[obj_name] if obj_name else DEFAULT_MASTER_OBJECTS[:]
        autofill_dq_dimension(overall)
        _render_dq_results(overall,dim_scores,all_annexure,dup_records,len(df),excel_bytes,obj_name or "Dataset", rule_entries=rule_entries)

    except Exception as e:
        status.empty()
        st.error(f"❌ Error during assessment: {e}")
        with st.expander("Technical Details"): st.code(traceback.format_exc())


def page_dq():
    _page_banner(
        icon="🔍",
        badge_text="Enterprise DQ Engine",
        title="Data Quality Assessment",
        subtitle="Upload your dataset, configure business rules across all dimensions, and generate comprehensive DQ scores with annexure-based reporting.",
        gradient="135deg,#3d1d63 0%,#5b2d90 55%,#b10f74 100%"
    )
    _page_nav("dq")

    has_results=st.session_state.get("dq_score") is not None
    st.markdown('<div class="dash-section-header"><div class="dash-section-dot"></div><h3>📂 Step 1 — Upload Dataset</h3><div class="dash-section-accent"></div></div>',unsafe_allow_html=True)
    uploaded_file=st.file_uploader("Upload CSV or Excel file",type=SUPPORTED_FORMATS,key="dq_data_upload",label_visibility="collapsed")

    if has_results and not uploaded_file:
        _render_dq_results(
            st.session_state["dq_score"],st.session_state.get("dq_dim_scores",{}),
            st.session_state.get("dq_annexure",[]),st.session_state.get("dq_dup_records",pd.DataFrame()),
            st.session_state.get("dq_total_records",0),st.session_state.get("dq_excel_bytes",b""),
            st.session_state.get("dq_object_name","Dataset"),
            rule_entries=st.session_state.get("dq_rule_entries_snapshot"))
        return

    if not uploaded_file:
        st.info("👆 Upload a CSV or Excel file to begin the assessment.")
        st.markdown("""
        <div class="dq-steps-row">
            <div class="dq-step active"><div class="dq-step-num">01</div><div class="dq-step-icon">📤</div><div class="dq-step-title">Upload Dataset</div><div class="dq-step-desc">CSV or Excel file</div></div>
            <div class="dq-step-line"></div>
            <div class="dq-step"><div class="dq-step-num">02</div><div class="dq-step-icon">⚙️</div><div class="dq-step-title">Configure Rules</div><div class="dq-step-desc">Business rules &amp; duplicates</div></div>
            <div class="dq-step-line"></div>
            <div class="dq-step"><div class="dq-step-num">03</div><div class="dq-step-icon">📊</div><div class="dq-step-title">Download Report</div><div class="dq-step-desc">Excel with annexures</div></div>
        </div>""", unsafe_allow_html=True)
        return

    sheet_name=None
    if uploaded_file.name.lower().endswith((".xlsx",".xls",".xlsm")):
        sheets=get_excel_sheet_names(uploaded_file)
        if len(sheets)>1: sheet_name=st.selectbox("Select Sheet",sheets,key="dq_sheet")

    try: df=load_dataset(uploaded_file,sheet_name)
    except Exception as e: st.error(f"❌ Failed to load file: {e}"); return

    all_columns=list(df.columns)
    st.success(f"✅ Loaded **{len(df):,}** rows × **{len(all_columns)}** columns")
    with st.expander("🔍 Preview Data (first 10 rows)",expanded=False):
        st.dataframe(df.head(10),use_container_width=True,height=150)

    obj_name=st.session_state.get("dq_object_name","Customer")

    # ── All dimensions enabled automatically — no checkboxes ──
    selected_dims = {dim: True for dim in DIMENSIONS}

    # ── Active dimensions indicator ──
    st.markdown(f"""
    <div style="background:linear-gradient(135deg,#f5f0fc,#fdf2f8);border:1.5px solid #d9cef0;
         border-radius:12px;padding:0.7rem 1.2rem;margin:1rem 0 0.5rem;
         display:flex;align-items:center;gap:0.8rem;flex-wrap:wrap;">
        <span style="font-size:0.82rem;font-weight:700;color:#3b1f72;">Active Dimensions:</span>
        {"".join(f'<span style="background:linear-gradient(135deg,#5b2d90,#7c4dbb);color:#fff;font-size:0.78rem;font-weight:700;padding:0.3rem 0.85rem;border-radius:999px;box-shadow:0 2px 6px rgba(91,45,144,0.2);">✓ {dim}</span>' for dim in DIMENSIONS)}
    </div>""", unsafe_allow_html=True)

    # ── Step 2: Business Rule Criteria Builder (Completeness + Validity + Standardization) ──
    st.markdown('<div class="dash-section-header" style="margin-top:1rem;"><div class="dash-section-dot magenta"></div><h3>⚙️ Step 2 — Dynamic Business Rule Criteria Builder</h3><div class="dash-section-accent"></div></div>',unsafe_allow_html=True)

    rules_by_dim = {}
    merged_cfg   = {}

    rules_by_dim_from_builder, cfg_from_builder = _build_dq_criteria_builder(all_columns, df)
    rules_by_dim.update(rules_by_dim_from_builder)
    merged_cfg.update(cfg_from_builder)

    # ── Uniqueness builder always shown ──
    r, c = _ui_uniqueness(all_columns)
    rules_by_dim["Uniqueness"] = r
    merged_cfg.update(c)

    st.markdown("---")
    _,center,_=st.columns([1,1.5,1])
    with center:
        run=st.button("🚀 Run Data Quality Assessment",type="primary",use_container_width=True,key="dq_run")
    if run:
        _run_dq_assessment(df,all_columns,selected_dims,rules_by_dim,merged_cfg,obj_name,
                           rule_entries=st.session_state.get("dq_rule_entries",[]))


# ══════════════════════════════════════════════════════════════════════════════
#  PAGE: MATURITY ASSESSMENT
# ══════════════════════════════════════════════════════════════════════════════
def _apply_editor_edits(dim, ek):
    ws=st.session_state.get(ek)
    if not ws: return
    er=ws.get("edited_rows",{})
    if not er: return
    df=st.session_state.mat_responses[dim].copy()
    for ri,changes in er.items():
        for col,val in changes.items(): df.at[int(ri),col]=val
    st.session_state.mat_responses[dim]=df

def _do_submit():
    cn=st.session_state.get("mat_client_name","").strip()
    if not cn:
        # Show a modal dialog for missing client name
        _show_client_name_dialog()
        return
    objects=st.session_state.mat_objects; dims=st.session_state.mat_dims
    responses=st.session_state.mat_responses
    bm=float(st.session_state.mat_benchmark); tg=float(st.session_state.mat_target)
    lt=float(st.session_state.mat_low_thr); dq_score=st.session_state.get("dq_score")
    ok,msg=validate_responses(responses,dims,objects)
    if not ok: st.error(f"⚠️ Validation failed: {msg}"); return
    with st.spinner("⚙️ Computing scores and building reports…"):
        dim_table,overall=compute_all_scores(objects,dims,responses)
        domain_display={dim:float(np.nanmean(dim_table.loc[dim].values)) for dim in dims}
        exec_score=float(np.nanmean(overall.values)) if len(overall) else 0.0
        slide_png=render_slide_png(client_name=cn,domain_scores=domain_display,
            exec_score=exec_score if np.isfinite(exec_score) else 0.0,benchmark=bm,target=tg)
        # ── Sanitise detail tables for PDF: convert all values to string & trim long text ──
        pdf_detail_tables = {}
        for d_key, d_df in responses.items():
            safe_df = d_df.copy()
            for col in safe_df.columns:
                safe_df[col] = safe_df[col].astype(str).str.slice(0, 120)
            pdf_detail_tables[d_key] = safe_df
        # Sanitise dim_table and overall for PDF
        pdf_dim_table = dim_table.copy()
        for col in pdf_dim_table.columns:
            pdf_dim_table[col] = pdf_dim_table[col].apply(lambda v: str(v)[:60] if pd.notna(v) else "")
        pdf_overall = {k: v for k, v in overall.items()}
        pdf_bytes=build_pdf_bytes(client_name=cn,slide_png=slide_png,dim_table=pdf_dim_table,
            overall=pdf_overall,detail_tables=pdf_detail_tables,dq_score=dq_score)
        mat_excel=to_excel_bytes(dim_table=dim_table,overall=overall,detail_tables=responses,low_thr=lt,objects=objects)
    st.session_state["mat_submitted"]=True
    st.session_state["mat_payload"]={"dim_table":dim_table,"overall":overall,"slide_png":slide_png,
        "mat_excel":mat_excel,"pdf_bytes":pdf_bytes,
        "client_name":cn,"ts":datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}
    st.rerun()


@st.dialog("⚠️ Client Name Required")
def _show_client_name_dialog():
    """Modal dialog blocking submission until Client Name is provided."""
    st.markdown(
        '<div style="text-align:center;padding:0.5rem 0;">'
        '<div style="font-size:2.5rem;margin-bottom:0.5rem;">🏢</div>'
        '<p style="font-size:1rem;color:#1a1a2e;font-weight:600;margin-bottom:0.3rem;">'
        'Please enter Client Name before submitting.</p>'
        '<p style="font-size:0.85rem;color:#7a7a9a;line-height:1.5;">'
        'Navigate to the <strong>sidebar Configuration panel</strong> and enter the '
        'organisation name in the <strong>Client Name</strong> field.</p>'
        '</div>',
        unsafe_allow_html=True,
    )
    if st.button("✅ OK, I'll fill it in", use_container_width=True, type="primary"):
        st.rerun()

def page_maturity():
    inject_gdg_light()
    if "mat_dims" not in st.session_state or not isinstance(st.session_state.get("mat_dims"),list): st.session_state["mat_dims"]=list(MATURITY_DIMS)
    if "mat_objects" not in st.session_state or not isinstance(st.session_state.get("mat_objects"),list): st.session_state["mat_objects"]=list(DEFAULT_MASTER_OBJECTS)
    if "mat_responses" not in st.session_state: st.session_state["mat_responses"]={}
    dq_score=st.session_state.get("dq_score"); submitted=st.session_state.get("mat_submitted",False)

    with st.sidebar:
        st.markdown('<div style="font-weight:700;font-size:0.88rem;color:#3b1f72;margin-bottom:0.4rem;">⚙️ Configuration</div>',unsafe_allow_html=True)
        st.markdown('<span style="font-size:0.82rem;font-weight:600;color:#1a1a2e;">Client Name <span style="color:#dc2626;">*</span></span>',unsafe_allow_html=True)
        cn_val=st.text_input("Client Name",value=st.session_state.get("mat_client_name",""),placeholder="Organisation name (required)",disabled=submitted,label_visibility="collapsed")
        st.session_state["mat_client_name"]=cn_val
        if not cn_val.strip() and not submitted:
            st.markdown('<div style="color:#dc2626;font-size:0.75rem;margin-top:-0.3rem;">⚠ Required</div>',unsafe_allow_html=True)

        st.divider()

        # ── Masters Applicable — prominently visible checkbox ──
        st.markdown(
            '<div style="background:linear-gradient(135deg,#f5f0fc,#fdf2f8);'
            'border:2px solid #b09dd6;border-radius:12px;'
            'padding:0.7rem 0.9rem;margin-bottom:0.7rem;'
            'box-shadow:0 2px 10px rgba(91,45,144,0.12);">'
            '<div style="display:flex;align-items:center;gap:0.5rem;margin-bottom:0.3rem;">'
            '<span style="font-size:1.1rem;">📋</span>'
            '<span style="font-size:0.88rem;font-weight:800;color:#3b1f72;">Masters Applicable</span>'
            '</div>'
            '<span style="font-size:0.76rem;color:#5b2d90;line-height:1.4;">'
            'Enable to evaluate maturity per Master Data Object</span>'
            '</div>',
            unsafe_allow_html=True,
        )
        masters_applicable = st.checkbox(
            "✅ Include Master Data Objects in assessment",
            value=st.session_state.get("mat_masters_applicable", True),
            key="mat_masters_applicable_cb",
            disabled=submitted,
            help="When enabled, maturity questions are evaluated per Master Data Object. Uncheck to use a single overall rating.",
        )
        st.session_state["mat_masters_applicable"] = masters_applicable

        if masters_applicable:
            all_obj_opts=list(dict.fromkeys(DEFAULT_MASTER_OBJECTS+st.session_state.mat_objects))
            st.session_state["mat_objects"]=st.multiselect("Master Data Objects",options=all_obj_opts,default=st.session_state.mat_objects,disabled=submitted)
        else:
            st.session_state["mat_objects"] = [cn_val.strip() or "Overall"]
            st.caption("Assessment will use a single overall rating column.")

        st.session_state["mat_dims"]=st.multiselect("Maturity Dimensions",options=MATURITY_DIMS,default=st.session_state.mat_dims,disabled=submitted)
        st.divider()
        st.session_state["mat_low_thr"]=st.slider("Exception threshold (≤)",1.0,5.0,float(st.session_state.get("mat_low_thr",2.0)),0.5,disabled=submitted)
        st.divider()
        st.session_state["mat_benchmark"]=st.number_input("Industry Benchmark",1.0,5.0,float(st.session_state.get("mat_benchmark",3.0)),0.1,disabled=submitted)
        st.session_state["mat_target"]=st.number_input("Target Score",1.0,5.0,float(st.session_state.get("mat_target",3.0)),0.1,disabled=submitted)

    if not st.session_state.mat_objects or not st.session_state.mat_dims:
        st.info("👉 Select at least one **Object** and one **Dimension** in the sidebar."); st.stop()

    prev_objs=st.session_state.get("_last_sync_objects"); prev_dims=st.session_state.get("_last_sync_dims")
    curr_objs=st.session_state.mat_objects; curr_dims=st.session_state.mat_dims
    needs_sync=(prev_objs is None or prev_dims is None or set(prev_objs)!=set(curr_objs) or set(prev_dims)!=set(curr_dims))
    if needs_sync:
        if st.session_state.get("_sync_pending"):
            sync_response_tables()
            for d in curr_dims: st.session_state.pop(f"mat_snap_{d}",None)
            st.session_state["_last_sync_objects"]=list(curr_objs); st.session_state["_last_sync_dims"]=list(curr_dims)
            st.session_state["_sync_pending"]=False
        else: st.session_state["_sync_pending"]=True; st.rerun()

    if dq_score is not None and not st.session_state.get("dq_autofilled"):
        autofill_dq_dimension(dq_score); st.session_state.pop("mat_snap_Data Quality",None); st.session_state["dq_autofilled"]=True

    if submitted and st.session_state.get("mat_payload"):
        p=st.session_state["mat_payload"]; cn=p["client_name"]
        hl,hs,hr=st.columns([6,1,1.4])
        with hr:
            st.markdown('<div class="mat-edit-btn-wrap">',unsafe_allow_html=True)
            if st.button("✏️ Edit Responses",key="mat_edit_top",use_container_width=True,type="primary"):
                st.session_state["mat_submitted"]=False; st.session_state["mat_payload"]={}; st.rerun()
            st.markdown('</div>',unsafe_allow_html=True)

        _page_banner(
            icon="✅",
            badge_text="Maturity Report",
            title="Data Maturity Assessment Report",
            subtitle="Executive summary of DAMA maturity dimensions with benchmark comparisons, domain scoring and downloadable reports.",
            gradient="135deg,#3d1d63 0%,#5b2d90 55%,#b10f74 100%"
        )
        _page_nav("maturity")
        if dq_score is not None:
            lvl=dq_score_to_maturity_level(dq_score)
            st.markdown(f'<div class="banner success"><strong>DQ Engine Score:</strong> {dq_score:.1f}% → <strong>Level:</strong> {lvl}</div>',unsafe_allow_html=True)

        st.markdown('<h3 style="margin-top:1.2rem;">📊 Summary Slide</h3>', unsafe_allow_html=True)
        import base64 as _b64
        _s64 = _b64.b64encode(p["slide_png"]).decode()
        st.markdown(f"""
<style>
.mat-slide-wrap {{
    width: 100%;
    display: block;
    margin: 0;
    padding: 0;
    box-sizing: border-box;
    border-radius: 14px;
    overflow: hidden;
    box-shadow: 0 6px 32px rgba(91,45,144,0.16);
    border: 1.5px solid #e8e2f5;
    background: #fff;
}}
.mat-slide-wrap img {{
    width: 100%;
    height: auto;
    display: block;
    margin: 0;
    padding: 0;
    border: none;
}}
</style>
<div class="mat-slide-wrap">
  <img src="data:image/png;base64,{_s64}" alt="Data Maturity Summary Slide" />
</div>
""", unsafe_allow_html=True)
        st.divider()

        st.markdown('<div class="mat-table-panel"><div class="mat-table-title">📐 Dimension-wise Maturity</div>',unsafe_allow_html=True)
        nr=p["dim_table"].shape[0]
        sd=(p["dim_table"].style.format("{:.2f}").background_gradient(cmap="Blues",axis=None,vmin=1,vmax=5)
            .set_table_styles([{"selector":"table","props":[("width","100%")]},{"selector":"td,th","props":[("text-align","center"),("padding","6px 10px")]}]))
        st.dataframe(sd,use_container_width=True,height=min(36*(nr+2),420))
        st.markdown('</div>',unsafe_allow_html=True)

        st.markdown('<div class="mat-table-panel"><div class="mat-table-title">🏆 Overall Maturity Score</div>',unsafe_allow_html=True)
        ov_df=pd.DataFrame(p["overall"]).T
        so=(ov_df.style.format("{:.2f}").background_gradient(cmap="Blues",axis=None,vmin=1,vmax=5)
            .set_table_styles([{"selector":"table","props":[("width","100%")]},{"selector":"td,th","props":[("text-align","center"),("padding","8px 12px")]}]))
        st.dataframe(so,use_container_width=True,height=80)
        st.markdown('</div>',unsafe_allow_html=True)

        st.divider()
        dim_vals={dim:float(np.nanmean(p["dim_table"].loc[dim].values)) for dim in p["dim_table"].index}
        bar_img=_mat_bar_png(dim_vals)
        if bar_img: st.image(bar_img,use_container_width=True)
        st.divider()

        st.markdown('<h3>📥 Download Reports</h3>',unsafe_allow_html=True)
        safe_cn=cn.replace(" ","_"); d1,d2=st.columns(2)
        with d1:
            st.markdown('<div class="dl-card"><div class="dl-card-icon">📄</div><div class="dl-card-title">PDF Maturity Report</div><div class="dl-card-desc">Executive maturity summary slide, domain scoring, and detailed question-level scoring.</div></div>',unsafe_allow_html=True)
            st.download_button("⬇ Download PDF",data=p["pdf_bytes"],file_name=get_timestamp_filename(f"Maturity_{safe_cn}","pdf"),mime="application/pdf",use_container_width=True,key="mat_pdf_dl")
            st.markdown('<p style="font-size:0.76rem;color:#7a7a9a;margin-top:0.25rem;line-height:1.45;">Executive maturity summary slide, domain scoring, and detailed question-level scoring.</p>',unsafe_allow_html=True)
        with d2:
            st.markdown('<div class="dl-card"><div class="dl-card-icon">📊</div><div class="dl-card-title">Excel Workbook</div><div class="dl-card-desc">Multi-sheet: dimension summary, overall maturity, question responses, exception sheets.</div></div>',unsafe_allow_html=True)
            st.download_button("⬇ Download Excel",data=p["mat_excel"],file_name=get_timestamp_filename(f"Maturity_{safe_cn}","xlsx"),mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True,key="mat_xl_dl")
            st.markdown('<p style="font-size:0.76rem;color:#7a7a9a;margin-top:0.25rem;line-height:1.45;">Includes DQ score, summary dashboard, column analysis, dimension scores, and annexures.</p>',unsafe_allow_html=True)
        st.stop()

    # ── Questionnaire ──
    _page_banner(
        icon="📈",
        badge_text="DAMA Framework",
        title="Data Maturity Assessment",
        subtitle="Evaluate maturity across DAMA dimensions — governance, quality, architecture, integration &amp; privacy — with weighted scoring and benchmarking.",
        gradient="135deg,#3d1d63 0%,#5b2d90 55%,#b10f74 100%"
    )
    _page_nav("maturity")

    # ── Download Template Button (top of questionnaire) ──────────────────
    def _build_bulk_template() -> bytes:
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        dims_   = st.session_state.mat_dims
        objects_= st.session_state.mat_objects

        hdr_font   = Font(name="Aptos", bold=True, size=10)
        body_font  = Font(name="Aptos", size=10)
        center_aln = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_aln   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        thin_border = Border(
            left  =Side(style="thin", color="B0B0B0"),
            right =Side(style="thin", color="B0B0B0"),
            top   =Side(style="thin", color="B0B0B0"),
            bottom=Side(style="thin", color="B0B0B0"),
        )

        FIXED_COLS = ["Question ID", "Section", "Question", "Weight"]

        for dim in dims_:
            ws = wb.create_sheet(title=dim[:31])
            questions = QUESTION_BANK.get(dim, [])

            all_cols = FIXED_COLS + objects_

            # Row 1 — plain instruction text (no highlight)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(all_cols))
            inst_cell = ws.cell(row=1, column=1)
            inst_cell.value = (
                f"Dimension: {dim} | "
                f"Allowed values: {', '.join(RATING_LABELS)}. "
                "Do NOT modify Question ID, Section, Question or Weight columns."
            )
            inst_cell.font      = Font(name="Aptos", italic=True, color="666666", size=9)
            inst_cell.alignment = left_aln

            # Row 2 — plain bold headers (no fill)
            for ci, col_name in enumerate(all_cols, 1):
                cell = ws.cell(row=2, column=ci, value=col_name)
                cell.font      = hdr_font
                cell.alignment = center_aln
                cell.border    = thin_border

            # Dropdown validation for rating columns
            dv = DataValidation(
                type="list",
                formula1=f'"{",".join(RATING_LABELS)}"',
                showDropDown=False,
                showErrorMessage=True,
                errorTitle="Invalid Rating",
                error=f"Please choose from: {', '.join(RATING_LABELS)}",
            )
            ws.add_data_validation(dv)

            # Data rows — plain, no fills
            for ri, q in enumerate(questions, 3):
                q_idx = ri - 3
                for ci, col_name in enumerate(FIXED_COLS, 1):
                    val = q.get({"Question ID": "id", "Section": "section",
                                 "Question": "question", "Weight": "weight"}.get(col_name, ""), "")
                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.font      = body_font
                    cell.alignment = left_aln if col_name == "Question" else center_aln
                    cell.border    = thin_border

                # Pull current rating from grid if available
                dim_df = st.session_state.get("mat_responses", {}).get(dim)
                for oi, obj in enumerate(objects_, len(FIXED_COLS) + 1):
                    current_rating = RATING_LABELS[0]
                    if dim_df is not None and obj in dim_df.columns and q_idx < len(dim_df):
                        cell_val = dim_df.iloc[q_idx].get(obj, RATING_LABELS[0])
                        if cell_val in RATING_LABELS:
                            current_rating = cell_val
                    cell = ws.cell(row=ri, column=oi, value=current_rating)
                    cell.font      = body_font
                    cell.alignment = center_aln
                    cell.border    = thin_border
                    dv.add(cell)

            # Column widths
            col_widths = {"Question ID": 14, "Section": 22, "Question": 55, "Weight": 8}
            for ci, col_name in enumerate(all_cols, 1):
                ltr = get_column_letter(ci)
                ws.column_dimensions[ltr].width = col_widths.get(col_name, 18)

            ws.freeze_panes = "A3"

        # Reference sheet — plain
        ref_ws = wb.create_sheet(title="Valid Ratings")
        ref_ws.cell(row=1, column=1, value="Allowed Rating Values").font = Font(name="Aptos", bold=True, size=11)
        ref_ws.cell(row=2, column=1, value="Score").font = Font(name="Aptos", bold=True)
        ref_ws.cell(row=2, column=2, value="Label").font = Font(name="Aptos", bold=True)
        for i, lbl in enumerate(RATING_LABELS, 1):
            ref_ws.cell(row=i + 2, column=1, value=i)
            ref_ws.cell(row=i + 2, column=2, value=lbl)
        ref_ws.column_dimensions["A"].width = 10
        ref_ws.column_dimensions["B"].width = 20

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── Template download bar ──────────────────────────────────────────
    st.markdown("""
    <div style="background:linear-gradient(135deg,#f5f0fc,#fdf2f8);border:1.5px solid #d9cef0;
         border-radius:14px;padding:1rem 1.4rem 0.8rem;margin-bottom:1rem;">
        <div style="display:flex;align-items:center;gap:0.6rem;margin-bottom:0.3rem;">
            <span style="font-size:1.3rem;">✏️</span>
            <span style="font-size:1rem;font-weight:800;color:#3b1f72;">Assessment Entry</span>
        </div>
        <p style="font-size:0.85rem;color:#5b2d90;margin:0;line-height:1.5;">
            Rate each question per Master Data Object using the dropdown selectors in the grid below,
            or use <strong>Upload Template</strong> to fill offline and import responses.
        </p>
    </div>""", unsafe_allow_html=True)

    # Download template button - always visible above the tabs
    sync_response_tables()

    # ── Tabs: Manual Entry | Upload Template ─────────────────────────
    mat_tab_manual, mat_tab_upload = st.tabs([
        "✏️ Manual Entry", "📤 Upload Template"
    ])

    # ══════════════════════════════════════════════════════════════════
    #  TAB 1 — MANUAL ENTRY (dimension tabs with grid editors)
    # ══════════════════════════════════════════════════════════════════
    with mat_tab_manual:
        # ── Dimension Tabs + Editors ──────────────────────────────────────
        dims=st.session_state.mat_dims; tabs=st.tabs(dims)
        for i,dim in enumerate(dims):
            with tabs[i]:
                ca,cr,cf,_=st.columns([1,1,1,3])
                with ca:
                    if st.button("➕ Add Question",key=f"add_{dim}",use_container_width=True):
                        df_=st.session_state.mat_responses[dim].copy()
                        nr_={**{"Question ID":f"CQ-{len(df_)+1}","Section":"Custom","Question":"Enter question text…","Weight":1.0},**{obj:RATING_LABELS[0] for obj in st.session_state.mat_objects}}
                        st.session_state.mat_responses[dim]=pd.concat([df_,pd.DataFrame([nr_])],ignore_index=True); st.rerun()
                with cr:
                    if st.button("➖ Remove Last",key=f"rem_{dim}",use_container_width=True):
                        df_=st.session_state.mat_responses[dim]
                        if len(df_)>1: st.session_state.mat_responses[dim]=df_.iloc[:-1].reset_index(drop=True); st.rerun()
                with cf:
                    fvk=f"fullview_{dim}"; is_fv=st.session_state.get(fvk,False)
                    if st.button("⛶ Full View" if not is_fv else "◱ Compact",key=f"fv_{dim}",use_container_width=True):
                        st.session_state[fvk]=not is_fv; st.rerun()

                col_cfg={
                    "Question ID": st.column_config.TextColumn("Question ID", disabled=True),
                    "Section": st.column_config.TextColumn("Section"),
                    "Question": st.column_config.TextColumn("Question", width="large"),
                    "Weight": st.column_config.NumberColumn("Weight", min_value=0.0, step=0.5),
                }
                for obj in st.session_state.mat_objects:
                    col_cfg[obj]=st.column_config.SelectboxColumn(obj,options=RATING_LABELS,required=True)
                ek=f"mat_editor_{dim}"; nrows=len(st.session_state.mat_responses[dim])
                gh = max(38*(nrows+2), 200) if st.session_state.get(f"fullview_{dim}") else min(max(38*(nrows+2), 200), 500)

                with st.container():
                    st.data_editor(
                        st.session_state.mat_responses[dim],
                        use_container_width=True,
                        hide_index=True,
                        column_config=col_cfg,
                        disabled=["Question ID"],
                        key=ek,
                        on_change=_apply_editor_edits,
                        args=(dim,ek),
                        height=gh,
                    )

    # ══════════════════════════════════════════════════════════════════
    #  TAB 2 — UPLOAD TEMPLATE (import filled Excel back into grid)
    # ══════════════════════════════════════════════════════════════════
    with mat_tab_upload:
        st.markdown(
            '<p style="font-size:0.85rem;color:#555;margin-bottom:0.5rem;line-height:1.5;">'
            'Download the Excel template, fill in ratings offline, then upload the completed file here '
            'to import all responses into the assessment grid.</p>',
            unsafe_allow_html=True,
        )

        # Step 1: Download template
        st.markdown(
            '<div style="background:#f9f8fc;border:1.5px solid #e8e2f5;border-radius:10px;'
            'padding:0.7rem 1rem;margin-bottom:0.8rem;">'
            '<span style="font-size:0.85rem;font-weight:700;color:#3b1f72;">Step 1:</span> '
            '<span style="font-size:0.84rem;color:#555;">Download the template with current questions and objects</span>'
            '</div>',
            unsafe_allow_html=True,
        )
        _ul_c1, _ul_c2, _ = st.columns([1.5, 1.5, 2])
        with _ul_c1:
            try:
                tmpl_bytes_ul = _build_bulk_template()
                ts_ul = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                st.download_button(
                    "⬇ Download Excel Template",
                    data=tmpl_bytes_ul,
                    file_name=f"Maturity_Template_{ts_ul}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="mat_tmpl_dl_upload_tab",
                )
            except Exception as e:
                st.warning(f"Template unavailable: {e}")

        # Step 2: Upload filled template
        st.markdown(
            '<div style="background:#f9f8fc;border:1.5px solid #e8e2f5;border-radius:10px;'
            'padding:0.7rem 1rem;margin-bottom:0.8rem;">'
            '<span style="font-size:0.85rem;font-weight:700;color:#3b1f72;">Step 2:</span> '
            '<span style="font-size:0.84rem;color:#555;">Upload the filled template to import responses</span>'
            '</div>',
            unsafe_allow_html=True,
        )
        uploaded_tmpl = st.file_uploader(
            "Upload completed Maturity Template (.xlsx)",
            type=["xlsx", "xls"],
            key="mat_tmpl_upload",
            label_visibility="collapsed",
        )

        if uploaded_tmpl is not None:
            try:
                import openpyxl as _opx
                wb = _opx.load_workbook(uploaded_tmpl, data_only=True)

                dims_ = st.session_state.mat_dims
                objects_ = st.session_state.mat_objects
                FIXED_COLS = ["Question ID", "Section", "Question", "Weight"]
                imported_dims = []
                skipped_dims = []
                total_cells_imported = 0
                total_new_rows = 0

                for dim in dims_:
                    sheet_name = dim[:31]
                    if sheet_name not in wb.sheetnames:
                        skipped_dims.append(dim)
                        continue

                    ws = wb[sheet_name]
                    # Read header row (row 2) to find column positions
                    headers = {}
                    for col_idx in range(1, ws.max_column + 1):
                        val = ws.cell(row=2, column=col_idx).value
                        if val:
                            headers[str(val).strip()] = col_idx

                    # Validate required fixed columns exist
                    if "Question ID" not in headers:
                        skipped_dims.append(dim)
                        continue

                    # Read ALL rows from the uploaded sheet (existing + new)
                    uploaded_rows = []
                    for row_idx in range(3, ws.max_row + 1):
                        qid = ws.cell(row=row_idx, column=headers.get("Question ID", 1)).value
                        if not qid:
                            continue
                        qid = str(qid).strip()
                        if not qid:
                            continue

                        row_data = {"Question ID": qid}

                        # Read fixed columns
                        if "Section" in headers:
                            v = ws.cell(row=row_idx, column=headers["Section"]).value
                            row_data["Section"] = str(v).strip() if v else ""
                        if "Question" in headers:
                            v = ws.cell(row=row_idx, column=headers["Question"]).value
                            row_data["Question"] = str(v).strip() if v else ""
                        if "Weight" in headers:
                            v = ws.cell(row=row_idx, column=headers["Weight"]).value
                            try:
                                row_data["Weight"] = float(v) if v is not None else 1.0
                            except (ValueError, TypeError):
                                row_data["Weight"] = 1.0

                        # Read object rating columns
                        for obj in objects_:
                            if obj in headers:
                                cell_val = ws.cell(row=row_idx, column=headers[obj]).value
                                if cell_val is not None:
                                    cell_str = str(cell_val).strip()
                                    if cell_str in RATING_LABELS:
                                        row_data[obj] = cell_str
                                    else:
                                        row_data[obj] = RATING_LABELS[0]
                                else:
                                    row_data[obj] = RATING_LABELS[0]
                            else:
                                row_data[obj] = RATING_LABELS[0]

                        uploaded_rows.append(row_data)

                    if not uploaded_rows:
                        skipped_dims.append(dim)
                        continue

                    # Apply to session state — update existing + append new rows
                    if dim in st.session_state.mat_responses:
                        df_dim = st.session_state.mat_responses[dim].copy()
                        existing_qids = set(df_dim["Question ID"].astype(str).str.strip().tolist())
                        cells_updated = 0
                        new_rows_list = []

                        for urow in uploaded_rows:
                            qid = urow["Question ID"]
                            if qid in existing_qids:
                                # Update existing row
                                mask = df_dim["Question ID"].astype(str).str.strip() == qid
                                for obj in objects_:
                                    if obj in urow and obj in df_dim.columns:
                                        df_dim.loc[mask, obj] = urow[obj]
                                        cells_updated += 1
                                # Also update Section, Question, Weight if changed
                                if "Section" in urow:
                                    df_dim.loc[mask, "Section"] = urow["Section"]
                                if "Question" in urow:
                                    df_dim.loc[mask, "Question"] = urow["Question"]
                                if "Weight" in urow:
                                    df_dim.loc[mask, "Weight"] = urow["Weight"]
                            else:
                                # New row added by user in template
                                new_row = {
                                    "Question ID": qid,
                                    "Section": urow.get("Section", "Custom"),
                                    "Question": urow.get("Question", ""),
                                    "Weight": urow.get("Weight", 1.0),
                                }
                                for obj in objects_:
                                    new_row[obj] = urow.get(obj, RATING_LABELS[0])
                                new_rows_list.append(new_row)
                                cells_updated += len(objects_)

                        # Append new rows
                        if new_rows_list:
                            df_new = pd.DataFrame(new_rows_list)
                            df_dim = pd.concat([df_dim, df_new], ignore_index=True)
                            total_new_rows += len(new_rows_list)

                        st.session_state.mat_responses[dim] = df_dim
                        total_cells_imported += cells_updated
                        imported_dims.append(dim)

                # Show results
                if imported_dims:
                    msg = f"✅ Successfully imported **{total_cells_imported:,}** rating(s) across **{len(imported_dims)}** dimension(s): {', '.join(imported_dims)}"
                    if total_new_rows:
                        msg += f" — including **{total_new_rows}** new question(s) added from template."
                    st.success(msg)
                if skipped_dims:
                    st.warning(
                        f"⚠️ Skipped {len(skipped_dims)} dimension(s) "
                        f"(sheet not found or missing headers): {', '.join(skipped_dims)}"
                    )
                if imported_dims:
                    st.markdown(
                        '<div style="background:#f0fdf4;border:1px solid #bbf7d0;border-radius:8px;'
                        'padding:0.5rem 1rem;font-size:0.84rem;color:#166534;margin-top:0.5rem;">'
                        '💡 Switch to the <strong>Manual Entry</strong> tab to review and fine-tune '
                        'the imported responses before submitting.</div>',
                        unsafe_allow_html=True,
                    )
                    if st.button("🔄 Refresh Grid", key="mat_refresh_after_upload", use_container_width=False, type="primary"):
                        st.rerun()

            except Exception as e:
                st.error(f"❌ Failed to import template: {e}")
                with st.expander("Technical Details"):
                    st.code(traceback.format_exc())

    st.divider()

    # ── Submit button — full width, always visible ──
    st.markdown("""
    <div style="background:linear-gradient(135deg,#f5f0fc,#fdf2f8);border:1.5px solid #d9cef0;
         border-radius:14px;padding:1rem 1.4rem;margin-bottom:1rem;">
        <div style="font-size:0.84rem;color:#5b2d90;font-weight:500;">
            ✅ Once you have rated all dimensions above, click <strong>Submit</strong> to generate 
            your maturity report with scores, charts, and downloadable PDF &amp; Excel outputs.
        </div>
    </div>""", unsafe_allow_html=True)

    if st.button("🚀 Submit & Generate Report", type="primary", use_container_width=True, key="mat_submit"):
        _do_submit()
# ══════════════════════════════════════════════════════════════════════════════
#  PAGE: POLICY HUB (Enterprise Version)
# ══════════════════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════════════════
#  PAGE: POLICY HUB (Enterprise Version)
# ══════════════════════════════════════════════════════════════════════════════
def page_policy_hub():

    # Top Banner (aligned with your UI pattern)
    _page_banner(
        icon="📋",
        badge_text="Enterprise Governance",
        title="Policy Hub & Procedures Management",
        subtitle="Centralized repository for enterprise data governance policies, procedures, and standards.",
        gradient="135deg,#3d1d63 0%,#5b2d90 55%,#b10f74 100%"
    )

    _page_nav("policy")

    UIComponents.render_action_hint_bar(
        title="Browse Modules",
        message="Expand each section below to explore the Policy Hub feature set.",
        color="#c084fc",
    )

    st.markdown("""
    <div class="ph-section-intro">
        <p>
        The <strong>Policy Hub</strong> is a centralized platform that helps organizations manage
        policies, procedures, and approvals in one place. Users can upload documents, track workflows,
        receive notifications, and ensure compliance across departments.
        </p>
    </div>
    """, unsafe_allow_html=True)

    # ─────────────────────────────────────────────────────────────
    # CUSTOM NUMBERED ACCORDIONS — Gradient Banner + PPTX Content
    # ─────────────────────────────────────────────────────────────
    st.markdown("""
    <style>
    .ph-accordion { margin-bottom: 1rem; border-radius: 14px; overflow: hidden; box-shadow: 0 2px 8px rgba(91,45,144,0.10); }
    .ph-accordion summary {
        display: flex; align-items: center; gap: 16px; padding: 18px 24px;
        cursor: pointer; list-style: none; font-size: 1.08rem; font-weight: 700;
        color: #fff; transition: opacity 0.2s; border: none;
    }
    .ph-accordion summary:hover { opacity: 0.92; }
    .ph-accordion summary::-webkit-details-marker { display: none; }
    .ph-accordion summary::after {
        content: '▸'; margin-left: auto; font-size: 1.3rem; color: #fff;
        transition: transform 0.25s;
    }
    .ph-accordion[open] summary::after { transform: rotate(90deg); }
    .ph-accordion .ph-num-badge {
        display: inline-flex; align-items: center; justify-content: center;
        width: 36px; height: 36px; border-radius: 10px; font-size: 1.1rem;
        font-weight: 800; flex-shrink: 0; background: rgba(255,255,255,0.95);
        color: #3d1d63; border: 2px solid rgba(255,255,255,1);
        box-shadow: 0 2px 6px rgba(0,0,0,0.15);
    }
    .ph-accordion .ph-acc-body { padding: 8px 24px 24px 76px; background: #fff; }
    .ph-accordion .ph-acc-feat {
        padding: 16px 20px; margin-bottom: 10px; border-radius: 10px;
        background: #f9f8fc; border: 1px solid #ede8f7;
        box-shadow: 0 1px 4px rgba(91,45,144,0.06);
        transition: box-shadow 0.2s, border-color 0.2s;
    }
    .ph-accordion .ph-acc-feat:hover {
        box-shadow: 0 3px 10px rgba(91,45,144,0.12);
        border-color: #d9cef0;
    }
    .ph-accordion .ph-acc-feat-title { font-weight: 700; color: #3d1d63; font-size: 0.96rem; margin-bottom: 5px; }
    .ph-accordion .ph-acc-feat-desc { color: #555; font-size: 0.88rem; line-height: 1.55; }
    .ph-accordion .ph-acc-benefit {
        margin-top: 18px; padding: 16px 20px; border-radius: 12px;
        background: linear-gradient(135deg, #f5f0fc 0%, #ede8f7 100%);
        border-left: 4px solid #5b2d90;
        box-shadow: 0 2px 6px rgba(91,45,144,0.08);
    }
    .ph-accordion .ph-acc-benefit-label { font-weight: 800; color: #5b2d90; font-size: 0.8rem; text-transform: uppercase; letter-spacing: 0.6px; }
    .ph-accordion .ph-acc-benefit-text { color: #3d1d63; font-size: 0.9rem; margin-top: 6px; font-weight: 500; line-height: 1.55; }
    .ph-accordion .ph-sub-section {
        margin-top: 20px; margin-bottom: 12px; font-weight: 800; color: #fff; font-size: 0.82rem;
        text-transform: uppercase; letter-spacing: 0.6px; padding: 8px 16px; border-radius: 8px;
        background: linear-gradient(135deg, #5b2d90 0%, #7c4dbb 100%);
        display: inline-block;
    }
    </style>

    <!-- 1. Workflow Automation -->
    <details class="ph-accordion">
        <summary style="background:linear-gradient(135deg,#3d1d63 0%,#5b2d90 55%,#b10f74 100%);">
            <span class="ph-num-badge">1</span>
            Workflow Automation
        </summary>
        <div class="ph-acc-body">
            <div class="ph-sub-section">Approval Workflow</div>
            <div class="ph-acc-feat">
                <div class="ph-acc-feat-title">Configurable Policy Workflows</div>
                <div class="ph-acc-feat-desc">Workflows based on the Delegation of Authority (DOA) matrix, with different routes per department or policy type.</div>
            </div>
            <div class="ph-acc-feat">
                <div class="ph-acc-feat-title">Sequential &amp; Parallel Approvals</div>
                <div class="ph-acc-feat-desc">Support for sequential and parallel approvals with due dates at each step.</div>
            </div>
            <div class="ph-acc-feat">
                <div class="ph-acc-feat-title">Email-Based Approvals</div>
                <div class="ph-acc-feat-desc">Approvers receive branded emails with summary and can approve/reject via secure links without complex navigation.</div>
            </div>
            <div class="ph-acc-feat">
                <div class="ph-acc-feat-title">Mandatory Comments &amp; Audit Trail</div>
                <div class="ph-acc-feat-desc">Mandatory comments on rejection or major change requests, captured in the audit trail alongside decisions and timestamps.</div>
            </div>
            <div class="ph-acc-feat">
                <div class="ph-acc-feat-title">Escalation Rules &amp; Dashboards</div>
                <div class="ph-acc-feat-desc">Escalation rules and reminders for overdue approvals, plus dashboards showing bottlenecks, average cycle time and pending items.</div>
            </div>
            <div class="ph-sub-section">Task Management</div>
            <div class="ph-acc-feat">
                <div class="ph-acc-feat-title">Integrated Task List</div>
                <div class="ph-acc-feat-desc">Policy-related work: drafting, reviews, approvals, periodic reviews, attestation follow-ups and remediation actions.</div>
            </div>
            <div class="ph-acc-feat">
                <div class="ph-acc-feat-title">Task Assignment &amp; Tracking</div>
                <div class="ph-acc-feat-desc">Assign tasks to individuals or groups with due dates, priority, status (not started / in progress / on hold / completed) and links to the relevant policy.</div>
            </div>
            <div class="ph-acc-feat">
                <div class="ph-acc-feat-title">Automation Rules &amp; Templates</div>
                <div class="ph-acc-feat-desc">Auto-create tasks on specific events (e.g., "policy due for annual review", "policy rejected — rework required").</div>
            </div>
            <div class="ph-acc-feat">
                <div class="ph-acc-feat-title">Task SLA Reporting</div>
                <div class="ph-acc-feat-desc">Reporting on task SLAs, completion trends and open items by department to support governance committees.</div>
            </div>
            <div class="ph-acc-benefit">
                <div class="ph-acc-benefit-label">Key Deliverables</div>
                <div class="ph-acc-benefit-text">Faster approvals and renewals with DOA-aligned workflows. Reduced regulatory risk through mandatory comments, escalations, and audit trails. Stronger compliance execution with automated tasks ensuring timely reviews and attestations.</div>
            </div>
        </div>
    </details>

    <!-- 2. Notifications & Reminders -->
    <details class="ph-accordion">
        <summary style="background:linear-gradient(135deg,#3d1d63 0%,#5b2d90 55%,#b10f74 100%);">
            <span class="ph-num-badge">2</span>
            Notifications &amp; Reminders
        </summary>
        <div class="ph-acc-body">
            <div class="ph-sub-section">Event-Based Notifications</div>
            <div class="ph-acc-feat">
                <div class="ph-acc-feat-title">Policy Lifecycle Alerts</div>
                <div class="ph-acc-feat-desc">Event-based notifications for created, updated, deleted, and published policies, new comments, @mentions, workflow transitions and overdue tasks.</div>
            </div>
            <div class="ph-acc-feat">
                <div class="ph-acc-feat-title">Configurable Frequency &amp; Channels</div>
                <div class="ph-acc-feat-desc">Choose between immediate, daily digest, or weekly summary notifications delivered via email and in-app alerts.</div>
            </div>
            <div class="ph-acc-feat">
                <div class="ph-acc-feat-title">Reminders &amp; Escalation</div>
                <div class="ph-acc-feat-desc">Reminders and escalation for overdue tasks; dashboard views for managers to see workload, bottlenecks and completion rates.</div>
            </div>
            <div class="ph-sub-section">Audit Trails</div>
            <div class="ph-acc-feat">
                <div class="ph-acc-feat-title">Full Audit Trail</div>
                <div class="ph-acc-feat-desc">Captures authentication events, content changes, approvals, permission changes, version restores and deletions with user, time and IP/device.</div>
            </div>
            <div class="ph-acc-feat">
                <div class="ph-acc-feat-title">Tamper-Resistant Logging</div>
                <div class="ph-acc-feat-desc">Aligned to banking/audit expectations, with long-term retention and export for external auditors.</div>
            </div>
            <div class="ph-acc-feat">
                <div class="ph-acc-feat-title">Activity Dashboards &amp; Reports</div>
                <div class="ph-acc-feat-desc">Dashboards and reports on activity — who changed what, when; policy review history; access to sensitive policies.</div>
            </div>
            <div class="ph-acc-benefit">
                <div class="ph-acc-benefit-label">Key Deliverables</div>
                <div class="ph-acc-benefit-text">Audit-ready governance with full traceability. Improved compliance oversight with dashboards showing policy history and access patterns. Faster issue resolution through event-based notifications.</div>
            </div>
        </div>
    </details>

    <!-- 3. Role-Based User Access -->
    <details class="ph-accordion">
        <summary style="background:linear-gradient(135deg,#3d1d63 0%,#5b2d90 55%,#b10f74 100%);">
            <span class="ph-num-badge">3</span>
            Role-Based User Access
        </summary>
        <div class="ph-acc-body">
            <div class="ph-sub-section">Access Management</div>
            <div class="ph-acc-feat">
                <div class="ph-acc-feat-title">Single Sign-On (SSO)</div>
                <div class="ph-acc-feat-desc">SSO via Microsoft 365 / Azure AD, using existing corporate credentials and security policies (including optional MFA).</div>
            </div>
            <div class="ph-acc-feat">
                <div class="ph-acc-feat-title">Role-Based Access Model</div>
                <div class="ph-acc-feat-desc">Admin, Editor, Viewer, plus optional custom roles controlling create/edit/publish/archive permissions.</div>
            </div>
            <div class="ph-acc-feat">
                <div class="ph-acc-feat-title">Department &amp; Policy-Wise Access Control</div>
                <div class="ph-acc-feat-desc">Each policy mapped to departments and policy groups; users only see content they are authorized for.</div>
            </div>
            <div class="ph-acc-feat">
                <div class="ph-acc-feat-title">Document &amp; Section-Level Permissions</div>
                <div class="ph-acc-feat-desc">Restrict certain annexures or clauses to small groups for confidential content.</div>
            </div>
            <div class="ph-acc-feat">
                <div class="ph-acc-feat-title">Comprehensive Access Logs</div>
                <div class="ph-acc-feat-desc">Real-time monitoring of suspicious activity and exportable reports for internal audit and regulators.</div>
            </div>
            <div class="ph-sub-section">Security</div>
            <div class="ph-acc-feat">
                <div class="ph-acc-feat-title">Microsoft 365 Security Stack</div>
                <div class="ph-acc-feat-desc">Encryption in transit (TLS) and at rest, secure tenant configuration and hardening.</div>
            </div>
            <div class="ph-acc-feat">
                <div class="ph-acc-feat-title">Cybersecurity Policy Alignment</div>
                <div class="ph-acc-feat-desc">RBAC, least-privilege, change control, periodic access reviews, segregation of duties aligned with central bank guidelines.</div>
            </div>
            <div class="ph-acc-feat">
                <div class="ph-acc-feat-title">Sensitivity Labels &amp; DLP</div>
                <div class="ph-acc-feat-desc">Optional field-level protection for highly sensitive information using SharePoint sensitivity labels and Data Loss Prevention.</div>
            </div>
            <div class="ph-acc-feat">
                <div class="ph-acc-feat-title">Backup &amp; Recovery</div>
                <div class="ph-acc-feat-desc">Business-continuity features consistent with Microsoft 365 SLAs, plus bank-defined retention for content.</div>
            </div>
            <div class="ph-acc-benefit">
                <div class="ph-acc-benefit-label">Key Deliverables</div>
                <div class="ph-acc-benefit-text">Stronger data protection with SSO, MFA, and role-based access. Audit-ready access controls with comprehensive logs. Reduced cyber and regulatory risk with enterprise-grade encryption and DLP.</div>
            </div>
        </div>
    </details>

    <!-- 4. White-Labelling -->
    <details class="ph-accordion">
        <summary style="background:linear-gradient(135deg,#3d1d63 0%,#5b2d90 55%,#b10f74 100%);">
            <span class="ph-num-badge">4</span>
            White-Labelling
        </summary>
        <div class="ph-acc-body">
            <div class="ph-sub-section">Personalization &amp; Branding</div>
            <div class="ph-acc-feat">
                <div class="ph-acc-feat-title">White-Labelled Portal</div>
                <div class="ph-acc-feat-desc">Portal aligned with the bank's branding — logo, colours, fonts, layouts, favicon and email templates.</div>
            </div>
            <div class="ph-acc-feat">
                <div class="ph-acc-feat-title">Role-Based Home Pages</div>
                <div class="ph-acc-feat-desc">Personalized dashboards showing relevant policies, tasks, approvals and frequently used links.</div>
            </div>
            <div class="ph-acc-feat">
                <div class="ph-acc-feat-title">Configurable Widgets</div>
                <div class="ph-acc-feat-desc">"Recently viewed", "Policies due for review", "My drafts" — users can add, remove and reorder widgets.</div>
            </div>
            <div class="ph-acc-feat">
                <div class="ph-acc-feat-title">User Preferences</div>
                <div class="ph-acc-feat-desc">Language, time zone, theme (light/dark) and default landing view (by department, by policy type, etc.).</div>
            </div>
            <div class="ph-acc-feat">
                <div class="ph-acc-feat-title">Personal Collections</div>
                <div class="ph-acc-feat-desc">Favourites, pinned documents and saved filters so power-users can build their own workspace on top of the portal.</div>
            </div>
            <div class="ph-acc-feat">
                <div class="ph-acc-feat-title">Branded Email Templates</div>
                <div class="ph-acc-feat-desc">Approval and notification emails aligned with brand identity for a consistent experience.</div>
            </div>
            <div class="ph-acc-benefit">
                <div class="ph-acc-benefit-label">Key Deliverables</div>
                <div class="ph-acc-benefit-text">Higher user adoption with role-based personalized dashboards. Faster task completion with pending actions visible on the home page. Consistent brand experience across the entire portal.</div>
            </div>
        </div>
    </details>
    """, unsafe_allow_html=True)

def page_case_management():
    _page_banner(
        icon="📁",
        badge_text="Issue Resolution Centre",
        title="Case Management",
        subtitle="Track, manage and resolve data quality issues and governance cases with full audit trail, escalation workflows and SLA monitoring.",
        gradient="135deg,#3d1d63 0%,#5b2d90 55%,#b10f74 100%"
    )
    _page_nav("case")

    # Coming Soon message
    st.markdown("""
    <div style="
        background: linear-gradient(135deg, #f9f6ff 0%, #fdf2f8 100%);
        border: 2px dashed #c4b0e0;
        border-radius: 20px;
        padding: 3.5rem 2.5rem;
        text-align: center;
        margin: 2rem auto;
        max-width: 680px;
    ">
        <div style="font-size: 4rem; margin-bottom: 1rem; filter: drop-shadow(0 4px 12px rgba(91,45,144,0.2));">🚧</div>
        <h2 style="margin: 0 0 0.6rem; font-size: 1.8rem; font-weight: 800;
            background: linear-gradient(135deg, #5b2d90, #b10f74);
            -webkit-background-clip: text; -webkit-text-fill-color: transparent;
            background-clip: text;">
            Coming Soon
        </h2>
        <p style="font-size: 1.05rem; color: #5b2d90; font-weight: 600; margin-bottom: 0.5rem;">
            Case Management is currently under development.
        </p>
        <p style="font-size: 0.92rem; color: #7a7a9a; line-height: 1.7; max-width: 480px; margin: 0 auto 2rem;">
            This module will allow you to raise, track, escalate and resolve data quality issues
            and governance cases — with SLA monitoring, role-based assignments and a full audit trail.
            Stay tuned for the upcoming release!
        </p>
        <div style="display: flex; justify-content: center; gap: 1rem; flex-wrap: wrap;">
            <div style="background: #fff; border: 1.5px solid #d9cef0; border-radius: 12px;
                padding: 0.75rem 1.25rem; min-width: 130px; box-shadow: 0 2px 10px rgba(91,45,144,0.08);">
                <div style="font-size: 1.3rem;">🎫</div>
                <div style="font-size: 0.78rem; font-weight: 700; color: #3b1f72; margin-top: 0.3rem;">Case Ticketing</div>
            </div>
            <div style="background: #fff; border: 1.5px solid #d9cef0; border-radius: 12px;
                padding: 0.75rem 1.25rem; min-width: 130px; box-shadow: 0 2px 10px rgba(91,45,144,0.08);">
                <div style="font-size: 1.3rem;">⏱️</div>
                <div style="font-size: 0.78rem; font-weight: 700; color: #3b1f72; margin-top: 0.3rem;">SLA Monitoring</div>
            </div>
            <div style="background: #fff; border: 1.5px solid #d9cef0; border-radius: 12px;
                padding: 0.75rem 1.25rem; min-width: 130px; box-shadow: 0 2px 10px rgba(91,45,144,0.08);">
                <div style="font-size: 1.3rem;">📊</div>
                <div style="font-size: 0.78rem; font-weight: 700; color: #3b1f72; margin-top: 0.3rem;">Audit Trail</div>
            </div>
            <div style="background: #fff; border: 1.5px solid #d9cef0; border-radius: 12px;
                padding: 0.75rem 1.25rem; min-width: 130px; box-shadow: 0 2px 10px rgba(91,45,144,0.08);">
                <div style="font-size: 1.3rem;">🔔</div>
                <div style="font-size: 0.78rem; font-weight: 700; color: #3b1f72; margin-top: 0.3rem;">Escalation Alerts</div>
            </div>
        </div>
    </div>

    """, unsafe_allow_html=True)

    _,btn_col,_ = st.columns([2,1.5,2])
    with btn_col:
        if st.button("← Back to Home", use_container_width=True, key="case_back_home"):
            st.session_state["page"] = "home"; st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════
try:
    AppConfig.TEMP_DIR.mkdir(parents=True, exist_ok=True)
    AppConfig.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
except Exception:
    pass

load_css()
_init_state()

{
    "home":     page_home,
    "dq":       page_dq,
    "maturity": page_maturity,
    "policy":   page_policy_hub,
    "case":     page_case_management,
}.get(st.session_state.get("page", "home"), page_home)()